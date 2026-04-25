import csv
import io
import mimetypes
import os
import re
import sqlite3
from datetime import timedelta
import sys
from pathlib import Path
from jinja2 import ChoiceLoader, FileSystemLoader

from flask import (
    Flask,
    Response,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
import oracledb
from werkzeug.utils import secure_filename

# 古いバージョンの Oracle DB に接続するため、
# Python-oracledb を thick モード（Oracle Client 経由）で使用する。
try:
    # Windows では、Oracle Client / Instant Client が PATH やレジストリに
    # 登録されていれば、引数なしで自動検出されます。
    oracledb.init_oracle_client()
except oracledb.ProgrammingError:
    # すでに初期化済みの場合などは無視
    pass

PLATFORM_ROOT = Path(__file__).resolve().parents[2]
COMMON_DIR = PLATFORM_ROOT / "common"

sys.path.insert(0, str(PLATFORM_ROOT))
from core.csv_loader import rows_to_csv
from core.sap_integrated_master import (
    append_supplier_use_lines_after_inputs,
    fetch_integrated_master_rows_from_sqlserver,
)
from core.sap_inventory_table import fetch_inventory_rows_from_sqlserver
from core.sap_item_table import fetch_item_table_rows_from_sqlserver
from core.sap_order_table import fetch_order_rows_from_sqlserver
from core.peb_excel_exports import (
    load_peb_inventory_rows_from_xlsx_bytes,
    load_peb_inventory_wip_rows_from_xlsx_bytes,
    load_peb_monthly_result_rows_from_xlsx_bytes,
    load_peb_order_rows_from_xlsx_bytes,
    load_peb_prd_plan_rows_from_xlsx_bytes,
)
from core.sqlserver_conn import connect_sqlserver
from config.bridge_customers import BRIDGE_CUSTOMERS
from config.settings import DB_PATH

app = Flask(__name__, static_folder=str(COMMON_DIR / "static"))
app.jinja_loader = ChoiceLoader(
    [FileSystemLoader(str(COMMON_DIR / "templates")), app.jinja_loader]
)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "asprova-bridge-dev-secret")
# ブラウザを閉じても接続フォーム用セッションを残す（デフォルトのセッション Cookie は終了時に消える）
_bridge_session_days = int(os.environ.get("BRIDGE_SESSION_DAYS", "30"))
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=max(1, min(_bridge_session_days, 365)))

MCFRAME_LOGO_PATH = os.environ.get(
    "MCFRAME_LOGO_PATH",
    str(COMMON_DIR / "static" / "mcframe-logo.webp"),
)

CONFIRM_LOGO_PATH = os.environ.get(
    "CONFIRM_LOGO_PATH",
    str(COMMON_DIR / "static" / "mcframe-logo.webp"),
)


HTML_INDEX = None


MASTER_SQL = """
SELECT
    b.P_ITM_CD AS P_ITM_CD,
    10 AS PROCESS_NO,
    '10' AS PROCESS_CD,
    'I' AS INST_TYP,
    'In' || ROW_NUMBER() OVER (
      PARTITION BY b.P_ITM_CD
      ORDER BY b.C_ITM_CD
    ) AS INST_CD,
    b.C_ITM_CD AS ITM_RESOURCE,
    TO_CHAR(b.C_REQ_QTY) AS PRODUCTION
FROM
    {schema}.SM_BOM_ALL b
WHERE
    b.BOM_PTN = 1

UNION ALL

SELECT
    hl.ITM_CD AS P_ITM_CD,
    10 AS PROCESS_NO,
    '10' AS PROCESS_CD,
    'U' AS INST_TYP,
    'M' AS INST_CD,
    hl.LINE_CD AS ITM_RESOURCE,
    TO_CHAR(hl.CYCLE_TIME) || 'mp' AS PRODUCTION
FROM
    {schema}.SM_HINLINE_ALL hl
WHERE
    hl.CO_CD = '{co_cd}'
"""


INTEGRATED_HEADERS = [
    "P_ITM_CD",
    "PROCESS_NO",
    "PROCESS_CD",
    "INST_TYP",
    "INST_CD",
    "ITM_RESOURCE",
    "PRODUCTION",
]


ITEM_TABLE_SQL = """
SELECT
    c.ITM_CD,
    c.ITM_NM,
    CASE TRIM(TO_CHAR(c.ITM_TYP))
        WHEN '1' THEN 'P'
        WHEN '2' THEN 'I'
        WHEN '5' THEN 'M'
        WHEN '6' THEN 'H'
        WHEN '7' THEN 'U'
        ELSE TRIM(TO_CHAR(c.ITM_TYP))
    END AS ITM_TYP,
    MAX(m.MAX_LOT_UNIT_QTY) AS MAX_LOT_UNIT_QTY
FROM
    {schema}.CM_HINMO_ALL c
    JOIN {schema}.SM_HINMOS_ALL m
      ON m.ITM_CD = c.ITM_CD
WHERE
    c.CO_CD = '{co_cd}'
    AND m.CO_CD = '{co_cd}'
    -- AND m.BOM_PTN = 1
GROUP BY
    c.ITM_CD,
    c.ITM_NM,
    c.ITM_TYP
ORDER BY
    c.ITM_CD
"""

ITEM_TABLE_HEADERS = [
    "ITM_CD",
    "ITM_NM",
    "ITM_TYP",
    "MAX_LOT_UNIT_QTY",
]

ORDER_TABLE_SQL = """
SELECT
    REQ_NO,
    ITM_CD,
    DLV_DT,
    REQ_QTY,
    CAST(NULL AS VARCHAR2(64)) AS CUST_CD
FROM
    {schema}.ST_SHOYO_ALL
ORDER BY
    REQ_NO,
    ITM_CD,
    DLV_DT
"""

ORDER_TABLE_HEADERS = [
    "REQ_NO",
    "ITM_CD",
    "DLV_DT",
    "REQ_QTY",
    "CUST_CD",
]

RESOURCE_TABLE_SQL = """
SELECT
    LINE_CD,
    LINE_NM,
    CASE
        WHEN LINE_CD LIKE 'M%' THEN 'INJECTION'
        ELSE ''
    END AS RESOURCE_GRP
FROM
    {schema}.SM_LINE_ALL
ORDER BY
    LINE_CD
"""

RESOURCE_TABLE_HEADERS = [
    "LINE_CD",
    "LINE_NM",
    "RESOURCE_GRP",
    "Sort_Order",
]

INVENTORY_TABLE_SQL = """
SELECT
    'INV' || LPAD(ROW_NUMBER() OVER (ORDER BY ITM_CD), 5, '0') AS INV_CD,
    ITM_CD,
    STK_QTY,
    CAST(NULL AS VARCHAR2(10)) AS INV_DT
FROM
    {schema}.ST_GNZAIKO_ALL
ORDER BY
    ITM_CD
"""

INVENTORY_TABLE_HEADERS = [
    "INV_CD",
    "ITM_CD",
    "STK_QTY",
    "INV_DT",
]

RESOURCE_LINE_CODE_CANDIDATES = ("LINE_CD", "Line Code", "LINE CODE")
RESOURCE_LINE_NAME_CANDIDATES = ("LINE_NM", "Line Name", "LINE NAME")
RESOURCE_CYCLE_TIME_CANDIDATES = ("CYCLE_TIME", "Cycle Time", "CYCLE TIME")
RESOURCE_CYCLE_TIME_HEADER = "CYCLE_TIME"
RESOURCE_CYCLE_TABLE = "bridge_resource_cycle_times"
RESOURCE_CYCLE_SOURCE_SESSION_KEY = "resource_cycle_source_name"
RESOURCE_CYCLE_DB_PATH = Path(DB_PATH)
RESOURCE_CYCLE_LEGACY_CO_CD = (
    os.environ.get("BRIDGE_RESOURCE_CYCLE_LEGACY_CO_CD", "NCI").strip().upper() or "NCI"
)


def _pick_first_matching(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    for key in candidates:
        if key in headers:
            return key
    return None


def _open_resource_cycle_db() -> sqlite3.Connection:
    RESOURCE_CYCLE_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(RESOURCE_CYCLE_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _resource_cycle_scope_key() -> str:
    """Line Cycle Time のデータ分離キー。mcframe は co_cd、excel は customer_id。"""
    erp = get_erp_system()
    customer_id = str(session.get("customer_id") or "").strip().lower()
    if erp == "excel" and customer_id:
        return customer_id
    raw = (session.get("mcframe_co_cd") or "").strip().upper()
    if raw and _MCFRAME_CO_CD_RE.match(raw):
        return raw
    if customer_id:
        return customer_id
    return RESOURCE_CYCLE_LEGACY_CO_CD


def _ensure_resource_cycle_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {RESOURCE_CYCLE_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            co_cd TEXT NOT NULL DEFAULT '',
            line_code TEXT NOT NULL,
            line_name TEXT NOT NULL DEFAULT '',
            cycle_time TEXT NOT NULL DEFAULT '',
            source_filename TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur = conn.execute(f"PRAGMA table_info({RESOURCE_CYCLE_TABLE})")
    existing_cols = {str(row["name"]) for row in cur.fetchall()}
    if "co_cd" not in existing_cols:
        conn.execute(
            f"ALTER TABLE {RESOURCE_CYCLE_TABLE} ADD COLUMN co_cd TEXT NOT NULL DEFAULT ''"
        )
        conn.execute(
            f"""
            UPDATE {RESOURCE_CYCLE_TABLE}
            SET co_cd = ?
            WHERE TRIM(COALESCE(co_cd, '')) = ''
            """,
            (RESOURCE_CYCLE_LEGACY_CO_CD,),
        )


def _fetch_resource_cycle_rows() -> list[dict]:
    co_cd = _resource_cycle_scope_key()
    with _open_resource_cycle_db() as conn:
        _ensure_resource_cycle_table(conn)
        cur = conn.execute(
            f"""
            SELECT id, line_code, line_name, cycle_time, sort_order
            FROM {RESOURCE_CYCLE_TABLE}
            WHERE co_cd = ?
            ORDER BY sort_order ASC, id ASC
            """,
            (co_cd,),
        )
        return [
            {
                "id": int(r["id"]),
                "line_code": str(r["line_code"] or "").strip(),
                "line_name": str(r["line_name"] or "").strip(),
                "cycle_time": str(r["cycle_time"] or "").strip(),
                "sort_order": int(r["sort_order"] or 0),
            }
            for r in cur.fetchall()
        ]


def _replace_resource_cycle_rows(rows: list[dict], source_filename: str, co_cd: str) -> None:
    with _open_resource_cycle_db() as conn:
        _ensure_resource_cycle_table(conn)
        conn.execute(f"DELETE FROM {RESOURCE_CYCLE_TABLE} WHERE co_cd = ?", (co_cd,))
        conn.executemany(
            f"""
            INSERT INTO {RESOURCE_CYCLE_TABLE}
            (co_cd, line_code, line_name, cycle_time, source_filename, sort_order, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            [
                (
                    co_cd,
                    str(row.get("line_code") or "").strip(),
                    str(row.get("line_name") or "").strip(),
                    str(row.get("cycle_time") or "").strip(),
                    source_filename,
                    idx,
                )
                for idx, row in enumerate(rows)
            ],
        )
        conn.commit()


def _update_resource_cycle_rows_by_id(updates: dict[int, dict[str, str | int]]) -> None:
    if not updates:
        return
    co_cd = _resource_cycle_scope_key()
    with _open_resource_cycle_db() as conn:
        _ensure_resource_cycle_table(conn)
        for row_id, payload in updates.items():
            cycle_time = str(payload.get("cycle_time") or "").strip()
            sort_order = int(payload.get("sort_order") or 0)
            conn.execute(
                f"""
                UPDATE {RESOURCE_CYCLE_TABLE}
                SET cycle_time = ?, sort_order = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND co_cd = ?
                """,
                (cycle_time, sort_order, row_id, co_cd),
            )
        conn.commit()


def _delete_resource_cycle_row_by_id(row_id: int) -> int:
    co_cd = _resource_cycle_scope_key()
    with _open_resource_cycle_db() as conn:
        _ensure_resource_cycle_table(conn)
        cur = conn.execute(
            f"DELETE FROM {RESOURCE_CYCLE_TABLE} WHERE id = ? AND co_cd = ?",
            (row_id, co_cd),
        )
        conn.commit()
        return int(cur.rowcount or 0)


def _load_cycle_time_map_from_sqlite() -> dict[str, str]:
    co_cd = _resource_cycle_scope_key()
    try:
        with _open_resource_cycle_db() as conn:
            _ensure_resource_cycle_table(conn)
            cur = conn.execute(
                f"""
                SELECT line_code, cycle_time
                FROM {RESOURCE_CYCLE_TABLE}
                WHERE co_cd = ?
                ORDER BY sort_order ASC, id ASC
                """,
                (co_cd,),
            )
            out: dict[str, str] = {}
            for row in cur.fetchall():
                line_code = str(row["line_code"] or "").strip()
                cycle_time = str(row["cycle_time"] or "").strip()
                if not line_code or not cycle_time:
                    continue
                out[line_code] = cycle_time
            return out
    except Exception:
        return {}


def _load_sort_order_map_from_sqlite() -> dict[str, int]:
    co_cd = _resource_cycle_scope_key()
    try:
        with _open_resource_cycle_db() as conn:
            _ensure_resource_cycle_table(conn)
            cur = conn.execute(
                f"""
                SELECT line_code, sort_order
                FROM {RESOURCE_CYCLE_TABLE}
                WHERE co_cd = ?
                ORDER BY sort_order ASC, id ASC
                """,
                (co_cd,),
            )
            out: dict[str, int] = {}
            for row in cur.fetchall():
                line_code = str(row["line_code"] or "").strip()
                if not line_code:
                    continue
                out[line_code] = int(row["sort_order"] or 0)
            return out
    except Exception:
        return {}


def _apply_cycle_time_to_integrated_records(records: list[dict]) -> list[dict]:
    cycle_map = _load_cycle_time_map_from_sqlite()
    if not cycle_map:
        return records
    for rec in records:
        itm_resource = str(rec.get("ITM_RESOURCE") or "").strip()
        if itm_resource and itm_resource in cycle_map:
            rec["PRODUCTION"] = cycle_map[itm_resource]
    return records


def _read_resource_cycle_rows_from_bytes(raw: bytes) -> list[dict]:
    errors: list[Exception] = []
    for enc in ("utf-8-sig", "cp932", "shift_jis"):
        try:
            text = raw.decode(enc)
            reader = csv.DictReader(io.StringIO(text))
            headers = list(reader.fieldnames or [])
            rows = [dict(r) for r in reader]
            line_key = _pick_first_matching(headers, RESOURCE_LINE_CODE_CANDIDATES)
            if not line_key:
                raise RuntimeError("CSV に LINE_CD（または Line Code）列がありません。")
            name_key = _pick_first_matching(headers, RESOURCE_LINE_NAME_CANDIDATES) or "LINE_NM"
            cycle_key = _pick_first_matching(headers, RESOURCE_CYCLE_TIME_CANDIDATES)
            if not cycle_key:
                cycle_key = RESOURCE_CYCLE_TIME_HEADER
            out: list[dict] = []
            for row in rows:
                out.append(
                    {
                        "line_code": str(row.get(line_key) or "").strip(),
                        "line_name": str(row.get(name_key) or "").strip(),
                        "cycle_time": str(row.get(cycle_key) or "").strip(),
                    }
                )
            return out
        except UnicodeDecodeError as exc:
            errors.append(exc)
            continue
    raise RuntimeError("CSV の文字コードを判別できませんでした。") from (
        errors[-1] if errors else None
    )


def get_connection():
    if get_erp_system() == "sap_b1":
        raise RuntimeError("SAP B1 は SQL Server 接続です。get_sqlserver_connection() を使用してください。")
    user = session.get("oracle_user")
    password = session.get("oracle_password")
    if not (user and password):
        raise RuntimeError("未接続です。先に「Connect」からデータベースへ接続してください。")
    dsn = session.get("oracle_dsn") or "orcl"
    return oracledb.connect(user=user, password=password, dsn=dsn)


def get_sqlserver_connection():
    """SAP B1: DNS=サーバー名、SCHEMA=データベース名、ID/PASSWORD=SQL 認証。"""
    if get_erp_system() != "sap_b1":
        raise RuntimeError("内部エラー: SQL Server 接続は SAP Business One モードのみです。")
    user = session.get("oracle_user")
    password = session.get("oracle_password")
    server = (session.get("oracle_dsn") or "").strip()
    database = (session.get("oracle_schema") or "").strip()
    if not (user and password and server and database):
        raise RuntimeError("未接続です。先に「Connect」からデータベースへ接続してください。")
    return connect_sqlserver(server, database, user, password, timeout=30)


def get_schema() -> str:
    schema = session.get("oracle_schema")
    if not schema:
        raise RuntimeError("未接続です。先に「Connect」からデータベースへ接続してください。")
    return schema


_SCHEMA_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]{0,29}$")
# SAP B1 / SQL Server: database name (SCHEMA 欄に DB 名を入力)
_SAP_B1_DB_RE = re.compile(r"^[A-Za-z0-9_.\-]{1,128}$")
# mcframe テーブルの CO_CD（Connect 画面の Company CD）。英数字 1〜20、保存時は大文字。
_MCFRAME_CO_CD_RE = re.compile(r"^[A-Z0-9]{1,20}$")


def get_mcframe_co_cd() -> str:
    """接続時にセッションへ保存した mcframe 用会社コード。未設定時は J0001。"""
    raw = (session.get("mcframe_co_cd") or "").strip().upper()
    if raw and _MCFRAME_CO_CD_RE.match(raw):
        return raw
    return "J0001"


def get_erp_system() -> str:
    raw = (session.get("erp_system") or "mcframe").strip().lower()
    return raw if raw in ("mcframe", "sap_b1", "excel") else "mcframe"


def _customer_profile_options() -> list[dict[str, str]]:
    options: list[dict[str, str]] = []
    for key, cfg in BRIDGE_CUSTOMERS.items():
        options.append(
            {
                "id": key,
                "label": str(cfg.get("label") or key),
                "erp_system": str(cfg.get("erp_system") or "mcframe"),
                "oracle_id": str(cfg.get("oracle_id") or ""),
                "oracle_pwd": str(cfg.get("oracle_pwd") or ""),
                "oracle_schema": str(cfg.get("oracle_schema") or ""),
                "oracle_dsn": str(cfg.get("oracle_dsn") or ""),
                "mcframe_co_cd": str(cfg.get("mcframe_co_cd") or ""),
                "excel_base_dir": str(cfg.get("excel_base_dir") or ""),
                "excel_integrated_file": str(cfg.get("excel_integrated_file") or ""),
                "excel_item_file": str(cfg.get("excel_item_file") or ""),
                "excel_order_file": str(cfg.get("excel_order_file") or ""),
                "excel_prd_plan_file": str(cfg.get("excel_prd_plan_file") or ""),
                "excel_resource_file": str(cfg.get("excel_resource_file") or ""),
                "excel_inventory_file": str(cfg.get("excel_inventory_file") or ""),
                "excel_inventory_wip_file": str(cfg.get("excel_inventory_wip_file") or ""),
            }
        )
    return options


def _sap_b1_not_supported_flash() -> None:
    flash(
        "SAP Business One 接続時はこの出力は未対応です（mcframe / Oracle 用のテーブル参照です）。",
        "error",
    )


def fetch_rows(sql: str):
    if get_erp_system() in ("sap_b1", "excel"):
        raise RuntimeError("この接続種別では Oracle SQL は使用しません。")
    with get_connection() as conn:
        with conn.cursor() as cur:
            schema = get_schema()
            cur.execute(
                sql.format_map(
                    {"schema": schema, "co_cd": get_mcframe_co_cd()}
                )
            )
            for row in cur:
                yield row


def _get_selected_customer_profile() -> dict[str, str] | None:
    customer_id = str(session.get("customer_id") or "").strip()
    if not customer_id:
        return None
    profile = BRIDGE_CUSTOMERS.get(customer_id)
    return profile if isinstance(profile, dict) else None


def _excel_file_key(kind: str) -> str:
    return {
        "integrated": "excel_integrated_file",
        "item": "excel_item_file",
        "order": "excel_order_file",
        "prd_plan": "excel_prd_plan_file",
        "resource": "excel_resource_file",
        "inventory": "excel_inventory_file",
        "inventory_wip": "excel_inventory_wip_file",
    }[kind]


def _default_excel_filename(kind: str) -> str:
    return {
        "integrated": "integrated_master.xlsx",
        "item": "item_table.xlsx",
        "order": "order_table.xlsx",
        "prd_plan": "prd_plan_table.xlsx",
        "resource": "resource_table.xlsx",
        "inventory": "inventory_table.xlsx",
        "inventory_wip": "inventory_wip_table.xlsx",
    }[kind]


def _resolve_excel_source_path(kind: str) -> Path:
    profile = _get_selected_customer_profile()
    if not profile:
        raise RuntimeError("Excel 取り込みでは Customer を選択してください。")
    base_dir = str(profile.get("excel_base_dir") or "").strip()
    if not base_dir:
        raise RuntimeError("選択中 Customer の excel_base_dir が未設定です。")
    file_key = _excel_file_key(kind)
    file_name = str(profile.get(file_key) or _default_excel_filename(kind)).strip()
    if not file_name:
        raise RuntimeError(f"選択中 Customer の {file_key} が未設定です。")
    return Path(base_dir) / file_name


def _iter_dict_rows_from_csv_bytes(raw: bytes) -> list[dict[str, str]]:
    errors: list[Exception] = []
    for enc in ("utf-8-sig", "cp932", "shift_jis"):
        try:
            text = raw.decode(enc)
            reader = csv.DictReader(io.StringIO(text))
            return [dict(r) for r in reader]
        except UnicodeDecodeError as exc:
            errors.append(exc)
            continue
    raise RuntimeError("CSV の文字コードを判別できませんでした。") from (
        errors[-1] if errors else None
    )


def _iter_dict_rows_from_xlsx_bytes(raw: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "Excel(.xlsx) の読み込みには openpyxl が必要です。"
        ) from exc
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        out: list[dict[str, str]] = []
        for values in rows:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = values[i] if i < len(values) else ""
                row[h] = "" if v is None else str(v).strip()
            out.append(row)
        return out
    finally:
        wb.close()


def _load_excel_export_rows(kind: str, headers: list[str], upload=None) -> list[tuple]:
    file_name = ""
    raw = b""
    if upload and getattr(upload, "filename", ""):
        file_name = str(upload.filename or "").strip()
        raw = upload.read()
        if not raw:
            raise RuntimeError("選択したソースファイルが空です。")
    else:
        path = _resolve_excel_source_path(kind)
        if not path.exists():
            raise RuntimeError(f"Excel/CSV ファイルが見つかりません: {path}")
        file_name = path.name
        raw = path.read_bytes()
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        dict_rows = _iter_dict_rows_from_csv_bytes(raw)
    elif suffix in (".xlsx", ".xlsm"):
        customer_id = str(session.get("customer_id") or "").strip().lower()
        if kind == "order" and customer_id == "peb":
            return load_peb_order_rows_from_xlsx_bytes(raw)
        if kind == "prd_plan" and customer_id == "peb":
            return load_peb_prd_plan_rows_from_xlsx_bytes(raw)
        if kind == "inventory" and customer_id == "peb":
            return load_peb_inventory_rows_from_xlsx_bytes(raw)
        if kind == "inventory_wip" and customer_id == "peb":
            return load_peb_inventory_wip_rows_from_xlsx_bytes(raw)
        dict_rows = _iter_dict_rows_from_xlsx_bytes(raw)
    else:
        raise RuntimeError(f"未対応の拡張子です: {suffix}（.csv/.xlsx/.xlsm のみ対応）")
    out: list[tuple] = []
    for row in dict_rows:
        out.append(tuple(str(row.get(h) or "").strip() for h in headers))
    return out


@app.route("/", methods=["GET"])
def index():
    show_line_cycle_master = (request.args.get("show_line_cycle") or "").strip() == "1"
    show_monthly_result = (request.args.get("show_monthly_result") or "").strip() == "1"
    line_cycle_rows: list[dict] = []
    if show_line_cycle_master:
        try:
            line_cycle_rows = _fetch_resource_cycle_rows()
        except Exception as exc:  # noqa: BLE001
            flash(f"サイクルタイムSQLiteの読み込みに失敗しました: {exc}", "error")
    return render_template(
        "bridge_index.html",
        oracle_connected=bool(session.get("oracle_connected")),
        erp_system=session.get("erp_system") or "mcframe",
        customer_profiles=_customer_profile_options(),
        selected_customer_id=str(session.get("customer_id") or ""),
        show_line_cycle_master=show_line_cycle_master,
        line_cycle_rows=line_cycle_rows,
        show_monthly_result=show_monthly_result,
        monthly_result_rows=[],
        monthly_result_source_name="",
    )


@app.route("/monthly-result", methods=["POST"])
def monthly_result():
    upload = request.files.get("monthly_result_excel")
    if not upload or not (upload.filename or "").strip():
        flash("Monthly Result のソース Excel ファイルを選択してください。", "error")
        return redirect(url_for("index", show_monthly_result="1"))
    try:
        raw = upload.read()
        if not raw:
            raise RuntimeError("アップロードされたファイルが空です。")
        rows = load_peb_monthly_result_rows_from_xlsx_bytes(raw)
        source_name = secure_filename(upload.filename or "monthly_result.xlsx")
    except Exception as exc:  # noqa: BLE001
        flash(f"Monthly Result の読み込みに失敗しました: {exc}", "error")
        return redirect(url_for("index", show_monthly_result="1"))

    show_line_cycle_master = (request.args.get("show_line_cycle") or "").strip() == "1"
    line_cycle_rows: list[dict] = []
    if show_line_cycle_master:
        try:
            line_cycle_rows = _fetch_resource_cycle_rows()
        except Exception as exc:  # noqa: BLE001
            flash(f"サイクルタイムSQLiteの読み込みに失敗しました: {exc}", "error")
    return render_template(
        "bridge_index.html",
        oracle_connected=bool(session.get("oracle_connected")),
        erp_system=session.get("erp_system") or "mcframe",
        customer_profiles=_customer_profile_options(),
        selected_customer_id=str(session.get("customer_id") or ""),
        show_line_cycle_master=show_line_cycle_master,
        line_cycle_rows=line_cycle_rows,
        show_monthly_result=True,
        monthly_result_rows=rows,
        monthly_result_source_name=source_name,
    )


def _send_logo_file(path: str) -> Response:
    mime, _ = mimetypes.guess_type(path)
    return send_file(path, mimetype=mime or "image/png")


@app.route("/assets/mcframe-logo.png", methods=["GET"])
def mcframe_logo():
    return _send_logo_file(MCFRAME_LOGO_PATH)


@app.route("/assets/confirm-logo.png", methods=["GET"])
def confirm_logo():
    return _send_logo_file(CONFIRM_LOGO_PATH)


@app.route("/connect", methods=["POST"])
def connect_oracle():
    customer_id = (request.form.get("customer_id") or "").strip()
    selected_profile = BRIDGE_CUSTOMERS.get(customer_id)
    oracle_id = request.form.get("oracle_id", "").strip()
    oracle_pwd = request.form.get("oracle_pwd", "")
    oracle_schema = request.form.get("oracle_schema", "").strip()
    oracle_dsn = request.form.get("oracle_dsn", "").strip()
    if selected_profile:
        if not oracle_id:
            oracle_id = str(selected_profile.get("oracle_id") or "").strip()
        if not oracle_pwd:
            oracle_pwd = str(selected_profile.get("oracle_pwd") or "")
        if not oracle_schema:
            oracle_schema = str(selected_profile.get("oracle_schema") or "").strip()
        if not oracle_dsn:
            oracle_dsn = str(selected_profile.get("oracle_dsn") or "").strip()
    erp_raw = (request.form.get("erp_system") or "mcframe").strip().lower()
    erp_system = erp_raw if erp_raw in ("mcframe", "sap_b1", "excel") else "mcframe"
    if selected_profile:
        profile_erp = str(selected_profile.get("erp_system") or "").strip().lower()
        if profile_erp in ("mcframe", "sap_b1", "excel") and profile_erp != erp_system:
            flash("選択したCustomerとSystemの組み合わせが一致しません。", "error")
            return redirect(url_for("index"))
    if erp_system != "excel" and (not oracle_id or not oracle_pwd or not oracle_schema or not oracle_dsn):
        flash("ID / PASSWORD / SCHEMA(DB Name) / DNS(Server Name) を入力してください。", "error")
        return redirect(url_for("index"))

    co_cd_stored: str | None = None
    if erp_system == "mcframe":
        co_raw = request.form.get("mcframe_co_cd", "").strip().upper()
        if not co_raw and selected_profile:
            co_raw = str(selected_profile.get("mcframe_co_cd") or "").strip().upper()
        if not co_raw:
            co_raw = (os.environ.get("BRIDGE_MCFRAME_CO_CD") or "J0001").strip().upper()
        if not _MCFRAME_CO_CD_RE.match(co_raw):
            flash(
                "Company CD は半角英数字のみ、1〜20文字で入力してください（未入力時は J0001 または"
                " 環境変数 BRIDGE_MCFRAME_CO_CD）。",
                "error",
            )
            return redirect(url_for("index"))
        co_cd_stored = co_raw

    if erp_system == "excel":
        schema_stored = ""
        oracle_dsn = ""
    elif erp_system == "sap_b1":
        if not _SAP_B1_DB_RE.match(oracle_schema):
            flash(
                "データベース名（SCHEMA 欄）は英数字・ドット・ハイフン・アンダースコアのみ、"
                "1〜128文字で入力してください。",
                "error",
            )
            return redirect(url_for("index"))
        schema_stored = oracle_schema
        try:
            conn = connect_sqlserver(oracle_dsn, schema_stored, oracle_id, oracle_pwd, timeout=15)
            conn.close()
        except Exception as exc:  # noqa: BLE001
            flash(f"接続に失敗しました: {exc}", "error")
            session["oracle_connected"] = False
            return redirect(url_for("index"))
    else:
        if not _SCHEMA_RE.match(oracle_schema):
            flash("Schema は英数字とアンダースコアのみ（先頭は英字、最大30文字）で入力してください。", "error")
            return redirect(url_for("index"))
        schema_stored = oracle_schema.upper()
        try:
            conn = oracledb.connect(user=oracle_id, password=oracle_pwd, dsn=oracle_dsn)
            conn.close()
        except Exception as exc:  # noqa: BLE001
            flash(f"接続に失敗しました: {exc}", "error")
            session["oracle_connected"] = False
            return redirect(url_for("index"))

    # 接続成功時のみ保持（モーダルは前回成功時の値をデフォルト表示）
    session.permanent = True
    session["erp_system"] = erp_system
    session["oracle_user"] = oracle_id if erp_system != "excel" else ""
    session["oracle_password"] = oracle_pwd if erp_system != "excel" else ""
    session["oracle_schema"] = schema_stored
    session["oracle_dsn"] = oracle_dsn
    session["oracle_connected"] = True
    session["customer_id"] = customer_id
    if erp_system == "mcframe" and co_cd_stored:
        session["mcframe_co_cd"] = co_cd_stored
    else:
        session.pop("mcframe_co_cd", None)
    flash("Connection successful.", "success")
    return redirect(url_for("index"))


@app.route("/download/integrated", methods=["POST"])
def download_integrated():
    try:
        if get_erp_system() == "excel":
            rows = _load_excel_export_rows("integrated", INTEGRATED_HEADERS, request.files.get("source_excel"))
            csv_data = rows_to_csv(rows, INTEGRATED_HEADERS)
        elif get_erp_system() == "sap_b1":
            conn = get_sqlserver_connection()
            try:
                recs = fetch_integrated_master_rows_from_sqlserver(conn)
                rec_dicts = [
                    {
                        "P_ITM_CD": r["P_ITM_CD"],
                        "PROCESS_NO": r["PROCESS_NO"],
                        "PROCESS_CD": r["PROCESS_CD"],
                        "INST_TYP": r["INST_TYP"],
                        "INST_CD": r["INST_CD"],
                        "ITM_RESOURCE": r["ITM_RESOURCE"],
                        "PRODUCTION": r["PRODUCTION"],
                    }
                    for r in recs
                ]
                rec_dicts = _apply_cycle_time_to_integrated_records(rec_dicts)
                rows = [tuple(d[h] for h in INTEGRATED_HEADERS) for d in rec_dicts]
                csv_data = rows_to_csv(rows, INTEGRATED_HEADERS)
            finally:
                conn.close()
        else:
            raw_rows = list(fetch_rows(MASTER_SQL))
            rec_dicts = [dict(zip(INTEGRATED_HEADERS, row)) for row in raw_rows]
            expanded = append_supplier_use_lines_after_inputs(rec_dicts)
            expanded = _apply_cycle_time_to_integrated_records(expanded)
            rows = [tuple(d[h] for h in INTEGRATED_HEADERS) for d in expanded]
            csv_data = rows_to_csv(rows, INTEGRATED_HEADERS)
    except Exception as exc:  # noqa: BLE001
        flash(f"エラーが発生しました: {exc}", "error")
        return redirect(url_for("index"))

    filename = "integrated_master.csv"
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/download/item-table", methods=["POST"])
def download_item_table():
    try:
        if get_erp_system() == "excel":
            rows = _load_excel_export_rows("item", ITEM_TABLE_HEADERS, request.files.get("source_excel"))
            csv_data = rows_to_csv(rows, ITEM_TABLE_HEADERS)
        elif get_erp_system() == "sap_b1":
            conn = get_sqlserver_connection()
            try:
                rows = fetch_item_table_rows_from_sqlserver(conn)
                csv_data = rows_to_csv(rows, ITEM_TABLE_HEADERS)
            finally:
                conn.close()
        else:
            rows = list(fetch_rows(ITEM_TABLE_SQL))
            csv_data = rows_to_csv(rows, ITEM_TABLE_HEADERS)
    except Exception as exc:  # noqa: BLE001
        flash(f"エラーが発生しました: {exc}", "error")
        return redirect(url_for("index"))

    filename = "item_table.csv"
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/download/order-table", methods=["POST"])
def download_order_table():
    try:
        if get_erp_system() == "excel":
            rows = _load_excel_export_rows("order", ORDER_TABLE_HEADERS, request.files.get("source_excel"))
            csv_data = rows_to_csv(rows, ORDER_TABLE_HEADERS)
        elif get_erp_system() == "sap_b1":
            conn = get_sqlserver_connection()
            try:
                rows = fetch_order_rows_from_sqlserver(conn)
                csv_data = rows_to_csv(rows, ORDER_TABLE_HEADERS)
            finally:
                conn.close()
        else:
            rows = list(fetch_rows(ORDER_TABLE_SQL))
            csv_data = rows_to_csv(rows, ORDER_TABLE_HEADERS)
    except Exception as exc:  # noqa: BLE001
        flash(f"エラーが発生しました: {exc}", "error")
        return redirect(url_for("index"))

    filename = "order_table.csv"
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/download/prd-plan-table", methods=["POST"])
def download_prd_plan_table():
    try:
        if get_erp_system() == "excel":
            rows = _load_excel_export_rows("prd_plan", ORDER_TABLE_HEADERS, request.files.get("source_excel"))
            csv_data = rows_to_csv(rows, ORDER_TABLE_HEADERS)
        else:
            raise RuntimeError("Prd Plan は Excel 取り込みのみ対応です。")
    except Exception as exc:  # noqa: BLE001
        flash(f"エラーが発生しました: {exc}", "error")
        return redirect(url_for("index"))

    filename = "prd_plan_table.csv"
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/download/resource-table", methods=["POST"])
def download_resource_table():
    if get_erp_system() == "sap_b1":
        _sap_b1_not_supported_flash()
        return redirect(url_for("index"))
    try:
        if get_erp_system() == "excel":
            rows = _load_excel_export_rows("resource", RESOURCE_TABLE_HEADERS, request.files.get("source_excel"))
        else:
            base_rows = list(fetch_rows(RESOURCE_TABLE_SQL))
            sort_map = _load_sort_order_map_from_sqlite()
            rows = []
            for line_cd, line_nm, resource_grp in base_rows:
                sort_order = sort_map.get(str(line_cd or "").strip())
                rows.append((line_cd, line_nm, resource_grp, "" if sort_order is None else sort_order))
        csv_data = rows_to_csv(rows, RESOURCE_TABLE_HEADERS)
    except Exception as exc:  # noqa: BLE001
        flash(f"エラーが発生しました: {exc}", "error")
        return redirect(url_for("index"))

    filename = "resource_table.csv"
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/download/inventory-table", methods=["POST"])
def download_inventory_table():
    try:
        if get_erp_system() == "excel":
            rows = _load_excel_export_rows("inventory", INVENTORY_TABLE_HEADERS, request.files.get("source_excel"))
            csv_data = rows_to_csv(rows, INVENTORY_TABLE_HEADERS)
        elif get_erp_system() == "sap_b1":
            conn = get_sqlserver_connection()
            try:
                rows = fetch_inventory_rows_from_sqlserver(conn)
                csv_data = rows_to_csv(rows, INVENTORY_TABLE_HEADERS)
            finally:
                conn.close()
        else:
            rows = list(fetch_rows(INVENTORY_TABLE_SQL))
            csv_data = rows_to_csv(rows, INVENTORY_TABLE_HEADERS)
    except Exception as exc:  # noqa: BLE001
        flash(f"エラーが発生しました: {exc}", "error")
        return redirect(url_for("index"))

    filename = "inventory_table.csv"
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/download/inventory-wip-table", methods=["POST"])
def download_inventory_wip_table():
    try:
        if get_erp_system() != "excel":
            raise RuntimeError("Inv. WIP は Excel 取り込みのみ対応です。")
        rows = _load_excel_export_rows("inventory_wip", INVENTORY_TABLE_HEADERS, request.files.get("source_excel"))
        csv_data = rows_to_csv(rows, INVENTORY_TABLE_HEADERS)
    except Exception as exc:  # noqa: BLE001
        flash(f"エラーが発生しました: {exc}", "error")
        return redirect(url_for("index"))

    filename = "inventory_wip_table.csv"
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/resource-cycle-times", methods=["GET"])
def resource_cycle_times():
    source_name = str(session.get(RESOURCE_CYCLE_SOURCE_SESSION_KEY) or "").strip()
    active_co_cd = _resource_cycle_scope_key()
    try:
        view_rows = _fetch_resource_cycle_rows()
    except Exception as exc:  # noqa: BLE001
        flash(f"サイクルタイムSQLiteの読み込みに失敗しました: {exc}", "error")
        view_rows = []
    return render_template(
        "bridge_cycle_time.html",
        oracle_connected=bool(session.get("oracle_connected")),
        erp_system=session.get("erp_system") or "mcframe",
        customer_profiles=_customer_profile_options(),
        selected_customer_id=str(session.get("customer_id") or ""),
        active_co_cd=active_co_cd,
        source_name=source_name,
        rows=view_rows,
    )


@app.route("/resource-cycle-times", methods=["POST"])
def save_resource_cycle_times():
    try:
        rows = _fetch_resource_cycle_rows()
    except Exception as exc:  # noqa: BLE001
        flash(f"サイクルタイムSQLiteの読み込みに失敗しました: {exc}", "error")
        return redirect(url_for("index", show_line_cycle="1"))
    try:
        if not rows:
            flash("保存対象がありません。先に Resource CSV を取り込んでください。", "error")
            return redirect(url_for("index", show_line_cycle="1"))
        updates: dict[int, dict[str, str | int]] = {}
        for row in rows:
            row_id = int(row["id"])
            cycle_time = (request.form.get(f"cycle_time_{row_id}") or "").strip()
            sort_raw = (request.form.get(f"sort_order_{row_id}") or "").strip()
            if sort_raw == "":
                sort_order = int(row.get("sort_order") or 0)
            else:
                try:
                    sort_order = int(sort_raw)
                except ValueError as exc:
                    raise RuntimeError(f"Sort Order は整数で入力してください（ID={row_id}）。") from exc
            updates[row_id] = {"cycle_time": cycle_time, "sort_order": sort_order}
        _update_resource_cycle_rows_by_id(updates)
        flash("サイクルタイムを保存しました。", "success")
    except Exception as exc:  # noqa: BLE001
        flash(f"サイクルタイムSQLiteの保存に失敗しました: {exc}", "error")
    return redirect(url_for("index", show_line_cycle="1"))


@app.route("/resource-cycle-times/delete-row", methods=["POST"])
def delete_resource_cycle_time_row():
    row_id_raw = (request.form.get("row_id") or "").strip()
    if not row_id_raw.isdigit():
        flash("削除対象の行番号が不正です。", "error")
        return redirect(url_for("index", show_line_cycle="1"))
    row_id = int(row_id_raw)
    try:
        deleted = _delete_resource_cycle_row_by_id(row_id)
        if deleted <= 0:
            raise RuntimeError("削除対象の行が見つかりません。")
        flash("行を削除しました。", "success")
    except Exception as exc:  # noqa: BLE001
        flash(f"行の削除に失敗しました: {exc}", "error")
    return redirect(url_for("index", show_line_cycle="1"))


@app.route("/resource-cycle-times/import", methods=["POST"])
def import_resource_cycle_times():
    upload = request.files.get("resource_csv")
    if not upload or not (upload.filename or "").strip():
        flash("取り込む Resource CSV ファイルを選択してください。", "error")
        return redirect(url_for("index", show_line_cycle="1"))
    try:
        raw = upload.read()
        if not raw:
            raise RuntimeError("アップロードされたファイルが空です。")
        rows = _read_resource_cycle_rows_from_bytes(raw)
        file_name = secure_filename(upload.filename or "resource_table.csv")
        if not file_name:
            file_name = "resource_table.csv"
        if not file_name.lower().endswith(".csv"):
            file_name = f"{file_name}.csv"
        _replace_resource_cycle_rows(rows, file_name, _resource_cycle_scope_key())
        session[RESOURCE_CYCLE_SOURCE_SESSION_KEY] = file_name
        flash(f"Resource CSV を取り込みました（SQLite保存）: {file_name}", "success")
    except Exception as exc:  # noqa: BLE001
        flash(f"Resource CSV の取り込みに失敗しました: {exc}", "error")
    return redirect(url_for("index", show_line_cycle="1"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5001")), debug=True)

