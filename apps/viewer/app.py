from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
import sqlite3
import csv
import os
import io
from datetime import datetime, timedelta
import calendar
import math
from collections import defaultdict
import sys
from pathlib import Path
from typing import Optional
from jinja2 import ChoiceLoader, FileSystemLoader

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    Workbook = None
from werkzeug.utils import secure_filename

PLATFORM_ROOT = Path(__file__).resolve().parents[2]
COMMON_DIR = PLATFORM_ROOT / "common"

sys.path.insert(0, str(PLATFORM_ROOT))
from config.settings import DATA_DIR, DB_PATH, UPLOAD_FOLDER
from config.bridge_customers import BRIDGE_CUSTOMERS
from core.asprova_parser import (
    detect_columns,
    parse_schedule_upload_row,
    result_csv_export_headers,
    schedule_row_to_result_csv_cells,
)
from core.input_instruction_parser import (
    detect_input_instruction_columns,
    parse_input_instruction_row,
)
from core.output_instruction_parser import (
    detect_output_instruction_columns,
    parse_output_instruction_row,
)
from core.csv_loader import csv_dict_reader_from_bytes
from core.sap_integrated_master import ensure_integrates_table

app = Flask(__name__, static_folder=str(COMMON_DIR / "static"))
app.jinja_loader = ChoiceLoader(
    [FileSystemLoader(str(COMMON_DIR / "templates")), app.jinja_loader]
)
app.secret_key = 'asprova-schedule-key-2024'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
ALLOWED_EXTENSIONS = {'csv'}


GANTT_PAGE_HEADERS = {
    'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0',
    'Pragma': 'no-cache',
    'X-Asprova-Gantt-Revision': '5',
    # 一部環境で X-* のみ除去される場合の予備（レスポンスヘッダ一覧に出るか確認用）
    'Asprova-Gantt-Revision': '5',
}


def _apply_gantt_cache_headers(response):
    for k, v in GANTT_PAGE_HEADERS.items():
        response.headers[k] = v
    return response


@app.after_request
def _no_store_gantt(response):
    path = (request.path or '').rstrip('/') or '/'
    if path == '/gantt' or request.endpoint == 'gantt':
        _apply_gantt_cache_headers(response)
    return response


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _sqlite_row_as_dict(row: sqlite3.Row) -> dict:
    """Normalize sqlite3.Row to a dict (reliable access to optional columns like actual_quantity)."""
    return {k: row[k] for k in row.keys()}


def _sched_row_to_gantt_task(r: sqlite3.Row) -> Optional[dict]:
    """
    One schedule row → gantt task dict.
    plan_* = DB plan (start_time, end_time, machine_name).
    start/end/machine = display (actual_start/end/resource when both actual times are set, else plan).
    """
    rd = _sqlite_row_as_dict(r)
    try:
        plan_s = datetime.strptime(r['start_time'], '%Y-%m-%d %H:%M:%S')
        plan_e = (
            datetime.strptime(r['end_time'], '%Y-%m-%d %H:%M:%S')
            if r['end_time']
            else plan_s + timedelta(hours=1)
        )
    except Exception:
        return None
    plan_machine = (
        (r['machine_name'] or '').strip()
        or (r['machine_id'] or '').strip()
        or 'Unknown'
    )

    act_s_raw = rd.get('actual_start')
    act_e_raw = rd.get('actual_end')
    act_res = rd.get('actual_resource')
    disp_s, disp_e = plan_s, plan_e
    disp_m = plan_machine
    if act_s_raw and act_e_raw:
        try:
            ast = datetime.strptime(str(act_s_raw).strip(), '%Y-%m-%d %H:%M:%S')
            aen = datetime.strptime(str(act_e_raw).strip(), '%Y-%m-%d %H:%M:%S')
            disp_s, disp_e = ast, aen
            if act_res is not None and str(act_res).strip() != '':
                disp_m = str(act_res).strip()
        except (TypeError, ValueError):
            disp_s, disp_e = plan_s, plan_e
            disp_m = plan_machine

    return {
        'id': r['id'],
        'machine': disp_m,
        'machine_id': r['machine_id'] or '',
        'plan_machine': plan_machine,
        'order_id': r['order_id'] or '',
        'order_item_code': r['order_item_code'] or '',
        'operation_id': r['operation_id'] or '',
        'next_operation_id': r['next_operation_id'] or '',
        'operation_code': r['operation_code'] or '',
        'next_operation_code': r['next_operation_code'] or '',
        'operation_out_item': r['operation_out_item'] or '',
        'item_id': r['item_id'] or '',
        'item_name': r['item_name'] or r['item_id'] or '',
        'process_name': r['process_name'] or '',
        'start': disp_s.isoformat(),
        'end': disp_e.isoformat(),
        'plan_start': plan_s.isoformat(),
        'plan_end': plan_e.isoformat(),
        'status': r['status'] or 'Scheduled',
        'quantity': r['quantity'],
        'actual_quantity': rd.get('actual_quantity'),
        'actual_start': act_s_raw,
        'actual_end': act_e_raw,
        'actual_resource': act_res,
        'setup_minutes': r['setup_minutes'],
        'work_group': rd.get('work_group') or '',
        'work_user_res_order': rd.get('work_user_res_order') or '',
        'delivery_date': rd.get('delivery_date') or '',
        'delivery_order_no': rd.get('delivery_order_no') or '',
        'delivery_item': rd.get('delivery_item') or '',
        'delivery_item_name': rd.get('delivery_item_name') or '',
        'min_skill': rd.get('min_skill') or '',
        'qc_skill': rd.get('qc_skill') or '',
    }


def _gantt_range_sql_clause():
    """WHERE fragment: plan window or actual window intersects [start_date, end_date)."""
    return (
        '((start_time < ? AND end_time > ?) OR '
        '(actual_start IS NOT NULL AND actual_end IS NOT NULL '
        'AND actual_start < ? AND actual_end > ?))'
    )


def _parse_work_user_res_order_val(v) -> float:
    """Numeric sort key for WorkUser_ResOrder; empty / invalid → +inf (sort last)."""
    if v is None:
        return float("inf")
    s = str(v).strip()
    if not s:
        return float("inf")
    try:
        x = float(s)
    except ValueError:
        return float("inf")
    if not math.isfinite(x):
        return float("inf")
    return x


def _gantt_machines_sorted_for_dropdown(machine_rows: list, tasks: list) -> list:
    """
    Resource list: ascending by min WorkUser_ResOrder among visible tasks for that row's
    display resource, then by resource name. Machines with no tasks in view sort last (A–Z).
    """
    names = sorted(
        {
            str(r["machine_name"]).strip()
            for r in machine_rows
            if r["machine_name"] is not None and str(r["machine_name"]).strip() != ""
        }
    )
    seen: set[str] = set()
    from_tasks: list[str] = []
    for t in tasks:
        m = t.get("machine")
        if not m or m in seen:
            continue
        seen.add(m)
        from_tasks.append(m)

    def min_ord(machine: str) -> float:
        vals = [
            _parse_work_user_res_order_val(x.get("work_user_res_order"))
            for x in tasks
            if x.get("machine") == machine
        ]
        return min(vals) if vals else float("inf")

    from_tasks.sort(key=lambda m: (min_ord(m), m))
    rest = sorted([m for m in names if m not in seen])
    ordered = from_tasks + rest
    return [{"machine_name": m} for m in ordered]


@app.context_processor
def inject_global_stats():
    # Makes schedule count available to all templates (e.g., for showing Clear button).
    try:
        conn = get_db()
        total = conn.execute('SELECT COUNT(*) as c FROM schedules').fetchone()['c']
        conn.close()
    except Exception:
        total = 0
    options = [
        {"id": k, "label": str(v.get("label") or k)}
        for k, v in BRIDGE_CUSTOMERS.items()
        if isinstance(v, dict)
    ]
    selected_customer_id = str(session.get("viewer_customer_id") or "").strip().lower()
    connected = bool(selected_customer_id and any(o["id"] == selected_customer_id for o in options))
    return {
        'schedule_total': total,
        'viewer_customer_profiles': options,
        'viewer_selected_customer_id': selected_customer_id,
        'viewer_customer_connected': connected,
    }


def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT,
            order_item_code TEXT,
            operation_id TEXT,
            next_operation_id TEXT,
            operation_code TEXT,
            next_operation_code TEXT,
            operation_out_item TEXT,
            item_id TEXT,
            item_name TEXT,
            machine_id TEXT,
            machine_name TEXT,
            start_time TEXT,
            end_time TEXT,
            quantity REAL,
            actual_quantity REAL,
            status TEXT,
            process_name TEXT,
            setup_minutes REAL,
            actual_start TEXT,
            actual_end TEXT,
            actual_resource TEXT,
            work_group TEXT,
            work_user_res_order TEXT,
            delivery_date TEXT,
            delivery_order_no TEXT,
            delivery_item TEXT,
            delivery_item_name TEXT,
            min_skill TEXT,
            qc_skill TEXT,
            uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS uploads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            row_count INTEGER,
            uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    ensure_integrates_table(conn)
    conn.commit()
    conn.close()

def ensure_db_schema():
    """
    Ensure newer columns exist when upgrading an existing schedule.db.
    SQLite doesn't support ADD COLUMN IF NOT EXISTS in all versions; use PRAGMA.
    """
    conn = get_db()
    try:
        cols = {r["name"] for r in conn.execute("PRAGMA table_info(schedules)").fetchall()}
        if "order_item_code" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN order_item_code TEXT")
        if "operation_id" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN operation_id TEXT")
        if "next_operation_id" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN next_operation_id TEXT")
        if "operation_code" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN operation_code TEXT")
        if "next_operation_code" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN next_operation_code TEXT")
        if "operation_out_item" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN operation_out_item TEXT")
        if "setup_minutes" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN setup_minutes REAL")
        if "actual_quantity" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN actual_quantity REAL")
        if "actual_start" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN actual_start TEXT")
        if "actual_end" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN actual_end TEXT")
        if "actual_resource" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN actual_resource TEXT")
        if "work_group" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN work_group TEXT")
        if "work_user_res_order" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN work_user_res_order TEXT")
        if "delivery_date" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN delivery_date TEXT")
        if "delivery_order_no" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN delivery_order_no TEXT")
        if "delivery_item" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN delivery_item TEXT")
        if "delivery_item_name" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN delivery_item_name TEXT")
        if "min_skill" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN min_skill TEXT")
        if "qc_skill" not in cols:
            conn.execute("ALTER TABLE schedules ADD COLUMN qc_skill TEXT")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS psi_input_instructions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_code TEXT NOT NULL,
                inst_time TEXT NOT NULL,
                quantity REAL,
                u_quantity REAL,
                qty_fixed_level REAL,
                qty_fixed_level_user_specified TEXT,
                pegging_method TEXT,
                object_id TEXT,
                object_status_flag_ext TEXT,
                flag_date TEXT,
                operation_code TEXT,
                source_filename TEXT,
                uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS psi_input_instruction_uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT,
                row_count INTEGER,
                uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS psi_output_instructions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_code TEXT NOT NULL,
                inst_time TEXT NOT NULL,
                quantity REAL,
                u_quantity REAL,
                qty_fixed_level REAL,
                qty_fixed_level_user_specified TEXT,
                pegging_method TEXT,
                object_id TEXT,
                object_status_flag_ext TEXT,
                flag_date TEXT,
                operation_code TEXT,
                source_filename TEXT,
                uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS psi_output_instruction_uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT,
                row_count INTEGER,
                uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        ensure_integrates_table(conn)
        conn.commit()
    finally:
        conn.close()


def get_latest_schedule_date():
    """
    Return latest date (YYYY-MM-DD) from schedules.start_time, or None if no rows.
    start_time is stored as 'YYYY-MM-DD HH:MM:SS', so substr(start_time, 1, 10) is safe.
    """
    conn = get_db()
    row = conn.execute("SELECT MAX(substr(start_time, 1, 10)) AS d FROM schedules").fetchone()
    conn.close()
    return row["d"] if row and row["d"] else None


def get_earliest_schedule_date():
    """
    Return earliest date (YYYY-MM-DD) from schedules.start_time, or None if no rows.
    """
    conn = get_db()
    row = conn.execute("SELECT MIN(substr(start_time, 1, 10)) AS d FROM schedules").fetchone()
    conn.close()
    return row["d"] if row and row["d"] else None


def get_earliest_psi_instruction_date():
    """Earliest inst_time date (YYYY-MM-DD) across PSI input/output instruction tables."""
    conn = get_db()
    dates = []
    for tbl in ("psi_output_instructions", "psi_input_instructions"):
        try:
            row = conn.execute(
                f"SELECT MIN(substr(inst_time, 1, 10)) AS d FROM {tbl}"
            ).fetchone()
            if row and row["d"]:
                dates.append(row["d"])
        except Exception:
            pass
    conn.close()
    return min(dates) if dates else None


@app.route('/')
def index():
    # Default landing page is the gantt view.
    return redirect(url_for('gantt'))


@app.route('/viewer/connect', methods=['POST'])
def viewer_connect():
    customer_id = (request.form.get('customer_id') or '').strip().lower()
    if customer_id and customer_id not in BRIDGE_CUSTOMERS:
        flash('選択したCustomerが不正です。', 'error')
        return redirect(request.referrer or url_for('gantt'))
    if customer_id:
        session['viewer_customer_id'] = customer_id
        flash('Viewer customer connected.', 'success')
    else:
        session.pop('viewer_customer_id', None)
        flash('Viewer customer cleared.', 'success')
    return redirect(request.referrer or url_for('gantt'))


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        ensure_db_schema()
        uploaded_files = request.files.getlist('files')
        candidates = [f for f in uploaded_files if f and f.filename and f.filename.strip() != '']
        if not candidates:
            flash('No file selected', 'error')
            return redirect(request.url)

        conn = get_db()
        total_rows = 0
        imported_files = []
        skipped = []

        for file in candidates:
            if not allowed_file(file.filename):
                skipped.append(f'{file.filename} (not CSV)')
                continue
            filename = secure_filename(file.filename)
            try:
                content_bytes = file.read()
                reader, _, _ = csv_dict_reader_from_bytes(content_bytes)
                headers = reader.fieldnames or []
            except Exception:
                skipped.append(f'{file.filename} (read error)')
                continue

            if not headers:
                skipped.append(f'{file.filename} (empty or invalid)')
                continue

            mapping = detect_columns(headers)
            rows = list(reader)
            count = 0
            for row in rows:
                rec = parse_schedule_upload_row(row, mapping)
                if not rec:
                    continue

                conn.execute('''
                    INSERT INTO schedules 
                    (order_id, order_item_code, operation_id, next_operation_id, operation_code, next_operation_code, operation_out_item, item_id, item_name, machine_id, machine_name, start_time, end_time, quantity, status, process_name, setup_minutes, actual_start, actual_end, actual_resource, work_group, work_user_res_order, delivery_date, delivery_order_no, delivery_item, delivery_item_name, min_skill, qc_skill)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    rec['order_id'],
                    rec['order_item_code'],
                    rec['operation_id'],
                    rec['next_operation_id'],
                    rec['operation_code'],
                    rec['next_operation_code'],
                    rec['operation_out_item'],
                    rec['item_id'],
                    rec['item_name'],
                    rec['machine_id'],
                    rec['machine_name'],
                    rec['start_time'],
                    rec['end_time'],
                    rec['quantity'],
                    rec['status'],
                    rec['process_name'],
                    rec['setup_minutes'],
                    rec.get('actual_start'),
                    rec.get('actual_end'),
                    rec.get('actual_resource'),
                    rec.get('work_group'),
                    rec.get('work_user_res_order'),
                    rec.get('delivery_date'),
                    rec.get('delivery_order_no'),
                    rec.get('delivery_item'),
                    rec.get('delivery_item_name'),
                    rec.get('min_skill'),
                    rec.get('qc_skill'),
                ))
                count += 1

            conn.execute('INSERT INTO uploads (filename, row_count) VALUES (?, ?)', (filename, count))
            total_rows += count
            imported_files.append(filename)

        conn.commit()
        conn.close()

        if skipped:
            flash('Skipped: ' + '; '.join(skipped[:12]) + ('…' if len(skipped) > 12 else ''), 'warning')
        if total_rows == 0:
            flash('No schedule rows imported from the selected file(s)', 'error')
            return redirect(url_for('upload'))
        flash(
            f'Successfully imported {total_rows} schedule record(s) from {len(imported_files)} file(s): '
            + ', '.join(imported_files[:8])
            + ('…' if len(imported_files) > 8 else ''),
            'success',
        )
        return redirect(url_for('upload'))

    ctx = get_schedule_context()
    conn = get_db()
    prior = conn.execute('SELECT DISTINCT filename FROM uploads ORDER BY filename').fetchall()
    conn.close()
    ctx['prior_upload_names_lower'] = [r['filename'].lower() for r in prior]
    return render_template('upload2.html', **ctx)


@app.route('/upload/input-instruction-psi', methods=['GET', 'POST'])
def upload_input_instruction_psi():
    """Import Input Instruction (PSI) CSV (e.g. inputinst.csv) into psi_input_instructions."""
    if request.method == 'POST':
        ensure_db_schema()
        uploaded_files = request.files.getlist('files')
        candidates = [f for f in uploaded_files if f and f.filename and f.filename.strip() != '']
        if not candidates:
            flash('No file selected', 'error')
            return redirect(request.url)

        conn = get_db()
        total_rows = 0
        imported_files = []
        skipped = []

        for file in candidates:
            if not allowed_file(file.filename):
                skipped.append(f'{file.filename} (not CSV)')
                continue
            filename = secure_filename(file.filename)
            try:
                content_bytes = file.read()
                reader, _, _ = csv_dict_reader_from_bytes(content_bytes)
                headers = reader.fieldnames or []
            except Exception:
                skipped.append(f'{file.filename} (read error)')
                continue

            if not headers:
                skipped.append(f'{file.filename} (empty or invalid)')
                continue

            mapping = detect_input_instruction_columns(headers)
            if not mapping.get('item_code') or not mapping.get('inst_time'):
                skipped.append(f'{file.filename} (not Input Instruction CSV — need item & time columns)')
                continue

            rows = list(reader)
            count = 0
            for row in rows:
                rec = parse_input_instruction_row(row, mapping)
                if not rec:
                    continue
                conn.execute(
                    """
                    INSERT INTO psi_input_instructions (
                        item_code, inst_time, quantity, u_quantity, qty_fixed_level,
                        qty_fixed_level_user_specified, pegging_method, object_id,
                        object_status_flag_ext, flag_date, operation_code, source_filename
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        rec['item_code'],
                        rec['inst_time'],
                        rec['quantity'],
                        rec['u_quantity'],
                        rec['qty_fixed_level'],
                        rec['qty_fixed_level_user_specified'],
                        rec['pegging_method'],
                        rec['object_id'],
                        rec['object_status_flag_ext'],
                        rec['flag_date'],
                        rec['operation_code'],
                        filename,
                    ),
                )
                count += 1

            conn.execute(
                'INSERT INTO psi_input_instruction_uploads (filename, row_count) VALUES (?, ?)',
                (filename, count),
            )
            total_rows += count
            imported_files.append(filename)

        conn.commit()
        conn.close()

        if skipped:
            flash('Skipped: ' + '; '.join(skipped[:12]) + ('…' if len(skipped) > 12 else ''), 'warning')
        if total_rows == 0:
            flash('No Input Instruction rows imported from the selected file(s)', 'error')
            return redirect(url_for('upload_input_instruction_psi'))
        flash(
            f'Successfully imported {total_rows} Input Instruction record(s) from {len(imported_files)} file(s): '
            + ', '.join(imported_files[:8])
            + ('…' if len(imported_files) > 8 else ''),
            'success',
        )
        return redirect(url_for('upload_input_instruction_psi'))

    ensure_db_schema()
    conn = get_db()
    try:
        ii_total = conn.execute('SELECT COUNT(*) AS c FROM psi_input_instructions').fetchone()['c']
    except Exception:
        ii_total = 0
    ii_uploads = conn.execute(
        'SELECT * FROM psi_input_instruction_uploads ORDER BY uploaded_at DESC LIMIT 8'
    ).fetchall()
    prior = conn.execute(
        'SELECT DISTINCT filename FROM psi_input_instruction_uploads ORDER BY filename'
    ).fetchall()
    conn.close()
    return render_template(
        'upload_input_instruction_psi.html',
        ii_total=ii_total,
        ii_uploads=ii_uploads,
        prior_upload_names_lower=[
            r['filename'].lower() for r in prior if r['filename']
        ],
    )


@app.route('/upload/output-instruction-psi', methods=['GET', 'POST'])
def upload_output_instruction_psi():
    """Import Output Instruction (PSI) CSV (e.g. outputinst.csv) into psi_output_instructions."""
    if request.method == 'POST':
        ensure_db_schema()
        uploaded_files = request.files.getlist('files')
        candidates = [f for f in uploaded_files if f and f.filename and f.filename.strip() != '']
        if not candidates:
            flash('No file selected', 'error')
            return redirect(request.url)

        conn = get_db()
        total_rows = 0
        imported_files = []
        skipped = []

        for file in candidates:
            if not allowed_file(file.filename):
                skipped.append(f'{file.filename} (not CSV)')
                continue
            filename = secure_filename(file.filename)
            try:
                content_bytes = file.read()
                reader, _, _ = csv_dict_reader_from_bytes(content_bytes)
                headers = reader.fieldnames or []
            except Exception:
                skipped.append(f'{file.filename} (read error)')
                continue

            if not headers:
                skipped.append(f'{file.filename} (empty or invalid)')
                continue

            mapping = detect_output_instruction_columns(headers)
            if not mapping.get('item_code') or not mapping.get('inst_time'):
                skipped.append(f'{file.filename} (not Output Instruction CSV — need item & time columns)')
                continue

            rows = list(reader)
            count = 0
            for row in rows:
                rec = parse_output_instruction_row(row, mapping)
                if not rec:
                    continue
                conn.execute(
                    """
                    INSERT INTO psi_output_instructions (
                        item_code, inst_time, quantity, u_quantity, qty_fixed_level,
                        qty_fixed_level_user_specified, pegging_method, object_id,
                        object_status_flag_ext, flag_date, operation_code, source_filename
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        rec['item_code'],
                        rec['inst_time'],
                        rec['quantity'],
                        rec['u_quantity'],
                        rec['qty_fixed_level'],
                        rec['qty_fixed_level_user_specified'],
                        rec['pegging_method'],
                        rec['object_id'],
                        rec['object_status_flag_ext'],
                        rec['flag_date'],
                        rec['operation_code'],
                        filename,
                    ),
                )
                count += 1

            conn.execute(
                'INSERT INTO psi_output_instruction_uploads (filename, row_count) VALUES (?, ?)',
                (filename, count),
            )
            total_rows += count
            imported_files.append(filename)

        conn.commit()
        conn.close()

        if skipped:
            flash('Skipped: ' + '; '.join(skipped[:12]) + ('…' if len(skipped) > 12 else ''), 'warning')
        if total_rows == 0:
            flash('No Output Instruction rows imported from the selected file(s)', 'error')
            return redirect(url_for('upload_output_instruction_psi'))
        flash(
            f'Successfully imported {total_rows} Output Instruction record(s) from {len(imported_files)} file(s): '
            + ', '.join(imported_files[:8])
            + ('…' if len(imported_files) > 8 else ''),
            'success',
        )
        return redirect(url_for('upload_output_instruction_psi'))

    ensure_db_schema()
    conn = get_db()
    try:
        oi_total = conn.execute('SELECT COUNT(*) AS c FROM psi_output_instructions').fetchone()['c']
    except Exception:
        oi_total = 0
    oi_uploads = conn.execute(
        'SELECT * FROM psi_output_instruction_uploads ORDER BY uploaded_at DESC LIMIT 8'
    ).fetchall()
    prior = conn.execute(
        'SELECT DISTINCT filename FROM psi_output_instruction_uploads ORDER BY filename'
    ).fetchall()
    conn.close()
    return render_template(
        'upload_output_instruction_psi.html',
        oi_total=oi_total,
        oi_uploads=oi_uploads,
        prior_upload_names_lower=[
            r['filename'].lower() for r in prior if r['filename']
        ],
    )


@app.route('/export/schedules.csv')
def export_schedules_csv():
    """Download result.csv: Work_Code, Actual_Start, Actual_End, Actual_Resource, actual_quantity."""
    ensure_db_schema()
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM schedules ORDER BY machine_name, start_time, id'
    ).fetchall()
    conn.close()

    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator='\r\n')
    writer.writerow(result_csv_export_headers())
    for row in rows:
        writer.writerow(schedule_row_to_result_csv_cells(row))

    data = buf.getvalue().encode('utf-8-sig')
    mem = io.BytesIO(data)
    mem.seek(0)
    return send_file(
        mem,
        mimetype='text/csv; charset=utf-8',
        as_attachment=True,
        download_name='result.csv',
    )


def get_schedule_context(view=None, date_str=None, machine_filter=None):
    """Build context dict for schedule section (used by schedule page and upload page)."""
    view = view or request.args.get('view', 'weekly')
    if date_str is not None:
        date_str = date_str or datetime.now().strftime('%Y-%m-%d')
    else:
        if 'date' in request.args:
            date_str = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
        else:
            date_str = get_earliest_schedule_date() or datetime.now().strftime('%Y-%m-%d')
    machine_filter = machine_filter if machine_filter is not None else request.args.get('machine', '')
    machine_filter = machine_filter.strip() if isinstance(machine_filter, str) else ''

    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()

    if view in ('weekly', '2week', '3week'):
        start_date = current_date - timedelta(days=current_date.weekday())
        span_days = {'weekly': 7, '2week': 14, '3week': 21}[view]
        end_date = start_date + timedelta(days=span_days)
    elif view == 'monthly':
        start_date = current_date.replace(day=1)
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1)
    else:
        start_date = current_date
        end_date = current_date + timedelta(days=1)

    conn = get_db()
    uploads = conn.execute('SELECT * FROM uploads ORDER BY uploaded_at DESC LIMIT 5').fetchall()
    total = conn.execute('SELECT COUNT(*) as c FROM schedules').fetchone()['c']
    machines = conn.execute(
        '''
        SELECT m AS machine_name FROM (
            SELECT DISTINCT TRIM(machine_name) AS m FROM schedules
            WHERE machine_name IS NOT NULL AND TRIM(machine_name) <> ''
            UNION
            SELECT DISTINCT TRIM(actual_resource) AS m FROM schedules
            WHERE actual_resource IS NOT NULL AND TRIM(actual_resource) <> ''
        )
        ORDER BY m
        '''
    ).fetchall()
    query = '''
        SELECT * FROM schedules 
        WHERE start_time >= ? AND start_time < ?
    '''
    params = [start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')]
    if machine_filter:
        query += ' AND machine_name = ?'
        params.append(machine_filter)
    query += ' ORDER BY machine_name, start_time'
    records = conn.execute(query, params).fetchall()
    conn.close()

    machine_schedules = {}
    for r in records:
        m = r['machine_name'] or 'Unknown'
        if m not in machine_schedules:
            machine_schedules[m] = []
        machine_schedules[m].append(dict(r))

    if view in ('weekly', '2week', '3week'):
        span_days = {'weekly': 7, '2week': 14, '3week': 21}[view]
        prev_date = (start_date - timedelta(days=span_days)).strftime('%Y-%m-%d')
        next_date = (start_date + timedelta(days=span_days)).strftime('%Y-%m-%d')
    elif view == 'monthly':
        if start_date.month == 1:
            prev_start = start_date.replace(year=start_date.year - 1, month=12, day=1)
        else:
            prev_start = start_date.replace(month=start_date.month - 1, day=1)
        if start_date.month == 12:
            next_start = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            next_start = start_date.replace(month=start_date.month + 1, day=1)
        prev_date = prev_start.strftime('%Y-%m-%d')
        next_date = next_start.strftime('%Y-%m-%d')
    else:
        prev_date = (start_date - timedelta(days=1)).strftime('%Y-%m-%d')
        next_date = (start_date + timedelta(days=1)).strftime('%Y-%m-%d')

    return dict(
        machine_schedules=machine_schedules,
        machines=machines,
        uploads=uploads,
        total=total,
        view=view,
        current_date=current_date,
        start_date=start_date,
        end_date=end_date,
        prev_date=prev_date,
        next_date=next_date,
        machine_filter=machine_filter,
    )


@app.route('/schedule')
def schedule():
    ctx = get_schedule_context()
    return render_template('schedule2.html', **ctx)


@app.route('/gantt')
def gantt():
    # Default to monthly when no view is specified (for header Gantt Chart link)
    view = request.args.get('view', 'monthly')
    machine_filter = (request.args.get('machine', '') or '').strip()
    item_filter = request.args.get('item', '')
    if 'date' in request.args:
        date_str = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    else:
        date_str = get_earliest_schedule_date() or datetime.now().strftime('%Y-%m-%d')

    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()

    if view in ('weekly', '2week', '3week'):
        start_date = current_date - timedelta(days=current_date.weekday())
        span_days = {'weekly': 7, '2week': 14, '3week': 21}[view]
        end_date = start_date + timedelta(days=span_days)
    elif view == 'monthly':
        start_date = current_date.replace(day=1)
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1)
    else:
        start_date = current_date
        end_date = current_date + timedelta(days=1)

    conn = get_db()
    machines_raw = conn.execute(
        '''
        SELECT m AS machine_name FROM (
            SELECT DISTINCT TRIM(machine_name) AS m FROM schedules
            WHERE machine_name IS NOT NULL
              AND TRIM(machine_name) <> ''
              AND TRIM(machine_name) <> 'DefaultInventoryResource'
            UNION
            SELECT DISTINCT TRIM(actual_resource) AS m FROM schedules
            WHERE actual_resource IS NOT NULL
              AND TRIM(actual_resource) <> ''
              AND TRIM(actual_resource) <> 'DefaultInventoryResource'
        )
        ORDER BY m
        '''
    ).fetchall()
    machines = [{"machine_name": r["machine_name"]} for r in machines_raw]

    # Distinct order item codes (WorkUser_OrderItem) for dropdown
    order_items = conn.execute(
        '''
        SELECT DISTINCT order_item_code 
        FROM schedules
        WHERE order_item_code IS NOT NULL AND order_item_code <> ''
        ORDER BY order_item_code
        '''
    ).fetchall()

    range_end = end_date.strftime('%Y-%m-%d %H:%M:%S')
    range_start = start_date.strftime('%Y-%m-%d %H:%M:%S')
    query = f'SELECT * FROM schedules WHERE {_gantt_range_sql_clause()}'
    params = [range_end, range_start, range_end, range_start]
    query += " AND TRIM(COALESCE(item_id, '')) <> ''"

    if machine_filter:
        query += ' AND (TRIM(machine_name) = ? OR TRIM(actual_resource) = ?)'
        params.extend([machine_filter, machine_filter])
    # item_filter is used only for link highlighting on frontend; keep data set complete

    query += ' ORDER BY machine_name, start_time'
    records = conn.execute(query, params).fetchall()
    conn.close()

    tasks = []
    for r in records:
        t = _sched_row_to_gantt_task(r)
        if t:
            tasks.append(t)

    # WorkUser_ResOrder → Resource 名の順（dropdown / 表示と一致）
    machines = _gantt_machines_sorted_for_dropdown(list(machines_raw), tasks)

    if view in ('weekly', '2week', '3week'):
        span_days = {'weekly': 7, '2week': 14, '3week': 21}[view]
        prev_date = (start_date - timedelta(days=span_days)).strftime('%Y-%m-%d')
        next_date = (start_date + timedelta(days=span_days)).strftime('%Y-%m-%d')
    elif view == 'monthly':
        if start_date.month == 1:
            prev_start = start_date.replace(year=start_date.year - 1, month=12, day=1)
        else:
            prev_start = start_date.replace(month=start_date.month - 1, day=1)
        if start_date.month == 12:
            next_start = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            next_start = start_date.replace(month=start_date.month + 1, day=1)
        prev_date = prev_start.strftime('%Y-%m-%d')
        next_date = next_start.strftime('%Y-%m-%d')
    else:
        prev_date = (start_date - timedelta(days=1)).strftime('%Y-%m-%d')
        next_date = (start_date + timedelta(days=1)).strftime('%Y-%m-%d')

    return (
        render_template(
            'gantt2.html',
            tasks=tasks,
            machines=machines,
            order_items=order_items,
            view=view,
            current_date=current_date,
            start_date=start_date,
            end_date=end_date,
            prev_date=prev_date,
            next_date=next_date,
            machine_filter=machine_filter,
            item_filter=item_filter,
        ),
        200,
        GANTT_PAGE_HEADERS,
    )


@app.route('/viewer-check')
@app.route('/__asprova_viewer_check')
def asprova_viewer_check():
    """このプロセスが読み込んだ app.py のパスとガント用ヘッダ定義を返す（別フォルダ起動の切り分け用）。"""
    return jsonify(
        ok=True,
        app_py=str(Path(__file__).resolve()),
        platform_root=str(PLATFORM_ROOT),
        gantt_page_revision='5',
        gantt_response_headers=dict(GANTT_PAGE_HEADERS),
    )


@app.route('/api/gantt_data')
def api_gantt_data():
    view = request.args.get('view', 'monthly')
    machine_filter = (request.args.get('machine', '') or '').strip()
    if 'date' in request.args:
        date_str = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    else:
        date_str = get_earliest_schedule_date() or datetime.now().strftime('%Y-%m-%d')

    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()

    if view in ('weekly', '2week', '3week'):
        start_date = current_date - timedelta(days=current_date.weekday())
        span_days = {'weekly': 7, '2week': 14, '3week': 21}[view]
        end_date = start_date + timedelta(days=span_days)
    elif view == 'monthly':
        start_date = current_date.replace(day=1)
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1)
    else:
        start_date = current_date
        end_date = current_date + timedelta(days=1)

    conn = get_db()
    range_end = end_date.strftime('%Y-%m-%d %H:%M:%S')
    range_start = start_date.strftime('%Y-%m-%d %H:%M:%S')
    query = f'SELECT * FROM schedules WHERE {_gantt_range_sql_clause()}'
    params = [range_end, range_start, range_end, range_start]
    query += " AND TRIM(COALESCE(item_id, '')) <> ''"
    if machine_filter:
        query += ' AND (TRIM(machine_name) = ? OR TRIM(actual_resource) = ?)'
        params.extend([machine_filter, machine_filter])
    query += ' ORDER BY machine_name, start_time'
    records = conn.execute(query, params).fetchall()
    conn.close()

    tasks = []
    for r in records:
        t = _sched_row_to_gantt_task(r)
        if t:
            tasks.append(t)
    return jsonify(tasks)


@app.route('/api/schedules/freeze_cutoff', methods=['POST'])
def api_schedules_freeze_cutoff():
    """
    Frozen: for rows with start_time on or before cutoff_date 23:59:59,
    set status 'D', fill null/empty actuals from plan, leave existing actuals unchanged.
    Optional JSON body key "machine": limit to that resource (machine_name or actual_resource).
    """
    data = request.get_json(silent=True) or {}
    raw = data.get('cutoff_date') or data.get('date')
    if not raw or not isinstance(raw, str):
        return jsonify({'ok': False, 'error': 'cutoff_date (YYYY-MM-DD) required'}), 400
    raw = raw.strip()
    try:
        datetime.strptime(raw, '%Y-%m-%d')
    except ValueError:
        return jsonify({'ok': False, 'error': 'Invalid cutoff_date'}), 400
    cutoff_end = f'{raw} 23:59:59'

    machine = data.get('machine') or data.get('resource') or ''
    machine = machine.strip() if isinstance(machine, str) else ''

    sql = '''
        UPDATE schedules SET
            status = 'D',
            actual_start = CASE
                WHEN actual_start IS NULL OR TRIM(COALESCE(actual_start, '')) = '' THEN start_time
                ELSE actual_start
            END,
            actual_end = CASE
                WHEN actual_end IS NULL OR TRIM(COALESCE(actual_end, '')) = '' THEN end_time
                ELSE actual_end
            END,
            actual_quantity = CASE
                WHEN actual_quantity IS NULL THEN quantity
                ELSE actual_quantity
            END
        WHERE start_time IS NOT NULL
          AND TRIM(start_time) <> ''
          AND start_time <= ?
    '''
    params = [cutoff_end]
    if machine:
        sql += ' AND (TRIM(machine_name) = ? OR TRIM(actual_resource) = ?)'
        params.extend([machine, machine])

    conn = get_db()
    cur = conn.execute(sql, params)
    n = cur.rowcount
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'updated': n, 'cutoff_date': raw, 'cutoff_end': cutoff_end})


@app.route('/api/schedules/<int:sched_id>/actual_quantity', methods=['POST'])
def api_schedule_actual_quantity(sched_id):
    """Update 実績数量 (actual quantity) for one schedule row."""
    data = request.get_json(silent=True) or {}
    raw = data.get('actual_quantity')
    conn = get_db()
    row = conn.execute('SELECT id FROM schedules WHERE id = ?', (sched_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'ok': False, 'error': 'Not found'}), 404
    if raw is None or (isinstance(raw, str) and raw.strip() == ''):
        conn.execute('UPDATE schedules SET actual_quantity = NULL WHERE id = ?', (sched_id,))
        stored = None
    else:
        try:
            val = float(raw)
        except (TypeError, ValueError):
            conn.close()
            return jsonify({'ok': False, 'error': 'Invalid number'}), 400
        conn.execute('UPDATE schedules SET actual_quantity = ? WHERE id = ?', (val, sched_id))
        stored = val
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'actual_quantity': stored})


def _parse_iso_datetime_plan(s):
    """Parse JSON/datetime string from client (ISO-8601, optional Z) to naive local datetime."""
    if not isinstance(s, str):
        raise ValueError('expected string')
    s = s.strip()
    if s.endswith('Z'):
        s = s[:-1] + '+00:00'
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is not None:
        dt = dt.replace(tzinfo=None)
    return dt


@app.route('/api/schedules/<int:sched_id>/plan_times', methods=['POST'])
def api_schedule_plan_times(sched_id):
    """Store gantt drag result in actual_start, actual_end, actual_resource (plan columns unchanged)."""
    data = request.get_json(silent=True) or {}
    start_raw = data.get('start')
    end_raw = data.get('end')
    machine_name = data.get('machine_name')
    if not start_raw or not end_raw:
        return jsonify({'ok': False, 'error': 'start and end required'}), 400
    try:
        s_dt = _parse_iso_datetime_plan(start_raw)
        e_dt = _parse_iso_datetime_plan(end_raw)
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Invalid datetime'}), 400
    if e_dt <= s_dt:
        return jsonify({'ok': False, 'error': 'end must be after start'}), 400

    conn = get_db()
    row = conn.execute('SELECT * FROM schedules WHERE id = ?', (sched_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'ok': False, 'error': 'Not found'}), 404

    start_txt = s_dt.strftime('%Y-%m-%d %H:%M:%S')
    end_txt = e_dt.strftime('%Y-%m-%d %H:%M:%S')
    if machine_name is not None and isinstance(machine_name, str) and machine_name.strip() != '':
        res_txt = machine_name.strip()
    else:
        res_txt = (row['machine_name'] or '').strip() or 'Unknown'

    conn.execute(
        'UPDATE schedules SET actual_start = ?, actual_end = ?, actual_resource = ? WHERE id = ?',
        (start_txt, end_txt, res_txt, sched_id),
    )
    conn.commit()
    row2 = conn.execute('SELECT * FROM schedules WHERE id = ?', (sched_id,)).fetchone()
    conn.close()
    task = _sched_row_to_gantt_task(row2) if row2 else None
    if not task:
        return jsonify({'ok': False, 'error': 'Row invalid after update'}), 500
    out = {'ok': True}
    out.update(task)
    return jsonify(out)


@app.route('/api/schedules/<int:sched_id>/clear_actual_results', methods=['POST'])
def api_schedule_clear_actual_results(sched_id):
    """
    Gantt Result Reset: clear actual_* overlays and snap resource label to plan
    (machine_id → machine_name when id is set). Plan start_time, end_time, quantity are unchanged.
    """
    conn = get_db()
    row = conn.execute('SELECT id FROM schedules WHERE id = ?', (sched_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'ok': False, 'error': 'Not found'}), 404
    conn.execute(
        '''
        UPDATE schedules SET
            actual_start = NULL,
            actual_end = NULL,
            actual_quantity = NULL,
            actual_resource = NULL,
            machine_name = COALESCE(NULLIF(TRIM(machine_id), ''), machine_name)
        WHERE id = ?
        ''',
        (sched_id,),
    )
    conn.commit()
    row2 = conn.execute('SELECT * FROM schedules WHERE id = ?', (sched_id,)).fetchone()
    conn.close()
    task = _sched_row_to_gantt_task(row2) if row2 else None
    if not task:
        return jsonify({'ok': False, 'error': 'Row invalid after update'}), 500
    out = {'ok': True}
    out.update(task)
    return jsonify(out)


@app.route('/export_monthly')
def export_monthly():
    """
    Export a monthly production plan to Excel.
    - Uses the same date logic as monthly view.
    - Optional query params: date=YYYY-MM-DD, machine=<name>
    """
    if Workbook is None:
        flash('Excel export requires openpyxl to be installed (pip install openpyxl).', 'error')
        return redirect(url_for('schedule'))

    view = request.args.get('view', 'monthly')
    # Always treat this export as monthly; ignore other views for now.
    # If no date is given, default to earliest schedule date so export is never "empty" when data exists.
    date_str = request.args.get('date') or get_earliest_schedule_date() or datetime.now().strftime('%Y-%m-%d')
    machine_filter = request.args.get('machine', '')

    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()

    # Monthly range: first day of month -> first day of next month
    start_date = current_date.replace(day=1)
    if start_date.month == 12:
        end_date = start_date.replace(year=start_date.year + 1, month=1)
    else:
        end_date = start_date.replace(month=start_date.month + 1)

    conn = get_db()
    query = '''
        SELECT * FROM schedules
        WHERE start_time >= ? AND start_time < ?
    '''
    params = [
        start_date.strftime('%Y-%m-%d'),
        end_date.strftime('%Y-%m-%d'),
    ]
    if machine_filter:
        query += ' AND machine_name = ?'
        params.append(machine_filter)
    query += ' ORDER BY machine_name, start_time'
    rows = conn.execute(query, params).fetchall()
    conn.close()

    # Create workbook
    wb = Workbook()

    # -------- Sheet 1: Calendar-style monthly plan (item x date, cell = machine + quantity) --------
    ws_cal = wb.active
    ws_cal.title = f'{start_date.strftime("%Y-%m")}_calendar'
    # Default zoom 80%
    ws_cal.sheet_view.zoomScale = 80
    ws_cal.sheet_view.zoomScaleNormal = 80

    header_font = Font(name='Meiryo', bold=True)
    header_fill = PatternFill('solid', fgColor='DDDDDD')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')

    # Collect days in the month
    day_list = []
    d = start_date
    while d < end_date:
        day_list.append(d)
        d += timedelta(days=1)

    # Aggregate quantity per (item, day, machine)
    agg = defaultdict(float)
    items = set()
    for r in rows:
        # Vertical axis key: Work_OperationOutMainItem (operation_out_item) 優先
        item_key = (
            r['operation_out_item']
            or r['order_item_code']
            or r['item_name']
            or r['item_id']
            or ''
        )
        if not item_key:
            continue
        items.add(item_key)
        m = r['machine_name'] or 'Unknown'
        if not r['start_time']:
            continue
        try:
            day = datetime.strptime(r['start_time'][:10], '%Y-%m-%d').date()
        except ValueError:
            continue
        qty = r['quantity'] if r['quantity'] is not None else 0
        agg[(item_key, day, m)] += qty

    items = sorted(items)

    # Header row: Item | 1/Mar etc. (per day)
    ws_cal.append(['Item'] + [f"{d.day}/{d.strftime('%b')}" for d in day_list])
    for col_idx in range(1, 2 + len(day_list)):
        cell = ws_cal.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # Data rows
    for item_key in items:
        row_vals = [item_key]
        for d in day_list:
            # Collect all machines for this item & day
            parts = []
            for (it, day, m), qty in agg.items():
                if it == item_key and day == d.date() and qty:
                    # machine and quantity on two lines overall (machine: qty)
                    parts.append(f'{m}: {qty:g}')
            cell_val = '\n'.join(parts) if parts else ''
            row_vals.append(cell_val)
        ws_cal.append(row_vals)

    # Style data cells
    default_font = Font(name='Meiryo')
    for row in ws_cal.iter_rows(min_row=2, max_row=ws_cal.max_row, min_col=1, max_col=1 + len(day_list)):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = default_font

    # Column widths
    # Item列はやや広めだが、指定に合わせて約40%狭くする（もともと22想定 -> 約13）
    ws_cal.column_dimensions['A'].width = 13
    for idx in range(len(day_list)):
        col_letter = ws_cal.cell(row=1, column=2 + idx).column_letter
        ws_cal.column_dimensions[col_letter].width = 11

    # -------- Sheet 2: Detail list (per operation) --------
    ws_det = wb.create_sheet(title=f'{start_date.strftime("%Y-%m")}_detail')
    ws_det.sheet_view.zoomScale = 80
    ws_det.sheet_view.zoomScaleNormal = 80
    headers = [
        'Machine',
        'Order ID',
        'Order Item Code',
        'Item',
        'Process',
        'Start',
        'End',
        'Quantity',
        'Status',
    ]
    ws_det.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws_det.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for r in rows:
        ws_det.append([
            r['machine_name'] or 'Unknown',
            r['order_id'] or '',
            r['order_item_code'] or '',
            r['item_name'] or r['item_id'] or '',
            r['process_name'] or '',
            r['start_time'] or '',
            r['end_time'] or '',
            r['quantity'] if r['quantity'] is not None else '',
            r['status'] or '',
        ])

    for row in ws_det.iter_rows(min_row=2, max_row=ws_det.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.font = default_font

    for col in ws_det.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ''
            max_len = max(max_len, len(val))
        ws_det.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))

    # -------- Sheet 3: PSI (Supply / Demand / Stock) --------
    ws_psi = wb.create_sheet(title=f'{start_date.strftime("%Y-%m")}_PSI')
    ws_psi.sheet_view.zoomScale = 80
    ws_psi.sheet_view.zoomScaleNormal = 80

    psi_header_fill = PatternFill('solid', fgColor='BCE597')  # RGB(188,229,151)
    psi_fill_even = PatternFill('solid', fgColor='FFFFFF')    # white
    psi_fill_odd = PatternFill('solid', fgColor='DDF2CA')     # RGB(221,242,202)

    # Sheet 3: same aggregates as web PSI (instructions, not schedules).
    (
        _,
        ordered_items_psi,
        _,
        supply_qty_psi,
        agg_psi,
        demand_qty_psi,
        demand_agg_psi,
    ) = _build_psi_data_for_month(start_date, end_date)
    day_labels = [_format_md_weekday_en(d) for d in day_list]

    ws_psi.append(['Item', 'Type'] + day_labels)
    for col_idx in range(1, 3 + len(day_list)):
        cell = ws_psi.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = psi_header_fill
        cell.border = thin_border
        cell.alignment = center

    for item_key in ordered_items_psi:
        stock_prev = 0.0
        for row_type in ('Supply', 'Demand', 'Stock'):
            row_vals = [item_key, row_type]
            for d in day_list:
                day_date = d.date()
                if row_type == 'Supply':
                    parts = []
                    for (it, day, m), qty in agg_psi.items():
                        if it == item_key and day == day_date and qty:
                            parts.append(f'{m}: {qty:g}')
                    cell_val = '\n'.join(parts) if parts else ''
                elif row_type == 'Demand':
                    parts = []
                    for (it, day, m), qty in demand_agg_psi.items():
                        if it == item_key and day == day_date and qty and float(qty) > 0:
                            parts.append(f'{m}: {qty:g}')
                    cell_val = '\n'.join(parts) if parts else ''
                else:
                    s = supply_qty_psi.get((item_key, day_date), 0.0)
                    d_qty = demand_qty_psi.get((item_key, day_date), 0.0)
                    stock = stock_prev + s - d_qty
                    stock_prev = stock
                    cell_val = stock if stock != 0 else ''
                row_vals.append(cell_val)
            ws_psi.append(row_vals)

    # Style PSI sheet
    for r_idx, row in enumerate(ws_psi.iter_rows(min_row=2, max_row=ws_psi.max_row, min_col=1, max_col=2 + len(day_list)), start=2):
        # Alternate fill per item (3 rows per item)
        item_group = (r_idx - 2) // 3
        fill = psi_fill_even if (item_group % 2 == 0) else psi_fill_odd
        for cell in row:
            cell.border = thin_border
            cell.fill = fill
            if cell.col_idx <= 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                # Wrap for Supply / Demand; numeric for Stock OK as center
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = default_font

    ws_psi.column_dimensions['A'].width = 22
    ws_psi.column_dimensions['B'].width = 10
    for idx in range(len(day_list)):
        col_letter = ws_psi.cell(row=1, column=3 + idx).column_letter
        ws_psi.column_dimensions[col_letter].width = 11

    # Serialize to memory and send
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'monthly_plan_{start_date.strftime("%Y%m")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/export_monthly_result')
def export_monthly_result():
    """
    Export the Monthly Result grid (Line × Qty/WH × calendar days) for the selected month.
    Query param: date=YYYY-MM-DD (any day in the month; defaults like monthly-result view).
    """
    if Workbook is None:
        flash('Excel export requires openpyxl to be installed (pip install openpyxl).', 'error')
        return redirect(url_for('monthly_result_view'))

    if 'date' in request.args:
        date_str = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    else:
        date_str = get_earliest_schedule_date() or datetime.now().strftime('%Y-%m-%d')
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()
    resource_filter = (request.args.get('machine', '') or '').strip()

    start_date = current_date.replace(day=1)
    if start_date.month == 12:
        end_date = start_date.replace(year=start_date.year + 1, month=1)
    else:
        end_date = start_date.replace(month=start_date.month + 1)

    day_list, line_items = _build_monthly_result_grid(start_date, end_date, resource_filter)
    day_labels = [_format_md_weekday_en(d) for d in day_list]

    wb = Workbook()
    ws = wb.active
    ws.title = f'{start_date.strftime("%Y-%m")}_MR'
    ws.sheet_view.zoomScale = 80
    ws.sheet_view.zoomScaleNormal = 80

    header_font = Font(name='Meiryo', bold=True)
    default_font = Font(name='Meiryo')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill('solid', fgColor='BCE597')
    fill_even = PatternFill('solid', fgColor='FFFFFF')
    fill_odd = PatternFill('solid', fgColor='DDF2CA')

    ws.append(['Resource', 'Type'] + day_labels)
    for col_idx in range(1, 3 + len(day_list)):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    row_idx = 2
    for group_i, item in enumerate(line_items):
        fill = fill_even if group_i % 2 == 0 else fill_odd
        if len(day_labels) > 0:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 1, end_column=1)
        c_line = ws.cell(row=row_idx, column=1)
        c_line.value = item['line_name']
        c_line.font = default_font
        c_line.border = thin_border
        c_line.alignment = Alignment(horizontal='left', vertical='center')
        c_line.fill = fill

        c_qty_lbl = ws.cell(row=row_idx, column=2)
        c_qty_lbl.value = 'Qty'
        c_qty_lbl.font = default_font
        c_qty_lbl.border = thin_border
        c_qty_lbl.alignment = center
        c_qty_lbl.fill = fill

        c_wh_lbl = ws.cell(row=row_idx + 1, column=2)
        c_wh_lbl.value = 'WH'
        c_wh_lbl.font = default_font
        c_wh_lbl.border = thin_border
        c_wh_lbl.alignment = center
        c_wh_lbl.fill = fill

        for di, qv in enumerate(item['qty_cells']):
            col = 3 + di
            cell = ws.cell(row=row_idx, column=col)
            cell.value = qv
            cell.font = default_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = fill

        for di, hv in enumerate(item['wh_cells']):
            col = 3 + di
            cell = ws.cell(row=row_idx + 1, column=col)
            cell.value = hv
            cell.font = default_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = fill

        row_idx += 2

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 10
    for idx in range(len(day_list)):
        col_letter = ws.cell(row=1, column=3 + idx).column_letter
        ws.column_dimensions[col_letter].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'monthly_result_{start_date.strftime("%Y%m")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/clear', methods=['POST'])
def clear_data():
    conn = get_db()
    conn.execute('DELETE FROM schedules')
    conn.execute('DELETE FROM uploads')
    conn.execute('DELETE FROM psi_input_instructions')
    conn.execute('DELETE FROM psi_input_instruction_uploads')
    conn.execute('DELETE FROM psi_output_instructions')
    conn.execute('DELETE FROM psi_output_instruction_uploads')
    conn.commit()
    conn.close()
    flash('All schedule/PSI data cleared', 'success')
    return redirect(url_for('index'))


@app.route('/sample_csv')
def sample_csv():
    """Generate a sample CSV for testing"""
    from flask import Response
    lines = [
        'OrderID,ItemID,ItemName,ResourceID,ResourceName,StartTime,EndTime,Quantity,Status,ProcessName',
        'ORD-001,PART-A,Widget Alpha,MC-01,CNC Machine 1,2024/03/15 08:00:00,2024/03/15 10:30:00,100,Scheduled,Milling',
        'ORD-002,PART-B,Gear Beta,MC-02,CNC Machine 2,2024/03/15 09:00:00,2024/03/15 11:00:00,50,In Progress,Turning',
        'ORD-003,PART-C,Frame Gamma,MC-01,CNC Machine 1,2024/03/15 11:00:00,2024/03/15 14:00:00,75,Scheduled,Drilling',
        'ORD-004,PART-A,Widget Alpha,MC-03,Assembly Line A,2024/03/15 14:00:00,2024/03/15 16:00:00,100,Scheduled,Assembly',
        'ORD-005,PART-D,Shaft Delta,MC-02,CNC Machine 2,2024/03/15 13:00:00,2024/03/15 17:00:00,200,Scheduled,Grinding',
        'ORD-006,PART-E,Housing Epsilon,MC-04,Lathe Machine 1,2024/03/15 08:30:00,2024/03/15 12:30:00,30,Completed,Lathing',
        'ORD-007,PART-B,Gear Beta,MC-03,Assembly Line A,2024/03/16 08:00:00,2024/03/16 10:00:00,50,Scheduled,Assembly',
        'ORD-008,PART-F,Bracket Zeta,MC-01,CNC Machine 1,2024/03/16 09:00:00,2024/03/16 13:00:00,120,Scheduled,Milling',
        'ORD-009,PART-G,Plate Eta,MC-04,Lathe Machine 1,2024/03/16 10:00:00,2024/03/16 14:00:00,45,Scheduled,Lathing',
        'ORD-010,PART-C,Frame Gamma,MC-02,CNC Machine 2,2024/03/16 14:00:00,2024/03/16 18:00:00,75,Scheduled,Drilling',
    ]
    csv_content = '\n'.join(lines)
    return Response(csv_content, mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=sample_schedule.csv'})


@app.route('/psi')
def psi_view():
    """
    Web PSI viewer.
    Shows, for a given month, per item_code from instruction uploads:
    - Supply: psi_output_instructions (cell lines keyed by operation_code)
    - Demand: psi_input_instructions (same)
    - Stock: running balance (Supply − Demand)
    """
    date_str = (
        request.args.get('date')
        or get_earliest_psi_instruction_date()
        or get_earliest_schedule_date()
        or datetime.now().strftime('%Y-%m-%d')
    )
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()

    start_date = current_date.replace(day=1)
    if start_date.month == 12:
        end_date = start_date.replace(year=start_date.year + 1, month=1)
    else:
        end_date = start_date.replace(month=start_date.month + 1)
    if start_date.month == 1:
        prev_start = start_date.replace(year=start_date.year - 1, month=12, day=1)
    else:
        prev_start = start_date.replace(month=start_date.month - 1, day=1)
    if start_date.month == 12:
        next_start = start_date.replace(year=start_date.year + 1, month=1, day=1)
    else:
        next_start = start_date.replace(month=start_date.month + 1, day=1)

    day_list, ordered_items, _, supply_qty, agg, demand_qty, demand_agg = _build_psi_data_for_month(
        start_date, end_date
    )
    beg_bal_by_item = _build_psi_begin_balances(start_date)
    day_labels = [_format_md_weekday_en(d) for d in day_list]

    # Build PSI rows for template
    psi_items = []
    for item_key in ordered_items:
        rows_for_item = []
        beg_bal = float(beg_bal_by_item.get(item_key, 0.0))
        stock_prev = beg_bal
        for row_type in ('Supply', 'Demand', 'Stock'):
            cells = []
            for d in day_list:
                day_date = d.date()
                if row_type == 'Supply':
                    parts = []
                    for (it, day, m), qty in agg.items():
                        if it == item_key and day == day_date and qty:
                            parts.append(f'{m}: {qty:g}')
                    cell_val = '\n'.join(parts) if parts else ''
                elif row_type == 'Demand':
                    parts = []
                    for (it, day, m), qty in demand_agg.items():
                        if it == item_key and day == day_date and qty and float(qty) > 0:
                            parts.append(f'{m}: {qty:g}')
                    cell_val = '\n'.join(parts) if parts else ''
                else:  # Stock
                    s = supply_qty.get((item_key, day_date), 0.0)
                    d_qty = demand_qty.get((item_key, day_date), 0.0)
                    stock = stock_prev + s - d_qty
                    stock_prev = stock
                    cell_val = f'{stock:g}'
                cells.append(cell_val)
            rows_for_item.append({'type': row_type, 'beg_bal': f'{beg_bal:g}' if row_type == 'Stock' else '', 'cells': cells})
        psi_items.append({'item': item_key, 'rows': rows_for_item})

    return render_template(
        'psi.html',
        day_labels=day_labels,
        psi_items=psi_items,
        month_label=start_date.strftime('%Y-%m'),
        current_date=current_date,
        psi_date=start_date.strftime('%Y-%m-%d'),
        start_date=start_date,
        prev_date=prev_start.strftime('%Y-%m-%d'),
        next_date=next_start.strftime('%Y-%m-%d'),
    )


def _parse_month_ym(ym: str) -> tuple[str, datetime, datetime]:
    raw = (ym or "").strip()
    month_start = datetime.strptime(raw + "-01", "%Y-%m-%d")
    if month_start.month == 12:
        month_end = datetime(month_start.year + 1, 1, 1)
    else:
        month_end = datetime(month_start.year, month_start.month + 1, 1)
    return raw, month_start, month_end


def _monthly_result_wh_counts_for_calendar_day(day: datetime) -> bool:
    """
    Monthly Result WH is split by calendar day. Weekends are treated as non-working:
    no WH is accumulated on Saturday/Sunday (avoids 24h per weekend day when ops span weeks).
    """
    return day.weekday() < 5


def _format_md_weekday_en(d: datetime) -> str:
    wd = ('Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun')[d.weekday()]
    return f"{d.month}/{d.day}({wd})"


def _build_resource_operation_detail_rows(
    start_date: datetime, end_date: datetime, resource_filter: str
) -> list[dict]:
    resource_filter = (resource_filter or "").strip()
    if not resource_filter:
        return []
    conn = get_db()
    try:
        rows = conn.execute(
            '''
            WITH base AS (
              SELECT
                TRIM(COALESCE(NULLIF(actual_resource, ''), NULLIF(machine_name, ''), 'Unknown')) AS resource_name,
                item_id,
                quantity,
                actual_quantity,
                status,
                CASE
                  WHEN actual_start IS NOT NULL AND actual_end IS NOT NULL THEN actual_start
                  ELSE start_time
                END AS start_ts,
                CASE
                  WHEN actual_start IS NOT NULL AND actual_end IS NOT NULL THEN actual_end
                  ELSE end_time
                END AS end_ts
              FROM schedules
              WHERE TRIM(COALESCE(item_id, '')) <> ''
                AND TRIM(COALESCE(machine_id, '')) NOT IN ('DefaultInventoryResource', 'Purchase')
            )
            SELECT item_id, quantity, actual_quantity, status, start_ts
            FROM base
            WHERE resource_name = ?
              AND start_ts IS NOT NULL
              AND end_ts IS NOT NULL
              AND start_ts < ?
              AND end_ts > ?
            ORDER BY start_ts, item_id
            ''',
            (
                resource_filter,
                end_date.strftime('%Y-%m-%d %H:%M:%S'),
                start_date.strftime('%Y-%m-%d %H:%M:%S'),
            ),
        ).fetchall()
    finally:
        conn.close()
    out: list[dict] = []
    for r in rows:
        st = datetime.strptime(str(r['start_ts']), '%Y-%m-%d %H:%M:%S')
        aq = r['actual_quantity']
        qty_disp = aq if aq is not None else r['quantity']
        planned_qty_disp = r['quantity'] if aq is not None else None
        out.append(
            {
                "start_md": f"{st.month}/{st.day}",
                "item_id": str(r['item_id'] or '').strip() or '—',
                "quantity": qty_disp,
                "status": str(r['status'] or '').strip() or '—',
                "planned_quantity": planned_qty_disp,
            }
        )
    return out


def _build_monthly_result_grid(
    start_date: datetime, end_date: datetime, resource_filter: str = ""
) -> tuple[list[datetime], list[dict]]:
    """
    Same aggregation as the Monthly Result page: per line, Qty on operation start day
    and WH (hours) spread across [start_ts, end_ts) clipped to the month.
    WH is not counted on Saturday/Sunday (non-working calendar days).
    """
    conn = get_db()
    try:
        sql = """
            WITH base AS (
              SELECT
                TRIM(COALESCE(NULLIF(actual_resource, ''), NULLIF(machine_name, ''), 'Unknown')) AS line_name,
                work_user_res_order,
                CASE
                  WHEN actual_quantity IS NOT NULL THEN actual_quantity
                  WHEN quantity IS NOT NULL THEN quantity
                  ELSE 0
                END AS qty_val,
                CASE
                  WHEN actual_start IS NOT NULL AND actual_end IS NOT NULL THEN actual_start
                  ELSE start_time
                END AS start_ts,
                CASE
                  WHEN actual_start IS NOT NULL AND actual_end IS NOT NULL THEN actual_end
                  ELSE end_time
                END AS end_ts
              FROM schedules
              WHERE TRIM(COALESCE(item_id, '')) <> ''
                AND TRIM(COALESCE(machine_id, '')) NOT IN ('DefaultInventoryResource', 'Purchase')
            )
            SELECT
              line_name,
              work_user_res_order,
              qty_val,
              start_ts,
              end_ts
            FROM base
            WHERE start_ts IS NOT NULL
              AND end_ts IS NOT NULL
              AND start_ts < ?
              AND end_ts > ?
              AND (? = '' OR line_name = ?)
            ORDER BY line_name, start_ts
        """
        resource_filter = (resource_filter or "").strip()
        params = (
            end_date.strftime('%Y-%m-%d %H:%M:%S'),
            start_date.strftime('%Y-%m-%d %H:%M:%S'),
            resource_filter,
            resource_filter,
        )
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()

    day_list: list[datetime] = []
    cur = start_date
    while cur < end_date:
        day_list.append(cur)
        cur += timedelta(days=1)
    day_keys = [d.strftime('%Y-%m-%d') for d in day_list]

    qty_by_line_day: dict[tuple[str, str], float] = defaultdict(float)
    wh_by_line_day: dict[tuple[str, str], float] = defaultdict(float)
    op_count_by_line: dict[str, int] = defaultdict(int)
    lines: set[str] = set()
    min_wro_by_line: dict[str, float] = {}
    for row in rows:
        line_name = str(row['line_name'] or '').strip() or 'Unknown'
        lines.add(line_name)
        op_count_by_line[line_name] += 1
        wro = _parse_work_user_res_order_val(row['work_user_res_order'])
        prev_min = min_wro_by_line.get(line_name)
        if prev_min is None or wro < prev_min:
            min_wro_by_line[line_name] = wro
        qty = float(row['qty_val'] or 0)
        start_ts = datetime.strptime(str(row['start_ts']), '%Y-%m-%d %H:%M:%S')
        end_ts = datetime.strptime(str(row['end_ts']), '%Y-%m-%d %H:%M:%S')
        day_key = start_ts.strftime('%Y-%m-%d')
        if day_key in day_keys:
            qty_by_line_day[(line_name, day_key)] += qty

        d0 = datetime(start_ts.year, start_ts.month, start_ts.day)
        if d0 < start_date:
            d0 = start_date
        while d0 < end_date and d0 < end_ts:
            d1 = d0 + timedelta(days=1)
            seg_s = max(start_ts, d0)
            seg_e = min(end_ts, d1, end_date)
            if seg_e > seg_s and _monthly_result_wh_counts_for_calendar_day(d0):
                wh_by_line_day[(line_name, d0.strftime('%Y-%m-%d'))] += (seg_e - seg_s).total_seconds() / 3600.0
            d0 = d1

    line_items: list[dict] = []
    lines_ordered = sorted(
        lines,
        key=lambda ln: (min_wro_by_line.get(ln, float('inf')), ln),
    )
    for line_name in lines_ordered:
        qty_cells = []
        wh_cells = []
        for k in day_keys:
            q = qty_by_line_day.get((line_name, k), 0.0)
            h = wh_by_line_day.get((line_name, k), 0.0)
            qty_cells.append(
                ""
                if q == 0
                else (f"{q:,.0f}" if abs(q - round(q)) < 1e-9 else f"{q:,.3f}".rstrip('0').rstrip('.'))
            )
            wh_cells.append("" if h == 0 else f"{h:,.2f}".rstrip('0').rstrip('.'))
        line_items.append(
            {
                "line_name": line_name,
                "op_count": op_count_by_line.get(line_name, 0),
                "qty_cells": qty_cells,
                "wh_cells": wh_cells,
            }
        )

    return day_list, line_items


def _build_monthly_result_resource_detail_rows(
    start_date: datetime, end_date: datetime, resource_filter: str
) -> list[dict]:
    out = _build_resource_operation_detail_rows(start_date, end_date, resource_filter)
    for i, row in enumerate(out):
        if i > 0 and out[i - 1]["start_md"] == row["start_md"]:
            row["start_rowspan"] = 0
        else:
            span = 1
            while i + span < len(out) and out[i + span]["start_md"] == row["start_md"]:
                span += 1
            row["start_rowspan"] = span

        if (
            i > 0
            and out[i - 1]["start_md"] == row["start_md"]
            and out[i - 1]["item_id"] == row["item_id"]
        ):
            row["item_rowspan"] = 0
        else:
            span = 1
            while (
                i + span < len(out)
                and out[i + span]["start_md"] == row["start_md"]
                and out[i + span]["item_id"] == row["item_id"]
            ):
                span += 1
            row["item_rowspan"] = span
    return out


@app.route('/export_gantt_resource_detail')
def export_gantt_resource_detail():
    if Workbook is None:
        flash('Excel export requires openpyxl to be installed (pip install openpyxl).', 'error')
        return redirect(url_for('gantt'))
    view = request.args.get('view', 'monthly')
    machine_filter = (request.args.get('machine', '') or '').strip()
    date_str = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()
    if view in ('weekly', '2week', '3week'):
        start_date = current_date - timedelta(days=current_date.weekday())
        span_days = {'weekly': 7, '2week': 14, '3week': 21}[view]
        end_date = start_date + timedelta(days=span_days)
    elif view == 'monthly':
        start_date = current_date.replace(day=1)
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1)
    else:
        start_date = current_date
        end_date = current_date + timedelta(days=1)
    rows = _build_resource_operation_detail_rows(start_date, end_date, machine_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resource_Detail'
    header_font = Font(name='Meiryo', bold=True)
    default_font = Font(name='Meiryo')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill('solid', fgColor='76E597')

    ws.append(['start (M/D)', 'item_id', 'quantity', 'status', 'planned quantity'])
    for col_idx in range(1, 6):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = center

    for r in rows:
        ws.append(
            [
                r['start_md'],
                r['item_id'],
                r['quantity'] if r['quantity'] is not None else '—',
                r['status'],
                r['planned_quantity'] if r['planned_quantity'] is not None else '—',
            ]
        )
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 6):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = default_font
            c.border = thin_border
            c.alignment = center
    for i, w in enumerate((14, 20, 14, 14, 18), start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_resource = ''.join(c if c.isalnum() else '_' for c in (machine_filter or 'resource'))
    filename = f'gantt_resource_detail_{safe_resource}_{start_date.strftime("%Y%m%d")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/export_monthly_result_resource_detail')
def export_monthly_result_resource_detail():
    if Workbook is None:
        flash('Excel export requires openpyxl to be installed (pip install openpyxl).', 'error')
        return redirect(url_for('monthly_result_view'))
    date_str = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    machine_filter = (request.args.get('machine', '') or '').strip()
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()
    start_date = current_date.replace(day=1)
    if start_date.month == 12:
        end_date = start_date.replace(year=start_date.year + 1, month=1)
    else:
        end_date = start_date.replace(month=start_date.month + 1)
    rows = _build_monthly_result_resource_detail_rows(start_date, end_date, machine_filter)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resource_Detail'
    header_font = Font(name='Meiryo', bold=True)
    default_font = Font(name='Meiryo')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill('solid', fgColor='76E597')

    ws.append(['start (M/D)', 'item_id', 'quantity', 'status', 'planned quantity'])
    for col_idx in range(1, 6):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = center

    for r in rows:
        ws.append(
            [
                r['start_md'],
                r['item_id'],
                r['quantity'] if r['quantity'] is not None else '—',
                r['status'],
                r['planned_quantity'] if r['planned_quantity'] is not None else '—',
            ]
        )
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 6):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = default_font
            c.border = thin_border
            c.alignment = center
    for i, w in enumerate((14, 20, 14, 14, 18), start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_resource = ''.join(c if c.isalnum() else '_' for c in (machine_filter or 'resource'))
    filename = f'monthly_resource_detail_{safe_resource}_{start_date.strftime("%Y%m")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/monthly-result')
def monthly_result_view():
    view = 'monthly'
    resource_filter = (request.args.get('machine', '') or '').strip()
    if 'date' in request.args:
        date_str = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    else:
        date_str = get_earliest_schedule_date() or datetime.now().strftime('%Y-%m-%d')
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()
    start_date = current_date.replace(day=1)
    if start_date.month == 12:
        end_date = start_date.replace(year=start_date.year + 1, month=1)
    else:
        end_date = start_date.replace(month=start_date.month + 1)
    if start_date.month == 1:
        prev_start = start_date.replace(year=start_date.year - 1, month=12, day=1)
    else:
        prev_start = start_date.replace(month=start_date.month - 1, day=1)
    if start_date.month == 12:
        next_start = start_date.replace(year=start_date.year + 1, month=1, day=1)
    else:
        next_start = start_date.replace(month=start_date.month + 1, day=1)
    prev_date = prev_start.strftime('%Y-%m-%d')
    next_date = next_start.strftime('%Y-%m-%d')

    day_list, line_items = _build_monthly_result_grid(start_date, end_date, resource_filter)
    detail_rows = _build_monthly_result_resource_detail_rows(start_date, end_date, resource_filter)
    conn = get_db()
    try:
        resources_raw = conn.execute(
            '''
            SELECT DISTINCT TRIM(COALESCE(NULLIF(actual_resource, ''), NULLIF(machine_name, ''), 'Unknown')) AS resource_name
            FROM schedules
            WHERE TRIM(COALESCE(item_id, '')) <> ''
              AND TRIM(COALESCE(machine_id, '')) NOT IN ('DefaultInventoryResource', 'Purchase')
            ORDER BY resource_name
            '''
        ).fetchall()
    finally:
        conn.close()
    resources = [str(r['resource_name'] or '').strip() for r in resources_raw if str(r['resource_name'] or '').strip()]

    return render_template(
        'monthly_result.html',
        view=view,
        current_date=current_date,
        start_date=start_date,
        end_date=end_date,
        prev_date=prev_date,
        next_date=next_date,
        day_labels=[_format_md_weekday_en(d) for d in day_list],
        line_items=line_items,
        detail_rows=detail_rows,
        resources=resources,
        machine_filter=resource_filter,
    )


def _build_psi_data_for_month(start_date, end_date):
    """
    Build PSI structures for a month from instruction uploads.

    Supply: psi_output_instructions aggregated by (item_code, calendar day,
    machine_name). Demand: psi_input_instructions the same way.
    The machine_name is resolved from schedules by operation_code.
    """
    start_s = start_date.strftime('%Y-%m-%d')
    end_s = end_date.strftime('%Y-%m-%d')

    conn = get_db()
    out_rows = conn.execute(
        '''
        SELECT item_code, inst_time, quantity, operation_code
        FROM psi_output_instructions
        WHERE inst_time >= ? AND inst_time < ?
        ''',
        (start_s, end_s),
    ).fetchall()
    in_rows = conn.execute(
        '''
        SELECT item_code, inst_time, quantity, operation_code
        FROM psi_input_instructions
        WHERE inst_time >= ? AND inst_time < ?
        ''',
        (start_s, end_s),
    ).fetchall()

    sched_rows = conn.execute(
        '''
        SELECT
            operation_code,
            COALESCE(NULLIF(TRIM(machine_name), ''), NULLIF(TRIM(actual_resource), ''), 'Unknown') AS machine_name
        FROM schedules
        WHERE start_time >= ? AND start_time < ?
          AND operation_code IS NOT NULL
          AND TRIM(operation_code) <> ''
        ''',
        (start_s, end_s),
    ).fetchall()
    conn.close()

    day_list = []
    d = start_date
    while d < end_date:
        day_list.append(d)
        d += timedelta(days=1)

    def _day_from_inst_time(inst_time):
        if not inst_time:
            return None
        s = str(inst_time).strip()
        if len(s) < 8:
            return None
        for fmt in ('%Y-%m-%d', '%Y/%m/%d'):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                continue
        return None

    op_to_machine_counts = defaultdict(lambda: defaultdict(int))
    for sr in sched_rows:
        op_code = str(sr['operation_code'] or '').strip()
        machine_name = str(sr['machine_name'] or '').strip() or 'Unknown'
        if not op_code:
            continue
        op_to_machine_counts[op_code][machine_name] += 1

    def _machine_dim(op_raw) -> str:
        op_code = str(op_raw or '').strip()
        if not op_code:
            return 'Unknown'
        counts = op_to_machine_counts.get(op_code)
        if counts:
            return max(counts.items(), key=lambda kv: kv[1])[0]
        return op_code

    items: set[str] = set()
    agg = defaultdict(float)
    for r in out_rows:
        item_key = (r['item_code'] or '').strip()
        if not item_key:
            continue
        items.add(item_key)
        day = _day_from_inst_time(r['inst_time'])
        if day is None:
            continue
        qty = float(r['quantity'] or 0)
        if item_key.upper() == 'CUSTOMER' and qty > 0:
            qty = -abs(qty)
        machine = _machine_dim(r['operation_code'])
        agg[(item_key, day, machine)] += qty

    demand_agg = defaultdict(float)
    for r in in_rows:
        item_key = (r['item_code'] or '').strip()
        if not item_key:
            continue
        items.add(item_key)
        day = _day_from_inst_time(r['inst_time'])
        if day is None:
            continue
        qty = float(r['quantity'] or 0)
        if qty <= 0:
            continue
        machine = _machine_dim(r['operation_code'])
        demand_agg[(item_key, day, machine)] += qty

    supply_qty = defaultdict(float)
    for (item_key, day, _op), qty in agg.items():
        if qty:
            supply_qty[(item_key, day)] += qty

    demand_qty = defaultdict(float)
    for (item_key, day, _op), qty in demand_agg.items():
        if qty:
            demand_qty[(item_key, day)] += qty

    def split_item_code(code: str):
        if not code:
            return None, None
        if '-' in code:
            base, suffix = code.rsplit('-', 1)
            try:
                num = int(suffix)
                return base, num
            except ValueError:
                return code, None
        return code, None

    items_sorted = sorted(items)
    families = {}
    for it in items_sorted:
        base, num = split_item_code(it)
        if base is None:
            continue
        families.setdefault(base, []).append((it, num))

    next_item_for: dict[str, str] = {}
    ordered_items = []
    for base in sorted(families.keys()):
        variants = families[base]
        numbered = [v for v in variants if v[1] is not None]
        base_codes = [v for v in variants if v[1] is None]
        numbered.sort(key=lambda x: x[1] if x[1] is not None else -1)
        base_code = base if base in items else (base_codes[0][0] if base_codes else None)
        family_order = [code for code, _ in numbered]
        if base_code and base_code not in family_order:
            family_order.append(base_code)
        ordered_items.extend(family_order)
        for code, num in numbered:
            if num is None:
                continue
            if code in next_item_for:
                continue
            candidate = f"{base}-{num + 10}"
            if candidate in items:
                next_item_for[code] = candidate
            elif base_code and base_code in items:
                next_item_for[code] = base_code
    remaining = [it for it in items_sorted if it not in ordered_items]
    ordered_items.extend(sorted(remaining))

    return day_list, ordered_items, next_item_for, supply_qty, agg, demand_qty, demand_agg


def _build_psi_begin_balances(month_start: datetime) -> dict[str, float]:
    """
    Beg Bal per item at month start from DefaultInventoryResource.
    Use the latest row on/before month_start by effective start timestamp.
    """
    start_s = month_start.strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    try:
        rows = conn.execute(
            '''
            WITH inv AS (
              SELECT
                TRIM(COALESCE(item_id, '')) AS item_id,
                CASE
                  WHEN actual_start IS NOT NULL AND actual_end IS NOT NULL THEN actual_start
                  ELSE start_time
                END AS ts,
                CASE
                  WHEN actual_quantity IS NOT NULL THEN actual_quantity
                  WHEN quantity IS NOT NULL THEN quantity
                  ELSE 0
                END AS qty
              FROM schedules
              WHERE TRIM(COALESCE(machine_id, '')) = 'DefaultInventoryResource'
                AND TRIM(COALESCE(item_id, '')) <> ''
            ),
            ranked AS (
              SELECT
                item_id,
                ts,
                qty,
                ROW_NUMBER() OVER (PARTITION BY item_id ORDER BY ts DESC) AS rn
              FROM inv
              WHERE ts IS NOT NULL
                AND ts <= ?
            )
            SELECT item_id, qty FROM ranked WHERE rn = 1
            ''',
            (start_s,),
        ).fetchall()
    finally:
        conn.close()
    out: dict[str, float] = {}
    for r in rows:
        item = str(r['item_id'] or '').strip()
        if not item:
            continue
        out[item] = float(r['qty'] or 0)
    return out


@app.route('/export_psi')
def export_psi():
    """
    Export only the PSI table (Supply / Demand / Stock) for the current month to Excel.
    Query param: date=YYYY-MM-DD (defaults to earliest instruction or schedule date, else today).
    """
    if Workbook is None:
        flash('Excel export requires openpyxl to be installed (pip install openpyxl).', 'error')
        return redirect(url_for('psi_view'))

    date_str = (
        request.args.get('date')
        or get_earliest_psi_instruction_date()
        or get_earliest_schedule_date()
        or datetime.now().strftime('%Y-%m-%d')
    )
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        current_date = datetime.now()

    start_date = current_date.replace(day=1)
    if start_date.month == 12:
        end_date = start_date.replace(year=start_date.year + 1, month=1)
    else:
        end_date = start_date.replace(month=start_date.month + 1)

    day_list, ordered_items, _, supply_qty, agg, demand_qty, demand_agg = _build_psi_data_for_month(
        start_date, end_date
    )
    beg_bal_by_item = _build_psi_begin_balances(start_date)
    day_labels = [_format_md_weekday_en(d) for d in day_list]

    wb = Workbook()
    ws_psi = wb.active
    ws_psi.title = f'{start_date.strftime("%Y-%m")}_PSI'
    ws_psi.sheet_view.zoomScale = 80
    ws_psi.sheet_view.zoomScaleNormal = 80

    header_font = Font(name='Meiryo', bold=True)
    default_font = Font(name='Meiryo')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')
    psi_header_fill = PatternFill('solid', fgColor='BCE597')
    psi_fill_even = PatternFill('solid', fgColor='FFFFFF')
    psi_fill_odd = PatternFill('solid', fgColor='DDF2CA')

    ws_psi.append(['Item', 'Type', 'Beg Bal'] + day_labels)
    for col_idx in range(1, 4 + len(day_list)):
        cell = ws_psi.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = psi_header_fill
        cell.border = thin_border
        cell.alignment = center

    for item_key in ordered_items:
        beg_bal = float(beg_bal_by_item.get(item_key, 0.0))
        stock_prev = beg_bal
        for row_type in ('Supply', 'Demand', 'Stock'):
            beg_bal_cell = beg_bal if row_type == 'Stock' else ''
            row_vals = [item_key, row_type, beg_bal_cell]
            for d in day_list:
                day_date = d.date()
                if row_type == 'Supply':
                    parts = []
                    for (it, day, m), qty in agg.items():
                        if it == item_key and day == day_date and qty:
                            parts.append(f'{m}: {qty:g}')
                    cell_val = '\n'.join(parts) if parts else ''
                elif row_type == 'Demand':
                    parts = []
                    for (it, day, m), qty in demand_agg.items():
                        if it == item_key and day == day_date and qty and float(qty) > 0:
                            parts.append(f'{m}: {qty:g}')
                    cell_val = '\n'.join(parts) if parts else ''
                else:
                    s = supply_qty.get((item_key, day_date), 0.0)
                    d_qty = demand_qty.get((item_key, day_date), 0.0)
                    stock = stock_prev + s - d_qty
                    stock_prev = stock
                    cell_val = stock
                row_vals.append(cell_val)
            ws_psi.append(row_vals)

    for r_idx, row in enumerate(
        ws_psi.iter_rows(min_row=2, max_row=ws_psi.max_row, min_col=1, max_col=3 + len(day_list)),
        start=2
    ):
        item_group = (r_idx - 2) // 3
        fill = psi_fill_even if (item_group % 2 == 0) else psi_fill_odd
        for cell in row:
            cell.border = thin_border
            cell.fill = fill
            if cell.col_idx <= 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = default_font

    ws_psi.column_dimensions['A'].width = 22
    ws_psi.column_dimensions['B'].width = 10
    ws_psi.column_dimensions['C'].width = 10
    for idx in range(len(day_list)):
        col_letter = ws_psi.cell(row=1, column=4 + idx).column_letter
        ws_psi.column_dimensions[col_letter].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'PSI_{start_date.strftime("%Y%m")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    init_db()
    ensure_db_schema()
    _viewer_app_py = Path(__file__).resolve()
    print()
    print('=== ASPROVA Viewer (SQLite / Gantt) ===')
    print(f'  使用中の app.py: {_viewer_app_py}')
    print('  動作確認: http://127.0.0.1:5000/viewer-check （または /__asprova_viewer_check）')
    print('  （別パスが出る場合は、今いるフォルダ違いの app.py を起動しています）')
    print('========================================')
    print()
    app.run(debug=True, host='0.0.0.0', port=5000)
