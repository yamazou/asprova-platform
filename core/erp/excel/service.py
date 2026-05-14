"""Excel ベース ERP 用 Bridge サービス実装。

このサービスは「顧客プロファイルから設定された Excel/CSV をデコードし、
顧客 Strategy が用意していれば専用ローダを使い、なければ汎用 ``dict_rows``
→ tuple 変換にフォールバック」する。openpyxl への依存はこのサブパッケージに
集約しておくことで、Excel 顧客を含めない納品では openpyxl 自体を切ることもできる。
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any, Optional

from core.customers.base import CustomerStrategy
from core.erp._base import BridgeErpService


# ---------------------------------------------------------------------------
# Customer profile キーと既定ファイル名
# ---------------------------------------------------------------------------


def _excel_file_key(kind: str) -> str:
    return {
        "integrated": "excel_integrated_file",
        "item": "excel_item_file",
        "order": "excel_order_file",
        "prd_plan": "excel_prd_plan_file",
        "resource": "excel_resource_file",
        "inventory": "excel_inventory_file",
        "inventory_wip": "excel_inventory_wip_file",
    }[kind]


def _default_excel_filename(kind: str) -> str:
    return {
        "integrated": "integrated_master.xlsx",
        "item": "item_table.xlsx",
        "order": "order_table.xlsx",
        "prd_plan": "prd_plan_table.xlsx",
        "resource": "resource_table.xlsx",
        "inventory": "inventory_table.xlsx",
        "inventory_wip": "inventory_wip_table.xlsx",
    }[kind]


# ---------------------------------------------------------------------------
# Headers (mcframe 側と整合)
# ---------------------------------------------------------------------------


_INTEGRATED_HEADERS = [
    "P_ITM_CD",
    "PROCESS_NO",
    "PROCESS_CD",
    "INST_TYP",
    "INST_CD",
    "ITM_RESOURCE",
    "PRODUCTION",
]
_ITEM_TABLE_HEADERS = ["ITM_CD", "ITM_NM", "ITM_TYP", "MAX_LOT_UNIT_QTY"]
_ORDER_TABLE_HEADERS = ["REQ_NO", "ITM_CD", "DLV_DT", "REQ_QTY", "CUST_CD"]
_RESOURCE_TABLE_HEADERS = ["LINE_CD", "LINE_NM", "RESOURCE_GRP", "Sort_Order"]
_INVENTORY_TABLE_HEADERS = ["INV_CD", "ITM_CD", "STK_QTY", "INV_DT"]


# ---------------------------------------------------------------------------
# Generic CSV / XLSX decoders (顧客に依存しない)
# ---------------------------------------------------------------------------


def _iter_dict_rows_from_csv_bytes(raw: bytes) -> list[dict[str, str]]:
    errors: list[Exception] = []
    for enc in ("utf-8-sig", "cp932", "shift_jis"):
        try:
            text = raw.decode(enc)
            reader = csv.DictReader(io.StringIO(text))
            return [dict(r) for r in reader]
        except UnicodeDecodeError as exc:
            errors.append(exc)
            continue
    raise RuntimeError("Could not determine the CSV character encoding.") from (
        errors[-1] if errors else None
    )


def _iter_dict_rows_from_xlsx_bytes(raw: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        out: list[dict[str, str]] = []
        for values in rows:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = values[i] if i < len(values) else ""
                row[h] = "" if v is None else str(v).strip()
            out.append(row)
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Service implementation
# ---------------------------------------------------------------------------


class ExcelBridgeService(BridgeErpService):
    """Excel ベース ERP 用 Bridge サービス。

    Args:
        customer: 顧客 Strategy。``load_excel_export_rows(kind, raw)`` を実装
                  していればそちらを優先して使う (PEB 等)。
        profile:  ``BRIDGE_CUSTOMERS[customer_id]`` の dict。``excel_base_dir``
                  ``excel_*_file`` を参照する。
    """

    def __init__(
        self,
        *,
        customer: CustomerStrategy,
        profile: dict[str, str],
    ) -> None:
        if not isinstance(profile, dict):
            raise RuntimeError("Please select a Customer for Excel import.")
        if not str(profile.get("excel_base_dir") or "").strip():
            raise RuntimeError("excel_base_dir is not configured for the selected Customer.")
        self._customer = customer
        self._profile = profile

    # -- internal helpers -----------------------------------------------------

    def _resolve_source_path(self, kind: str) -> Path:
        base_dir = str(self._profile.get("excel_base_dir") or "").strip()
        file_key = _excel_file_key(kind)
        file_name = str(
            self._profile.get(file_key) or _default_excel_filename(kind)
        ).strip()
        if not file_name:
            raise RuntimeError(f"{file_key} is not configured for the selected Customer.")
        return Path(base_dir) / file_name

    def _load_rows(
        self, kind: str, headers: list[str], upload: Any = None
    ) -> list[tuple]:
        file_name = ""
        raw = b""
        if upload and getattr(upload, "filename", ""):
            file_name = str(upload.filename or "").strip()
            raw = upload.read()
            if not raw:
                raise RuntimeError("The selected source file is empty.")
        else:
            path = self._resolve_source_path(kind)
            if not path.exists():
                raise RuntimeError(
                    f"Excel/CSV file not found: {path}"
                )
            file_name = path.name
            raw = path.read_bytes()
        suffix = Path(file_name).suffix.lower()
        if suffix == ".csv":
            dict_rows = _iter_dict_rows_from_csv_bytes(raw)
        elif suffix in (".xlsx", ".xlsm"):
            # 顧客固有の Excel ローダがあればそれを使い、無ければ汎用処理にフォールバック。
            custom_rows = self._customer.load_excel_export_rows(kind, raw)
            if custom_rows is not None:
                return custom_rows
            dict_rows = _iter_dict_rows_from_xlsx_bytes(raw)
        else:
            raise RuntimeError(
                f"Unsupported file extension: {suffix} (.csv/.xlsx/.xlsm only)."
            )
        return [
            tuple(str(row.get(h) or "").strip() for h in headers)
            for row in dict_rows
        ]

    # -- BridgeErpService implementation -------------------------------------

    def fetch_integrated_records(
        self, *, upload: Any = None
    ) -> list[dict[str, Any]]:
        rows = self._load_rows("integrated", _INTEGRATED_HEADERS, upload)
        return [dict(zip(_INTEGRATED_HEADERS, row)) for row in rows]

    def fetch_item_rows(self, *, upload: Any = None) -> list[tuple]:
        return self._load_rows("item", _ITEM_TABLE_HEADERS, upload)

    def fetch_order_rows(self, *, upload: Any = None) -> list[tuple]:
        return self._load_rows("order", _ORDER_TABLE_HEADERS, upload)

    def fetch_prd_plan_rows(self, *, upload: Any = None) -> list[tuple]:
        return self._load_rows("prd_plan", _ORDER_TABLE_HEADERS, upload)

    def fetch_resource_rows(
        self,
        *,
        upload: Any = None,
        sort_order_map: Optional[dict[str, int]] = None,
    ) -> list[tuple]:
        return self._load_rows("resource", _RESOURCE_TABLE_HEADERS, upload)

    def fetch_inventory_rows(self, *, upload: Any = None) -> list[tuple]:
        return self._load_rows("inventory", _INVENTORY_TABLE_HEADERS, upload)

    def fetch_inventory_wip_rows(self, *, upload: Any = None) -> list[tuple]:
        return self._load_rows("inventory_wip", _INVENTORY_TABLE_HEADERS, upload)
