"""PEB Integrated Master (Asprova IMaster_* CSV) を master data.xlsx から生成する。"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from typing import Any

from core.integrated_master import format_integrated_i_production

PEB_IMASTER_HEADERS: tuple[str, ...] = (
    "IMaster_FinalItemCode",
    "IMaster_ProcNo",
    "IMaster_ProcCode",
    "IMaster_InstructionType",
    "IMaster_InstructionCode",
    "IMaster_ItemCodeOrResourceCode",
    "IMaster_Task2Expr",
    "IMaster_TimeConstraintMin",
)

_SHEET_MASTER_CODE = "master code"
_SHEET_MAIN_LINE_BY_MODEL = "Master main line by model"
_SHEET_INFUSION_BY_INK = "master infusion line by Ink"

_PROC_MAIN_LINE = ("30", "Main Line")
_PROC_QA_OUTPUT = ("40", "QA Output")
_PROC_INFUSION = ("20", "Infusion")
_QA_OUTPUT_TASK2_EXPR = "21H"
_INST_U = "U"
_INST_I = "I"
_INST_CD_U = "M"


def _normalize_header_name(raw: object) -> str:
    text = "" if raw is None else str(raw).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _normalize_item_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return str(int(float(s)))
    return s


def _format_pack_type(value: object) -> str:
    if value is None or str(value).strip() == "":
        return ""
    text = format_integrated_i_production(value)
    if not text:
        return ""
    try:
        d = Decimal(text.replace(",", ""))
    except InvalidOperation:
        return text
    if d == d.to_integral_value():
        return f"{int(d)}.00"
    return text


def _format_infusion_cycle_task2(value: object) -> str:
    """``master code`` の Cycle time infusion → ``257PH`` 等形式。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        return f"{int(round(float(value)))}PH"
    s = str(value).strip().replace(",", "")
    if not s:
        return ""
    if re.fullmatch(r"\d+PH", s, re.IGNORECASE):
        return f"{int(s[:-2])}PH"
    m = re.match(r"^(\d+(?:\.\d+)?)", s)
    if m:
        return f"{int(round(float(m.group(1))))}PH"
    return s


def _format_ink_cycle_task2(value: object) -> str:
    """Ink シート ``Cycle time (per hour)`` → ``30PH``（``30`` / ``30sp`` など）。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    s = ""
    if isinstance(value, int):
        s = str(value)
    elif isinstance(value, float):
        s = str(int(value)) if value == int(value) else str(value).strip()
    else:
        s = str(value).strip().replace(",", "")
    if not s:
        return ""
    if re.fullmatch(r"\d+PH", s, re.IGNORECASE):
        return f"{int(s[:-2])}PH"
    m = re.match(r"^(\d+)", s)
    if m:
        return f"{int(m.group(1))}PH"
    return s


def _format_time_constraint_min(value: object) -> str:
    """``master code`` の IMaster_TimeConstraintMin（例: ``3D``, ``4D``）。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        if value == int(value):
            return f"{int(value)}D"
        return str(value).strip()
    s = str(value).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return f"{int(float(s))}D"
    return s


def _format_cycle_time_task2(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    d = None
    if isinstance(value, (int, float)):
        d = Decimal(str(value))
    else:
        try:
            d = Decimal(s.replace(",", ""))
        except InvalidOperation:
            return s
    if d is not None and d == d.to_integral_value():
        n = int(d)
        if re.fullmatch(r"\d+PH", s, re.IGNORECASE):
            return s.upper() if s.endswith("ph") else s
        return f"{n}PH"
    return s


def _header_index_map(header_row: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, v in enumerate(header_row):
        key = _normalize_header_name(v)
        if key:
            out[key] = i
    return out


def _cell(row: tuple, idx_map: dict[str, int], *names: str) -> Any:
    for name in names:
        key = _normalize_header_name(name)
        i = idx_map.get(key)
        if i is None or i >= len(row):
            continue
        return row[i]
    return None


def _find_header_row(rows: list[tuple], required: tuple[str, ...]) -> int:
    req = {_normalize_header_name(n) for n in required}
    for i, row in enumerate(rows[:40]):
        keys = {_normalize_header_name(v) for v in row if v is not None}
        if req.issubset(keys):
            return i
    return 0


def _load_model_mainline_map(rows: list[tuple]) -> dict[str, tuple[str, str]]:
    """Model → (Mainline, Cycle Time)。"""
    if not rows:
        return {}
    header_idx = _find_header_row(rows, ("Model", "Mainline", "Cycle Time"))
    header_row = rows[header_idx]
    idx = _header_index_map(header_row)
    out: dict[str, tuple[str, str]] = {}
    for row in rows[header_idx + 1 :]:
        model = _normalize_item_code(_cell(row, idx, "Model"))
        if not model:
            continue
        mainline = str(_cell(row, idx, "Mainline", "Main line") or "").strip()
        cycle = _cell(row, idx, "Cycle Time", "Cycle time")
        out[model] = (mainline, _format_cycle_time_task2(cycle))
    return out


def _load_ink_line_map(rows: list[tuple]) -> dict[str, list[str]]:
    """Ink Code → ラインコード一覧（master infusion line by Ink）。"""
    if not rows:
        return {}
    header_idx = _find_header_row(rows, ("Ink Code",))
    header_row = rows[header_idx]
    idx = _header_index_map(header_row)
    out: dict[str, list[str]] = {}
    ink_col = idx.get("inkcode")
    desc_col = idx.get("inkdescription") or idx.get("description")
    start_col = 2
    if desc_col is not None:
        start_col = int(desc_col) + 1
    elif ink_col is not None:
        start_col = int(ink_col) + 2
    for row in rows[header_idx + 1 :]:
        ink = _normalize_item_code(_cell(row, idx, "Ink Code", "Ink code"))
        if not ink:
            continue
        lines: list[str] = []
        for i in range(start_col, len(row)):
            line = str(row[i] or "").strip()
            if line:
                lines.append(line)
        if lines:
            out[ink] = lines
    return out


def _qa_output_resource_code(mainline: str) -> str:
    """Main line の Mainline コードに `` QA Output`` を付与（例: ``MD`` → ``MD QA Output``）。"""
    code = str(mainline or "").strip()
    if not code:
        return ""
    return f"{code} QA Output"


def _imaster_row(
    *,
    final_item: str,
    proc_no: str,
    proc_code: str,
    inst_type: str,
    inst_code: str,
    item_or_resource: str,
    task2: str,
    time_constraint_min: str = "",
) -> dict[str, str]:
    return {
        "IMaster_FinalItemCode": final_item,
        "IMaster_ProcNo": proc_no,
        "IMaster_ProcCode": proc_code,
        "IMaster_InstructionType": inst_type,
        "IMaster_InstructionCode": inst_code,
        "IMaster_ItemCodeOrResourceCode": item_or_resource,
        "IMaster_Task2Expr": task2,
        "IMaster_TimeConstraintMin": (
            time_constraint_min if proc_code == _PROC_QA_OUTPUT[1] else ""
        ),
    }


def build_peb_integrated_master_records(raw: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc

    wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    try:
        if _SHEET_MASTER_CODE not in wb.sheetnames:
            raise RuntimeError(
                f'PEB Integrated Master requires a "{_SHEET_MASTER_CODE}" worksheet.'
            )
        if _SHEET_MAIN_LINE_BY_MODEL not in wb.sheetnames:
            raise RuntimeError(
                f'PEB Integrated Master requires a "{_SHEET_MAIN_LINE_BY_MODEL}" worksheet.'
            )

        ws_ml = wb[_SHEET_MAIN_LINE_BY_MODEL]
        model_rows = [tuple(r) for r in ws_ml.iter_rows(values_only=True)]
        model_map = _load_model_mainline_map(model_rows)

        ink_map: dict[str, list[str]] = {}
        if _SHEET_INFUSION_BY_INK in wb.sheetnames:
            ws_ink = wb[_SHEET_INFUSION_BY_INK]
            ink_rows = [tuple(r) for r in ws_ink.iter_rows(values_only=True)]
            ink_map = _load_ink_line_map(ink_rows)

        ws_mc = wb[_SHEET_MASTER_CODE]
        mc_rows = [tuple(r) for r in ws_mc.iter_rows(values_only=True)]
        if not mc_rows:
            return []

        header_idx = _find_header_row(
            mc_rows, ("FGcode", "Model", "process", "Infusion code", "PackType")
        )
        header_row = mc_rows[header_idx]
        idx = _header_index_map(header_row)
        data_rows = mc_rows[header_idx + 1 :]

        by_fg: dict[str, list[tuple]] = defaultdict(list)
        for row in data_rows:
            process = str(_cell(row, idx, "process") or "").strip().lower()
            if process != "main line":
                continue
            fg = _normalize_item_code(_cell(row, idx, "FGcode", "FG Code"))
            if not fg:
                continue
            by_fg[fg].append(row)

        records: list[dict[str, str]] = []
        proc_no, proc_code = _PROC_MAIN_LINE
        qa_proc_no, qa_proc_code = _PROC_QA_OUTPUT

        for fg in sorted(by_fg.keys()):
            items = by_fg[fg]
            for i, row in enumerate(items, start=1):
                infusion = _normalize_item_code(
                    _cell(row, idx, "Infusion code", "Infusion Code")
                )
                if not infusion:
                    continue
                pack_type = _format_pack_type(_cell(row, idx, "PackType", "Pack Type"))
                records.append(
                    _imaster_row(
                        final_item=fg,
                        proc_no=proc_no,
                        proc_code=proc_code,
                        inst_type=_INST_I,
                        inst_code=f"In{i}",
                        item_or_resource=infusion,
                        task2=pack_type,
                    )
                )

            seen_u: set[tuple[str, str]] = set()
            for row in items:
                model = _normalize_item_code(_cell(row, idx, "Model"))
                if not model:
                    continue
                lookup = model_map.get(model)
                if not lookup:
                    continue
                mainline, cycle_task2 = lookup
                if not mainline:
                    continue
                key = (mainline, cycle_task2)
                if key in seen_u:
                    continue
                seen_u.add(key)
                records.append(
                    _imaster_row(
                        final_item=fg,
                        proc_no=proc_no,
                        proc_code=proc_code,
                        inst_type=_INST_U,
                        inst_code=_INST_CD_U,
                        item_or_resource=mainline,
                        task2=cycle_task2,
                    )
                )
                qa_resource = _qa_output_resource_code(mainline)
                if qa_resource:
                    qa_time_constraint = _format_time_constraint_min(
                        _cell(
                            row,
                            idx,
                            "IMaster_TimeConstraintMin",
                            "IMaster TimeConstraintMin",
                            "Time Constraint Min",
                        )
                    )
                    records.append(
                        _imaster_row(
                            final_item=fg,
                            proc_no=qa_proc_no,
                            proc_code=qa_proc_code,
                            inst_type=_INST_U,
                            inst_code=_INST_CD_U,
                            item_or_resource=qa_resource,
                            task2=_QA_OUTPUT_TASK2_EXPR,
                            time_constraint_min=qa_time_constraint,
                        )
                    )

        if ink_map:
            inf_proc_no, inf_proc_code = _PROC_INFUSION
            seen_infusion: set[str] = set()
            for row in data_rows:
                process = str(_cell(row, idx, "process") or "").strip().lower()
                if process != "main line":
                    continue
                infusion_cd = _normalize_item_code(
                    _cell(row, idx, "Infusion code", "Infusion Code")
                )
                if not infusion_cd or infusion_cd in seen_infusion:
                    continue
                ink = _normalize_item_code(_cell(row, idx, "Ink Code", "Ink code"))
                if not ink or ink not in ink_map:
                    continue
                seen_infusion.add(infusion_cd)
                task2_inf = _format_infusion_cycle_task2(
                    _cell(row, idx, "Cycle time infusion", "Cycle time infusion")
                )
                seen_lines: set[str] = set()
                for line in ink_map[ink]:
                    if not line or line in seen_lines:
                        continue
                    seen_lines.add(line)
                    records.append(
                        _imaster_row(
                            final_item=infusion_cd,
                            proc_no=inf_proc_no,
                            proc_code=inf_proc_code,
                            inst_type=_INST_U,
                            inst_code=_INST_CD_U,
                            item_or_resource=line,
                            task2=task2_inf,
                        )
                    )

        return records
    finally:
        wb.close()
