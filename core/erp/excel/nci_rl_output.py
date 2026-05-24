"""NCI R/L Output: Delivery date Excel → IntegratedMaster_OutputInstruction.csv."""

from __future__ import annotations

import io
import re

from core.erp.excel.peb_integrated_master import PEB_IMASTER_HEADERS

NCI_RL_OUTPUT_HEADERS: tuple[str, ...] = PEB_IMASTER_HEADERS

_RL_OUTPUT_SHEETS = frozenset(
    {
        "BRIO PLANT 2",
        "BR-V",
        "HR-V & HR-V EXP",
        "WR-V + WR-V EXP",
    }
)
_PROC_NO = "10"
_INST_TYPE = "O"
_INST_CODE = "Out"
_MCFRAME_CODE_RE = re.compile(r"^(NCI_.+?_)(\d{5})(_.+)$")
_PART_CODE_HEADER_KEYS = frozenset({"partcode"})
_PART_NAME_HEADER_KEYS = frozenset({"partname"})
_MCFRAME_HEADER_KEYS = frozenset({"mcframeitemcode"})
_NOTES_HEADER_KEYS = frozenset({"notes"})
_ITEM_CD_HEADER_KEYS = frozenset({"itemcd"})


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name or "").strip())


def _normalize_header_name(raw: object) -> str:
    text = "" if raw is None else str(raw).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _normalize_part_code(raw: object) -> str:
    return re.sub(r"\s+", " ", str(raw or "").strip())


def _is_mcframe_item_code(value: object) -> bool:
    text = str(value or "").strip()
    return text.startswith("NCI_")


def _is_rl_pair(name_a: str, name_b: str) -> bool:
    """PART NAME が R/L の一字差のみのとき True。"""
    a = str(name_a or "").strip().upper()
    b = str(name_b or "").strip().upper()
    if not a or not b or a == "TOTAL" or b == "TOTAL":
        return False
    if len(a) != len(b):
        return False
    diffs = [(ca, cb) for ca, cb in zip(a, b, strict=True) if ca != cb]
    if len(diffs) != 1:
        return False
    ca, cb = diffs[0]
    return {ca, cb} == {"R", "L"}


def _mcframe_code_variant(code: str) -> tuple[str, str, str] | None:
    """``(prefix, part_number, suffix)`` — R/L で共通のバリアント判定に使う。"""
    match = _MCFRAME_CODE_RE.match(str(code or "").strip())
    if not match:
        return None
    return match.group(1), match.group(2), match.group(3)


def _mcframe_codes_same_variant(code_a: str, code_b: str) -> bool:
    """品目番号以外（色・型番サフィックス等）が同一の R/L ペアか。"""
    va = _mcframe_code_variant(code_a)
    vb = _mcframe_code_variant(code_b)
    if va is None or vb is None:
        return False
    prefix_a, _num_a, suffix_a = va
    prefix_b, _num_b, suffix_b = vb
    return prefix_a == prefix_b and suffix_a == suffix_b


def _is_valid_rl_output_pair(
    name_a: str, name_b: str, code_a: str, code_b: str
) -> bool:
    if code_a == code_b:
        return False
    return _is_rl_pair(name_a, name_b) and _mcframe_codes_same_variant(code_a, code_b)


def _output_record(final_item: str, other_item: str) -> dict[str, str]:
    return {
        "IMaster_FinalItemCode": final_item,
        "IMaster_ProcNo": _PROC_NO,
        "IMaster_ProcCode": _PROC_NO,
        "IMaster_InstructionType": _INST_TYPE,
        "IMaster_InstructionCode": _INST_CODE,
        "IMaster_ItemCodeOrResourceCode": other_item,
        "IMaster_Task2Expr": "",
        "IMaster_TimeConstraintMin": "",
    }


def _load_workbook_rows(raw: bytes) -> list[tuple]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise RuntimeError("The selected Excel file has no worksheet.")
        return [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _find_header_indexes(
    header_row: tuple,
    *,
    part_code_keys: frozenset[str] = _PART_CODE_HEADER_KEYS,
    part_name_keys: frozenset[str] = _PART_NAME_HEADER_KEYS,
    mcframe_keys: frozenset[str] = _MCFRAME_HEADER_KEYS,
    notes_keys: frozenset[str] = _NOTES_HEADER_KEYS,
    item_cd_keys: frozenset[str] = _ITEM_CD_HEADER_KEYS,
) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _normalize_header_name(cell)
        if key in part_code_keys and "part_code" not in indexes:
            indexes["part_code"] = idx
        elif key in part_name_keys and "part_name" not in indexes:
            indexes["part_name"] = idx
        elif key in mcframe_keys and "mcframe" not in indexes:
            indexes["mcframe"] = idx
        elif key in notes_keys and "notes" not in indexes:
            indexes["notes"] = idx
        elif key in item_cd_keys and "item_cd" not in indexes:
            indexes["item_cd"] = idx
    return indexes


def _find_delivery_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    for i, row in enumerate(rows[:40]):
        indexes = _find_header_indexes(row)
        if "part_code" in indexes and "part_name" in indexes:
            return i, indexes
    raise RuntimeError(
        'Delivery worksheet is missing a header row with "PART CODE" and "PART NAME".'
    )


def _find_item_master_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    for i, row in enumerate(rows[:40]):
        indexes = _find_header_indexes(row)
        if "notes" in indexes and "item_cd" in indexes:
            return i, indexes
    raise RuntimeError(
        'Item master file is missing a header row with "Notes" and "Item CD".'
    )


def build_nci_part_code_lookup(item_raw: bytes) -> dict[str, str]:
    """Item.xlsx 等の Notes (PART CODE) → Item CD (mcframe) マップを構築する。"""
    rows = _load_workbook_rows(item_raw)
    if not rows:
        raise RuntimeError("The item master file is empty.")

    header_idx, indexes = _find_item_master_header_row(rows)
    notes_idx = indexes["notes"]
    item_idx = indexes["item_cd"]

    lookup: dict[str, str] = {}
    for row in rows[header_idx + 1 :]:
        if notes_idx >= len(row) or item_idx >= len(row):
            continue
        part_code = _normalize_part_code(row[notes_idx])
        item_cd = str(row[item_idx] or "").strip()
        if not part_code or not _is_mcframe_item_code(item_cd):
            continue
        lookup.setdefault(part_code, item_cd)
    if not lookup:
        raise RuntimeError(
            'No mcframe Item CD rows were found under the "Notes" column in the item master file.'
        )
    return lookup


def _resolve_item_code(
    row: tuple,
    indexes: dict[str, int],
    lookup: dict[str, str] | None,
) -> str:
    mcframe_idx = indexes.get("mcframe")
    if mcframe_idx is not None and mcframe_idx < len(row):
        inline = row[mcframe_idx]
        if _is_mcframe_item_code(inline):
            return str(inline).strip()

    part_code_idx = indexes.get("part_code")
    if lookup is None or part_code_idx is None or part_code_idx >= len(row):
        return ""

    part_code = _normalize_part_code(row[part_code_idx])
    if not part_code:
        return ""
    return lookup.get(part_code, "")


def _collect_part_entries(
    rows: list[tuple],
    lookup: dict[str, str] | None,
) -> list[tuple[str, str]]:
    """PART NAME と解決済み mcframe Item code を上から収集する。"""
    header_idx, indexes = _find_delivery_header_row(rows)
    part_name_idx = indexes["part_name"]
    entries: list[tuple[str, str]] = []

    for row in rows[header_idx + 1 :]:
        part_name = str(
            row[part_name_idx] if len(row) > part_name_idx else ""
        ).strip()
        if not part_name or part_name.upper() == "TOTAL":
            continue
        item_cd = _resolve_item_code(row, indexes, lookup)
        if not item_cd:
            continue
        entries.append((part_name, item_cd))
    return entries


def _dedupe_output_records(records: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple[str, str]] = set()
    out: list[dict[str, str]] = []
    for record in records:
        key = (
            record["IMaster_FinalItemCode"],
            record["IMaster_ItemCodeOrResourceCode"],
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(record)
    return out


def _parse_rl_output_sheet(
    rows: list[tuple],
    lookup: dict[str, str] | None,
) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    entries = _collect_part_entries(rows, lookup)
    paired: set[tuple[str, str]] = set()
    for i in range(len(entries) - 1):
        name_a, code_a = entries[i]
        for j in range(i + 1, len(entries)):
            name_b, code_b = entries[j]
            if not _is_valid_rl_output_pair(name_a, name_b, code_a, code_b):
                continue
            pair_key = tuple(sorted((code_a, code_b)))
            if pair_key in paired:
                break
            paired.add(pair_key)
            records.append(_output_record(code_a, code_b))
            records.append(_output_record(code_b, code_a))
            break
    return records


def build_nci_rl_output_records(
    delivery_raw: bytes,
    item_raw: bytes | None = None,
) -> list[dict[str, str]]:
    """Delivery date Excel と Item マスタから R/L 出力指示行を生成する。"""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc

    lookup = build_nci_part_code_lookup(item_raw) if item_raw else None

    wb = load_workbook(filename=io.BytesIO(delivery_raw), read_only=True, data_only=True)
    try:
        sheet_by_name = {_normalize_sheet_name(n): n for n in wb.sheetnames}
        records: list[dict[str, str]] = []
        parsed_any = False

        for target in _RL_OUTPUT_SHEETS:
            actual = sheet_by_name.get(_normalize_sheet_name(target))
            if actual is None:
                continue
            ws = wb[actual]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            sheet_records = _parse_rl_output_sheet(rows, lookup)
            if sheet_records:
                parsed_any = True
                records.extend(sheet_records)

        if not parsed_any:
            raise RuntimeError(
                "No R/L part pairs were found on the expected vehicle worksheets."
            )
        return _dedupe_output_records(records)
    finally:
        wb.close()
