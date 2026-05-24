"""NCI Excel マスタ: BOM / ItemLine / Item / Line → Bridge CSV。"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from decimal import Decimal
from typing import Any

from core.erp.inventory_aggregate import (
    _format_stk_qty,
    _normalize_itm_cd,
    _parse_stk_qty,
)

from core.integrated_master import (
    INTEGRATED_HEADERS,
    append_supplier_use_lines_after_inputs,
    build_integrated_records,
    format_integrated_i_production,
)

NCI_ITEM_TABLE_HEADERS: tuple[str, ...] = (
    "ITM_CD",
    "ITM_NM",
    "ITM_TYP",
    "MAX_LOT_UNIT_QTY",
)

_ITEM_TYPE_MAP: dict[str, str] = {
    "1": "P",
    "2": "I",
    "5": "M",
    "6": "H",
    "7": "U",
}


def _normalize_header_name(raw: object) -> str:
    text = "" if raw is None else str(raw).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _load_workbook_rows(raw: bytes) -> list[tuple]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise RuntimeError("The selected Excel file has no worksheet.")
        return [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _find_header_indexes(header_row: tuple, *key_groups: tuple[str, frozenset[str]]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for name, keys in key_groups:
        for idx, cell in enumerate(header_row):
            if cell is not None and _normalize_header_name(cell) in keys:
                indexes[name] = idx
                break
    return indexes


def _find_header_row(
    rows: list[tuple],
    required: dict[str, frozenset[str]],
    *,
    label: str,
) -> tuple[int, dict[str, int]]:
    for i, row in enumerate(rows[:40]):
        indexes = _find_header_indexes(
            row,
            *[(name, keys) for name, keys in required.items()],
        )
        if all(name in indexes for name in required):
            return i, indexes
    missing = ", ".join(f'"{name}"' for name in required)
    raise RuntimeError(f'{label} is missing header column(s): {missing}.')


def _format_line_production(cycle_time: object, standard_load: object) -> str:
    if cycle_time is not None and str(cycle_time).strip() not in ("", "None"):
        text = str(cycle_time).strip()
        if text.lower().endswith("mp"):
            return text
        if text.replace(".", "", 1).isdigit():
            return f"{text}mp"
        return text
    if standard_load is not None and str(standard_load).strip() not in ("", "None"):
        sl = str(standard_load).strip()
        if sl.replace(".", "", 1).isdigit():
            return f"{sl}mp"
        return sl
    return ""


def _is_truthy(value: object) -> bool:
    return str(value or "").strip().lower() in ("true", "1", "yes", "y")


def _bom_flat_rows(bom_raw: bytes) -> list[dict[str, Any]]:
    rows = _load_workbook_rows(bom_raw)
    if not rows:
        return []

    header_idx, idx = _find_header_row(
        rows,
        {
            "parent": frozenset({"parentitemcd"}),
            "child": frozenset({"childitemcd"}),
            "pattern": frozenset({"bompattern"}),
            "operation": frozenset({"operationcd"}),
        },
        label="BOM file",
    )
    header_row = rows[header_idx]
    qty_idx = None
    for qkey in ("childinputqty", "childrequiredqty"):
        for col_i, cell in enumerate(header_row):
            if cell is not None and _normalize_header_name(cell) == qkey:
                qty_idx = col_i
                break
        if qty_idx is not None:
            break
    if qty_idx is None:
        raise RuntimeError('BOM file is missing "Child input qty." or "Child required qty." column.')
    idx["qty"] = qty_idx
    parent_i = idx["parent"]
    child_i = idx["child"]
    qty_i = idx["qty"]
    pattern_i = idx.get("pattern")
    operation_i = idx.get("operation")

    flat: list[dict[str, Any]] = []
    seq = 0
    for row in rows[header_idx + 1 :]:
        parent = str(row[parent_i] if len(row) > parent_i else "").strip()
        child = str(row[child_i] if len(row) > child_i else "").strip()
        if not parent or not child:
            continue
        if pattern_i is not None and len(row) > pattern_i:
            pattern = str(row[pattern_i] if row[pattern_i] is not None else "").strip()
            if pattern and pattern not in ("1", "1.0"):
                continue
        qty = row[qty_i] if len(row) > qty_i else None
        op = ""
        if operation_i is not None and len(row) > operation_i:
            op = str(row[operation_i] or "").strip()
        seq += 1
        flat.append(
            {
                "parent_item": parent,
                "component_code": child,
                "quantity": qty,
                "visual_order": op or seq,
            }
        )
    return flat


def _item_line_use_rows(item_line_raw: bytes) -> list[dict[str, Any]]:
    rows = _load_workbook_rows(item_line_raw)
    if not rows:
        return []

    header_idx, idx = _find_header_row(
        rows,
        {
            "item": frozenset({"itemcd"}),
            "line": frozenset({"linecd"}),
            "cycle": frozenset({"cycletime"}),
            "standard": frozenset({"standardload"}),
            "main": frozenset({"mainline"}),
        },
        label="ItemLine file",
    )
    item_i = idx["item"]
    line_i = idx["line"]
    cycle_i = idx["cycle"]
    standard_i = idx["standard"]
    main_i = idx.get("main")

    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if main_i is not None and len(row) > main_i and not _is_truthy(row[main_i]):
            continue
        item_cd = str(row[item_i] if len(row) > item_i else "").strip()
        line_cd = str(row[line_i] if len(row) > line_i else "").strip()
        if not item_cd or not line_cd:
            continue
        cycle = row[cycle_i] if len(row) > cycle_i else None
        standard = row[standard_i] if len(row) > standard_i else None
        production = _format_line_production(cycle, standard)
        out.append(
            {
                "P_ITM_CD": item_cd,
                "PROCESS_NO": 10,
                "PROCESS_CD": "10",
                "INST_TYP": "U",
                "INST_CD": "M",
                "ITM_RESOURCE": line_cd,
                "PRODUCTION": production,
            }
        )
    return out


def build_nci_integrated_master_records(
    bom_raw: bytes,
    item_line_raw: bytes,
) -> list[dict[str, str]]:
    """BOM (I) + ItemLine (U) + supplier 追補行を mcframe 形式で生成する。"""
    flat = _bom_flat_rows(bom_raw)
    if not flat:
        raise RuntimeError("No BOM rows were found in the selected BOM file.")

    records = build_integrated_records(flat)
    # mcframe Oracle 出力に合わせ、親品目ごとの PROD 行のみ除外する。
    records = [
        r
        for r in records
        if str(r.get("ITM_RESOURCE") or "").strip().upper() != "PROD"
    ]

    line_rows = _item_line_use_rows(item_line_raw)
    if not line_rows:
        raise RuntimeError("No Item Line rows were found in the selected ItemLine file.")

    merged: list[dict[str, Any]] = [dict(r) for r in records]
    merged.extend(line_rows)
    final = append_supplier_use_lines_after_inputs(merged)
    return [
        {h: str(row.get(h) or "").strip() for h in INTEGRATED_HEADERS}
        for row in final
    ]


def _map_item_type(raw: object) -> str:
    key = str(raw or "").strip()
    if key in _ITEM_TYPE_MAP:
        return _ITEM_TYPE_MAP[key]
    return key


def load_nci_item_table_rows_from_xlsx_bytes(raw: bytes) -> list[tuple[str, ...]]:
    rows = _load_workbook_rows(raw)
    if not rows:
        return []

    header_idx, idx = _find_header_row(
        rows,
        {
            "item": frozenset({"itemcd"}),
            "name": frozenset({"itemname"}),
            "type": frozenset({"itemtype"}),
        },
        label="Item master file",
    )
    item_i = idx["item"]
    name_i = idx["name"]
    type_i = idx["type"]

    out: list[tuple[str, ...]] = []
    for row in rows[header_idx + 1 :]:
        item_cd = str(row[item_i] if len(row) > item_i else "").strip()
        if not item_cd:
            continue
        item_nm = str(row[name_i] if len(row) > name_i else "").strip()
        item_typ = _map_item_type(row[type_i] if len(row) > type_i else "")
        out.append((item_cd, item_nm, item_typ, ""))
    if not out:
        raise RuntimeError("No item rows were found in the selected Item master file.")
    return out


def load_nci_resource_table_rows_from_xlsx_bytes(
    raw: bytes,
    *,
    sort_order_map: dict[str, int] | None = None,
) -> list[tuple[str, ...]]:
    rows = _load_workbook_rows(raw)
    if not rows:
        return []

    header_idx, idx = _find_header_row(
        rows,
        {
            "line": frozenset({"linecd"}),
            "name": frozenset({"linename"}),
        },
        label="Line master file",
    )
    line_i = idx["line"]
    name_i = idx["name"]
    sort_map = sort_order_map or {}

    out: list[tuple[str, ...]] = []
    for row in rows[header_idx + 1 :]:
        line_cd = str(row[line_i] if len(row) > line_i else "").strip()
        if not line_cd:
            continue
        line_nm = str(row[name_i] if len(row) > name_i else "").strip()
        resource_grp = "INJECTION" if line_cd.upper().startswith("M") else ""
        sort_order = sort_map.get(line_cd)
        out.append(
            (
                line_cd,
                line_nm,
                resource_grp,
                "" if sort_order is None else str(sort_order),
            )
        )
    if not out:
        raise RuntimeError("No line rows were found in the selected Line master file.")
    return out


_NCI_INVENTORY_SHEET = "Current stock list"
_NCI_ALLOCATABLE_THRESHOLD = Decimal(100000)


def _adjust_nci_allocatable_qty(total: Decimal) -> Decimal:
    """Allocatable qty 合計から 100,000 を繰り返し差し引く（100,000 以上の間）。"""
    qty = total
    while qty >= _NCI_ALLOCATABLE_THRESHOLD:
        qty -= _NCI_ALLOCATABLE_THRESHOLD
    return qty


def _nci_inventory_column_indices(header_row: tuple) -> tuple[int, int]:
    item_idx: int | None = None
    qty_idx: int | None = None
    for i, cell in enumerate(header_row):
        key = _normalize_header_name(cell)
        if key == "itemcd":
            item_idx = i
        elif key in ("allocatableqty", "allocatablequantity"):
            qty_idx = i
    if item_idx is None or qty_idx is None:
        raise RuntimeError(
            "The worksheet is missing required columns. "
            "Required: Item CD, Allocatable qty."
        )
    return item_idx, qty_idx


def _load_nci_inventory_rows_from_sheet_rows(all_rows: list[tuple]) -> list[tuple[str, str, str, str]]:
    if not all_rows:
        raise RuntimeError("The selected Excel file has no data rows.")
    header_row = all_rows[0]
    item_idx, qty_idx = _nci_inventory_column_indices(header_row)
    totals: defaultdict[str, Decimal] = defaultdict(lambda: Decimal(0))
    for values in all_rows[1:]:
        if not values:
            continue
        itm_cd = _normalize_itm_cd(
            values[item_idx] if item_idx < len(values) else ""
        )
        if not itm_cd:
            continue
        totals[itm_cd] += _parse_stk_qty(
            values[qty_idx] if qty_idx < len(values) else ""
        )
    if not totals:
        raise RuntimeError("No inventory rows were found in the selected file.")
    out: list[tuple[str, str, str, str]] = []
    for itm_cd in sorted(totals.keys()):
        stk = _adjust_nci_allocatable_qty(totals[itm_cd])
        out.append(("", itm_cd, _format_stk_qty(stk), ""))
    return out


def load_nci_inventory_rows_from_xlsx_bytes(raw: bytes) -> list[tuple[str, str, str, str]]:
    """Inventory.xlsx の Current stock list → (INV_CD, ITM_CD, STK_QTY, INV_DT) 行。"""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        if _NCI_INVENTORY_SHEET in wb.sheetnames:
            ws = wb[_NCI_INVENTORY_SHEET]
            return _load_nci_inventory_rows_from_sheet_rows(
                [tuple(r) for r in ws.iter_rows(values_only=True)]
            )
        for name in wb.sheetnames:
            rows = [tuple(r) for r in wb[name].iter_rows(values_only=True)]
            if not rows:
                continue
            try:
                _nci_inventory_column_indices(rows[0])
            except RuntimeError:
                continue
            return _load_nci_inventory_rows_from_sheet_rows(rows)
        raise RuntimeError(
            "No worksheet with Item CD and Allocatable qty. was found."
        )
    finally:
        wb.close()
