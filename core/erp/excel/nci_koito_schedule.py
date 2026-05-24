"""NCI KOITO 納入予定: Delivery date Excel → delivery_schedule_KOITO.csv."""

from __future__ import annotations

import io
import re
from datetime import datetime

KOITO_SCHEDULE_HEADERS: tuple[str, ...] = (
    "ORDER CD",
    "Item CD",
    "Date",
    "Qty",
    "Kanban",
    "Customer",
)

KOITO_CUSTOMER = "KOITO"
_ORDER_CD_PREFIX = "KOITO"
_ORDER_CD_START = 1
_ORDER_CD_WIDTH = 5
_SHEET_DELIVERY = "Sheet 1"
_ITEM_CD_HEADER = "itemcd"
_KANBAN_HEADER_KEYS = frozenset({"kodekanban", "kanban"})


def _normalize_header_name(raw: object) -> str:
    text = "" if raw is None else str(raw).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _format_koito_date(value: object) -> str:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    return str(value).strip()


def _format_koito_qty(value: object) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if value == 0:
            return None
        if float(value) == int(value):
            return str(int(value))
        text = str(value).strip()
        return text if text and text not in ("0", "0.0") else None
    text = str(value).strip().replace(",", "")
    if not text or text in ("0", "0.0"):
        return None
    try:
        num = float(text)
        if num == 0:
            return None
        if num == int(num):
            return str(int(num))
    except ValueError:
        pass
    return text


def _format_order_cd(sequence: int) -> str:
    return f"{_ORDER_CD_PREFIX}{sequence:0{_ORDER_CD_WIDTH}d}"


def _format_kanban(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_column_index(header_row: tuple, header_keys: frozenset[str]) -> int | None:
    for i, cell in enumerate(header_row):
        if cell is not None and _normalize_header_name(cell) in header_keys:
            return i
    return None


def _find_delivery_header_row(rows: list[tuple]) -> int:
    for i, row in enumerate(rows[:20]):
        for cell in row:
            if cell is not None and _normalize_header_name(cell) == _ITEM_CD_HEADER:
                return i
    raise RuntimeError(
        f'The worksheet "{_SHEET_DELIVERY}" is missing an "Item CD" header row.'
    )


def build_koito_delivery_schedule_rows(
    raw: bytes,
) -> list[tuple[str, str, str, str, str, str]]:
    """Delivery date Excel の日次数量を ``delivery_schedule_KOITO.csv`` 行に展開する。"""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        if _SHEET_DELIVERY in wb.sheetnames:
            ws = wb[_SHEET_DELIVERY]
        else:
            ws = wb.active
            if ws is None:
                raise RuntimeError("The selected Excel file has no worksheet.")

        all_rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        if not all_rows:
            return []

        header_idx = _find_delivery_header_row(all_rows)
        header_row = all_rows[header_idx]
        item_idx = next(
            i
            for i, cell in enumerate(header_row)
            if cell is not None and _normalize_header_name(cell) == _ITEM_CD_HEADER
        )
        kanban_idx = _find_column_index(header_row, _KANBAN_HEADER_KEYS)
        date_columns: list[tuple[int, datetime]] = []
        for i, cell in enumerate(header_row):
            if isinstance(cell, datetime):
                date_columns.append((i, cell))

        if not date_columns:
            raise RuntimeError(
                f'The worksheet "{ws.title}" has no date columns in the header row.'
            )

        out: list[tuple[str, str, str, str, str, str]] = []
        order_seq = _ORDER_CD_START
        for row in all_rows[header_idx + 1 :]:
            if item_idx >= len(row):
                continue
            item_raw = row[item_idx]
            if item_raw is None or not str(item_raw).strip():
                continue
            item_cd = str(item_raw).strip()
            kanban = ""
            if kanban_idx is not None and kanban_idx < len(row):
                kanban = _format_kanban(row[kanban_idx])

            for col_idx, dt in date_columns:
                if col_idx >= len(row):
                    continue
                qty = _format_koito_qty(row[col_idx])
                if qty is None:
                    continue
                out.append(
                    (
                        _format_order_cd(order_seq),
                        item_cd,
                        _format_koito_date(dt),
                        qty,
                        kanban,
                        KOITO_CUSTOMER,
                    )
                )
                order_seq += 1
        return out
    finally:
        wb.close()
