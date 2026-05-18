from __future__ import annotations

import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from core.erp.inventory_aggregate import aggregate_inventory_rows_by_itm_cd


def _normalize_header_name(raw: object) -> str:
    text = "" if raw is None else str(raw).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _format_excel_date(value: object) -> str:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(text, fmt)
            return f"{dt.month}/{dt.day}/{dt.year}"
        except ValueError:
            continue
    return text


def _normalize_peb_item_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return str(int(float(s)))
    if re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9\-_.]*", s):
        return s.upper()
    return s


_ITEM_CODE_HEADER_PRIORITY = (
    "itemcode",
    "itmcd",
    "material",
    "item",
    "partno",
    "partnumber",
    "part",
    "sku",
)
_DESC_HEADER_PRIORITY = ("itemdesc", "itemdescription", "description", "desc")
_INVENTORY_TRAN_EXCLUDE = frozenset(
    {
        "committed",
        "onorder",
        "on order",
        "intransit",
        "in transit",
        "reserved",
        "allocated",
    }
)


def _parse_peb_qty_cell(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    s = str(value).strip()
    if not s:
        return None
    if s.lower() in ("on hand", "onhand", "stock", "-", "n/a", "na"):
        return None
    s = s.replace(",", "")
    m = re.match(r"^([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)", s)
    if m:
        try:
            return Decimal(m.group(1))
        except InvalidOperation:
            return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _format_peb_qty(total: Decimal) -> str:
    if total.is_zero():
        return "0"
    if total == total.to_integral_value():
        return str(int(total))
    s = format(total.normalize(), "f")
    return s.rstrip("0").rstrip(".") or "0"


def _header_index_map(header_row: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, v in enumerate(header_row):
        key = _normalize_header_name(v)
        if key:
            out[key] = i
    return out


def _peb_pick_item_column(header_row: tuple) -> tuple[int, int | None]:
    keys = [_normalize_header_name(v) for v in header_row]
    code_idx: int | None = None
    for pref in _ITEM_CODE_HEADER_PRIORITY:
        for i, key in enumerate(keys):
            if key == pref:
                code_idx = i
                break
        if code_idx is not None:
            break
    desc_idx: int | None = None
    for pref in _DESC_HEADER_PRIORITY:
        for i, key in enumerate(keys):
            if key == pref:
                desc_idx = i
                break
        if desc_idx is not None:
            break
    if code_idx is not None:
        return code_idx, desc_idx
    if desc_idx is not None:
        return desc_idx, None
    return 0, 1 if len(header_row) > 1 else None


def _peb_inventory_tran_row_included(tran_val: str) -> bool:
    """Exclude only known non-stock transaction rows; do not require 'On Hand'."""
    if not tran_val:
        return True
    t = re.sub(r"\s+", " ", tran_val.lower().strip())
    if t in _INVENTORY_TRAN_EXCLUDE:
        return False
    return True


def _peb_is_header_like_item(item: str) -> bool:
    key = re.sub(r"[^a-z0-9]+", "", item.lower())
    return key in (
        "itemcode",
        "itemdesc",
        "item",
        "material",
        "partno",
        "description",
        "total",
    )


def _peb_pick_tran_column(header_row: tuple) -> int | None:
    for i, v in enumerate(header_row):
        if _normalize_header_name(v) in ("tran", "transaction"):
            return i
    return None


_SINGLE_QTY_HEADER_PRIORITY = (
    "available",
    "onhand",
    "stkqty",
    "stockqty",
    "qty",
    "quantity",
    "stock",
)
_NON_QTY_HEADER_KEYS = frozenset(
    {
        "material",
        "itemcode",
        "itmcd",
        "item",
        "partno",
        "partnumber",
        "part",
        "sku",
        "itemdesc",
        "itemdescription",
        "description",
        "desc",
        "uom",
        "unit",
        "unitofmeasure",
        "location",
        "locatio",
        "loc",
        "bin",
        "warehouse",
        "wh",
        "tran",
        "transaction",
    }
)


def _header_row_has_item_column(header_row: tuple) -> bool:
    keys = [_normalize_header_name(v) for v in header_row]
    return any(k in _ITEM_CODE_HEADER_PRIORITY for k in keys if k)


def _header_row_has_qty_column(header_row: tuple) -> bool:
    for v in header_row:
        key = _normalize_header_name(v)
        if key in _SINGLE_QTY_HEADER_PRIORITY:
            return True
        text = "" if v is None else str(v).strip()
        if not text:
            continue
        token = text.split()[0]
        dt = _format_excel_date(token)
        if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", dt):
            return True
    return False


def _find_inventory_header_row(all_rows: list[tuple]) -> int:
    """タイトル行のあとにヘッダーがあるシート向けに、品目・数量列を含む行を探す。"""
    for i, row in enumerate(all_rows[:40]):
        if _header_row_has_item_column(row) and _header_row_has_qty_column(row):
            return i
    return 0


def _read_worksheet_rows(ws) -> list[tuple]:
    """結合セルの値を各セルに展開して行リストを返す。"""
    merged_fill: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_fill[(r, c)] = val
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    out: list[tuple] = []
    for r in range(1, max_row + 1):
        row_vals: list[object] = []
        for c in range(1, max_col + 1):
            row_vals.append(merged_fill.get((r, c), ws.cell(r, c).value))
        out.append(tuple(row_vals))
    return out


def _peb_pick_qty_columns(header_row: tuple, skip: set[int]) -> list[int]:
    explicit: list[int] = []
    for i, v in enumerate(header_row):
        if i in skip:
            continue
        key = _normalize_header_name(v)
        if key in _SINGLE_QTY_HEADER_PRIORITY:
            explicit.append(i)
    if explicit:
        for pref in _SINGLE_QTY_HEADER_PRIORITY:
            for i in explicit:
                if _normalize_header_name(header_row[i]) == pref:
                    return [i]
        return explicit
    date_cols: list[int] = []
    for i, v in enumerate(header_row):
        if i in skip:
            continue
        text = "" if v is None else str(v).strip()
        if not text:
            continue
        token = text.split()[0]
        dt = _format_excel_date(token)
        if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", dt):
            date_cols.append(i)
    if date_cols:
        return date_cols
    fallback = [
        i
        for i in range(len(header_row))
        if i not in skip
        and _normalize_header_name(header_row[i])
        and _normalize_header_name(header_row[i]) not in _NON_QTY_HEADER_KEYS
    ]
    if fallback:
        return fallback
    return [i for i in range(len(header_row)) if i not in skip]


def _peb_row_qty_total(values: tuple, qty_indices: list[int]) -> Decimal | None:
    total = Decimal(0)
    found = False
    for i in qty_indices:
        if i >= len(values):
            continue
        q = _parse_peb_qty_cell(values[i])
        if q is not None:
            total += q
            found = True
    return total if found else None


def _peb_inv_dt_from_header(header_row: tuple, qty_indices: list[int]) -> str:
    for i in reversed(qty_indices):
        if i >= len(header_row):
            continue
        h = header_row[i]
        if isinstance(h, datetime):
            return _format_excel_date(h)
        dt = _format_excel_date(h)
        if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", dt):
            return dt
    return ""


def load_peb_order_rows_from_xlsx_bytes(raw: bytes) -> list[tuple[str, str, str, str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise RuntimeError("The selected Excel file has no worksheet.")
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []

        header_index = {_normalize_header_name(v): i for i, v in enumerate(header_row)}

        def _idx(*names: str) -> int | None:
            for n in names:
                idx = header_index.get(_normalize_header_name(n))
                if idx is not None:
                    return idx
            return None

        idx_item = _idx("Actual Item", "Item Code", "ITM_CD")
        idx_qty = _idx("Qty", "REQ_QTY")
        idx_dlv = _idx("exfact", "Exfact Date", "Exfact", "DLV_DT")
        idx_req = _idx("Index", "REQ_NO")

        if idx_item is None or idx_qty is None or idx_dlv is None:
            raise RuntimeError(
                "The worksheet is missing required columns. "
                "Required: Actual Item, exfact, Qty."
            )

        out: list[tuple[str, str, str, str, str]] = []
        seq = 1
        for values in rows:
            item = _normalize_peb_item_code(
                values[idx_item] if idx_item < len(values) else ""
            )
            qty_raw = values[idx_qty] if idx_qty < len(values) else ""
            dlv_raw = values[idx_dlv] if idx_dlv < len(values) else ""
            if not item:
                continue

            qty_text = "" if qty_raw is None else str(qty_raw).strip()
            if not qty_text:
                continue

            req_raw = values[idx_req] if (idx_req is not None and idx_req < len(values)) else ""
            req_no = str(req_raw).strip() if req_raw is not None else ""
            if not req_no:
                req_no = str(seq)

            out.append(
                (req_no, item, _format_excel_date(dlv_raw), qty_text, "Shipping")
            )
            seq += 1
        return out
    finally:
        wb.close()


def _load_peb_inventory_rows_from_sheet(
    ws,
    *,
    inv_cd_prefix: str,
    inv_dt: str,
    require_tran: bool,
) -> list[tuple[str, str, str, str]]:
    all_rows = _read_worksheet_rows(ws)
    if not all_rows:
        return []
    header_idx = _find_inventory_header_row(all_rows)
    header_row = all_rows[header_idx]
    data_rows = all_rows[header_idx + 1 :]

    item_idx, desc_idx = _peb_pick_item_column(header_row)
    tran_idx = _peb_pick_tran_column(header_row) if require_tran else None
    skip_cols = {item_idx}
    if desc_idx is not None:
        skip_cols.add(desc_idx)
    if tran_idx is not None:
        skip_cols.add(tran_idx)
    qty_cols = _peb_pick_qty_columns(header_row, skip_cols)
    if not qty_cols:
        raise RuntimeError("No stock quantity column was found.")

    if not inv_dt:
        inv_dt = _peb_inv_dt_from_header(header_row, qty_cols)

    out: list[tuple[str, str, str, str]] = []
    seq = 1
    last_item = ""
    for values in data_rows:
        raw_item = _normalize_peb_item_code(
            values[item_idx] if item_idx < len(values) else ""
        )
        if raw_item:
            last_item = raw_item
        item = last_item
        if not item or _peb_is_header_like_item(item):
            continue
        if require_tran:
            tran_val = str(
                values[tran_idx]
                if (tran_idx is not None and tran_idx < len(values))
                else ""
            ).strip()
            if not _peb_inventory_tran_row_included(tran_val):
                continue
        qty_total = _peb_row_qty_total(values, qty_cols)
        if qty_total is None:
            continue
        inv_cd = f"{inv_cd_prefix}{seq:05d}"
        out.append((inv_cd, item, _format_peb_qty(qty_total), inv_dt))
        seq += 1
    return aggregate_inventory_rows_by_itm_cd(
        out,
        inv_cd_prefix=inv_cd_prefix,
    )


def _load_peb_current_stock_rows_from_sheet(
    ws,
    *,
    inv_cd_prefix: str = "INV",
    inv_dt: str = "",
) -> list[tuple[str, str, str, str]]:
    """Current Stock: Material → ITM_CD, Available → STK_QTY（ITM_CD 単位に集計）。"""
    all_rows = _read_worksheet_rows(ws)
    if not all_rows:
        return []
    header_idx = _find_inventory_header_row(all_rows)
    header_row = all_rows[header_idx]
    idx = _header_index_map(header_row)
    data_rows = all_rows[header_idx + 1 :]

    item_idx = idx.get("material")
    if item_idx is None:
        item_idx, _ = _peb_pick_item_column(header_row)
    qty_idx = idx.get("available")
    if qty_idx is None:
        qty_cols = _peb_pick_qty_columns(header_row, {item_idx} if item_idx is not None else set())
        qty_idx = qty_cols[0] if qty_cols else None

    if item_idx is None or qty_idx is None:
        raise RuntimeError(
            "The worksheet is missing required columns. Required: Material, Available."
        )

    out: list[tuple[str, str, str, str]] = []
    seq = 1
    last_item = ""
    for values in data_rows:
        raw_item = _normalize_peb_item_code(
            values[item_idx] if item_idx < len(values) else ""
        )
        if raw_item:
            last_item = raw_item
        item = last_item
        if not item or _peb_is_header_like_item(item):
            continue
        qty_total = _peb_row_qty_total(values, [qty_idx])
        if qty_total is None:
            continue
        inv_cd = f"{inv_cd_prefix}{seq:05d}"
        out.append((inv_cd, item, _format_peb_qty(qty_total), inv_dt))
        seq += 1
    return aggregate_inventory_rows_by_itm_cd(
        out,
        inv_cd_prefix=inv_cd_prefix,
    )


def load_peb_inventory_rows_from_xlsx_bytes(raw: bytes) -> list[tuple[str, str, str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc
    wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise RuntimeError("The selected Excel file has no worksheet.")
        inv_dt = ""
        m = re.match(r"^Stock_list_(\d{4})(\d{2})(\d{2})", str(ws.title or ""))
        if m:
            inv_dt = f"{int(m.group(2))}/{int(m.group(3))}/{int(m.group(1))}"
        return _load_peb_current_stock_rows_from_sheet(
            ws,
            inv_cd_prefix="INV",
            inv_dt=inv_dt,
        )
    finally:
        wb.close()


def load_peb_inventory_wip_rows_from_xlsx_bytes(raw: bytes) -> list[tuple[str, str, str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc
    wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    try:
        ws_name = next((n for n in wb.sheetnames if str(n).startswith("Stock_list_")), None)
        if not ws_name:
            raise RuntimeError("PEB WIP inventory import requires a sheet starting with 'Stock_list_'.")
        inv_dt = ""
        m = re.match(r"^Stock_list_(\d{4})(\d{2})(\d{2})", ws_name)
        if m:
            inv_dt = f"{int(m.group(2))}/{int(m.group(3))}/{int(m.group(1))}"
        return _load_peb_inventory_rows_from_sheet(
            wb[ws_name],
            inv_cd_prefix="WIP",
            inv_dt=inv_dt,
            require_tran=False,
        )
    finally:
        wb.close()


def load_peb_prd_plan_rows_from_xlsx_bytes(raw: bytes) -> list[tuple[str, str, str, str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws_name = next((n for n in wb.sheetnames if str(n).startswith("ProductionPlan")), None)
        if not ws_name:
            raise RuntimeError("PEB Prd Plan import requires a 'ProductionPlan' sheet.")
        ws = wb[ws_name]
        rows = ws.iter_rows(values_only=True)
        _ = next(rows, None)  # title row
        header_row = next(rows, None)
        if not header_row:
            return []

        fg_idx = None
        tran_idx = None
        date_cols: list[tuple[int, str]] = []
        for i, v in enumerate(header_row):
            key = _normalize_header_name(v)
            if fg_idx is None and key in ("fgcode", "itemcode", "itmcd", "itm_cd"):
                fg_idx = i
            elif key in ("tran", "transaction"):
                tran_idx = i

            text = "" if v is None else str(v).strip()
            if not text:
                continue
            token = text.split()[0]
            dt = _format_excel_date(token)
            if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", dt):
                date_cols.append((i, dt))

        if fg_idx is None:
            raise RuntimeError("No FG Code column was found in the ProductionPlan sheet.")
        if not date_cols:
            raise RuntimeError("No date header column was found in the ProductionPlan sheet.")

        out: list[tuple[str, str, str, str, str]] = []
        seq = 1
        for values in rows:
            itm_cd = str(values[fg_idx] if fg_idx < len(values) else "").strip()
            if not itm_cd:
                continue
            tran_val = str(values[tran_idx] if (tran_idx is not None and tran_idx < len(values)) else "").strip().lower()
            if tran_val and tran_val != "production":
                continue

            for col_idx, dlv_dt in date_cols:
                qty_raw = values[col_idx] if col_idx < len(values) else ""
                qty_text = "" if qty_raw is None else str(qty_raw).strip()
                if not qty_text:
                    continue
                try:
                    if float(qty_text) == 0:
                        continue
                except ValueError:
                    pass
                req_no = f"PRDPLAN-{seq:06d}"
                out.append((req_no, itm_cd, dlv_dt, qty_text, ""))
                seq += 1
        return out
    finally:
        wb.close()


def _last_numeric_value(cells: tuple[object, ...]) -> float | None:
    for value in reversed(cells):
        if isinstance(value, (int, float)):
            return float(value)
    return None


def _format_numeric_text(value: float) -> str:
    if value.is_integer():
        return f"{int(value):,}"
    text = f"{value:,.3f}"
    return text.rstrip("0").rstrip(".")


def load_peb_monthly_result_rows_from_xlsx_bytes(raw: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        if "Monthly_ym" not in wb.sheetnames:
            raise RuntimeError("PEB Monthly Result import requires a 'Monthly_ym' sheet.")
        ws = wb["Monthly_ym"]
        qty_by_line: dict[str, float] = {}
        wh_by_line: dict[str, float] = {}
        mode: str | None = None

        for row in ws.iter_rows(values_only=True):
            key_col2 = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()
            key = str(row[2] if len(row) > 2 and row[2] is not None else "").strip()
            if not key:
                if key_col2 == "QTY":
                    mode = "qty"
                elif key_col2 == "WH":
                    mode = "wh"
                continue
            lowered = key.lower()
            if lowered == "sum of qty":
                mode = "qty-header"
                continue
            if lowered == "sum of working hour":
                mode = "wh-header"
                continue
            if key == "Grand Total":
                continue
            if mode not in ("qty", "wh"):
                continue

            total = _last_numeric_value(row[3:])
            if total is None:
                continue
            if mode == "qty":
                qty_by_line[key] = qty_by_line.get(key, 0.0) + total
            else:
                wh_by_line[key] = wh_by_line.get(key, 0.0) + total

        keys = sorted(set(qty_by_line.keys()) | set(wh_by_line.keys()))
        return [
            {
                "line_name": k,
                "production_qty": _format_numeric_text(qty_by_line.get(k, 0.0)),
                "working_hours": _format_numeric_text(wh_by_line.get(k, 0.0)),
            }
            for k in keys
        ]
    finally:
        wb.close()
