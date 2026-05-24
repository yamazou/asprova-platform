"""NCI HPM 納入予定: Delivery date Excel → delivery_schedule_HPM.csv."""

from __future__ import annotations

import io
import re
from datetime import datetime

HPM_SCHEDULE_HEADERS: tuple[str, ...] = (
    "ORDER CD",
    "Item CD",
    "Date",
    "Qty",
    "Kanban",
    "Customer",
)

_ORDER_CD_PREFIX = "HPM"
_ORDER_CD_START = 1
_ORDER_CD_WIDTH = 5
_SKIP_SHEETS = frozenset({"itemmaster", "sheet4"})
_MCFRAME_ITEM_KEYS = frozenset({"mcframeitemcode"})
_MODEL_HEADER_KEYS = frozenset({"model"})
_COLOR_HEADER_KEYS = frozenset({"color"})


def _normalize_header_name(raw: object) -> str:
    text = "" if raw is None else str(raw).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _format_schedule_date(value: object) -> str:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    return str(value).strip()


def _format_schedule_qty(value: object) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if value == 0:
            return None
        if float(value) == int(value):
            return str(int(value))
        text = str(value).strip()
        return text if text and text not in ("0", "0.0") else None
    text = str(value).strip().replace(",", "")
    if not text or text in ("0", "0.0"):
        return None
    try:
        num = float(text)
        if num == 0:
            return None
        if num == int(num):
            return str(int(num))
    except ValueError:
        pass
    return text


def _format_order_cd(sequence: int) -> str:
    return f"{_ORDER_CD_PREFIX}{sequence:0{_ORDER_CD_WIDTH}d}"


def _format_kanban(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.upper() == "TOTAL":
        return ""
    return text


def _is_mcframe_item_code(value: object) -> bool:
    text = str(value or "").strip()
    return text.startswith("NCI_")


def _find_column_index(header_row: tuple, header_keys: frozenset[str]) -> int | None:
    for i, cell in enumerate(header_row):
        if cell is not None and _normalize_header_name(cell) in header_keys:
            return i
    return None


def _find_mcframe_header_row(rows: list[tuple]) -> int | None:
    for i, row in enumerate(rows[:25]):
        for cell in row:
            if cell is not None and _normalize_header_name(cell) in _MCFRAME_ITEM_KEYS:
                return i
    return None


def _row_qty_by_date(
    row: tuple,
    date_columns: list[tuple[int, datetime]],
) -> dict[int, str]:
    out: dict[int, str] = {}
    for col_idx, _dt in date_columns:
        if col_idx >= len(row):
            continue
        qty = _format_schedule_qty(row[col_idx])
        if qty is not None:
            out[col_idx] = qty
    return out


def _emit_rows(
    *,
    out: list[tuple[str, str, str, str, str, str]],
    order_seq: int,
    item_cd: str,
    kanban: str,
    customer: str,
    qty_by_col: dict[int, str],
    date_columns: list[tuple[int, datetime]],
) -> int:
    if not item_cd or not kanban:
        return order_seq
    for col_idx, dt in date_columns:
        qty = qty_by_col.get(col_idx)
        if qty is None:
            continue
        out.append(
            (
                _format_order_cd(order_seq),
                item_cd,
                _format_schedule_date(dt),
                qty,
                kanban,
                customer,
            )
        )
        order_seq += 1
    return order_seq


def _parse_hpm_schedule_sheet(
    rows: list[tuple],
    *,
    customer: str,
    out: list[tuple[str, str, str, str, str, str]],
    order_seq: int,
) -> int:
    header_idx = _find_mcframe_header_row(rows)
    if header_idx is None:
        return order_seq

    header_row = rows[header_idx]
    item_idx = _find_column_index(header_row, _MCFRAME_ITEM_KEYS)
    if item_idx is None:
        return order_seq

    model_idx = _find_column_index(header_row, _MODEL_HEADER_KEYS)
    color_idx = _find_column_index(header_row, _COLOR_HEADER_KEYS)
    if model_idx is not None:
        kanban_idx = model_idx
    elif color_idx is not None:
        kanban_idx = color_idx
    else:
        return order_seq

    date_columns: list[tuple[int, datetime]] = []
    for i, cell in enumerate(header_row):
        if isinstance(cell, datetime):
            date_columns.append((i, cell))
    if not date_columns:
        return order_seq

    current_item = ""

    for row in rows[header_idx + 1 :]:
        part_name = str(row[2] if len(row) > 2 else "").strip().upper()
        if part_name == "TOTAL":
            continue

        mc_raw = row[item_idx] if item_idx < len(row) else None
        if _is_mcframe_item_code(mc_raw):
            current_item = str(mc_raw).strip()

        kanban = _format_kanban(row[kanban_idx] if kanban_idx < len(row) else None)
        if not kanban or not current_item:
            continue

        row_qty = _row_qty_by_date(row, date_columns)
        if not row_qty:
            continue
        order_seq = _emit_rows(
            out=out,
            order_seq=order_seq,
            item_cd=current_item,
            kanban=kanban,
            customer=customer,
            qty_by_col=row_qty,
            date_columns=date_columns,
        )

    return order_seq


def build_hpm_delivery_schedule_rows(
    raw: bytes,
) -> list[tuple[str, str, str, str, str, str]]:
    """Delivery date Excel（複数シート）から ``delivery_schedule_HPM.csv`` 行を生成する。"""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "openpyxl is required to read Excel (.xlsx) files."
        ) from exc

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        out: list[tuple[str, str, str, str, str, str]] = []
        order_seq = _ORDER_CD_START
        parsed_any = False

        for sheet_name in wb.sheetnames:
            if _normalize_header_name(sheet_name) in _SKIP_SHEETS:
                continue
            ws = wb[sheet_name]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            before = order_seq
            order_seq = _parse_hpm_schedule_sheet(
                rows,
                customer=str(sheet_name).strip(),
                out=out,
                order_seq=order_seq,
            )
            if order_seq > before:
                parsed_any = True

        if not parsed_any:
            raise RuntimeError(
                "No schedule worksheets with mcframe Item code were found in the workbook."
            )
        return out
    finally:
        wb.close()
