"""
Recreate dbo.ITT1_TMP on SQL Server (database SW) from ITT1_TMP.xlsx layout.

Drops the existing table (all columns removed), creates the standard staging
schema, then loads all rows from the Excel file.

Example (PowerShell):

  $env:SQLSERVER_PASSWORD = 'your-password'
  py -3 scripts/load_itt1_tmp_from_excel.py ^
    --server "HOST\\SQLEXPRESS" --database SW --user sa ^
    --xlsx "C:\\Users\\you\\Downloads\\ITT1_TMP.xlsx"
"""

from __future__ import annotations

import argparse
import os
import sys
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook

# Column order and SQL types (matches scripts/sqlserver_sap_bom_oitt_itt1.sql).
COL_SPECS: list[tuple[str, str]] = [
    ("#", "int"),
    ("Parent Item", "nvarchar"),
    ("Component Element Number", "int"),
    ("Visual Order", "int"),
    ("Component Code", "nvarchar"),
    ("Quantity", "decimal"),
    ("Warehouse", "nvarchar"),
    ("Price", "decimal"),
    ("Currency", "nvarchar"),
    ("Price List", "int"),
    ("Original Price", "decimal"),
    ("Original Currency", "nvarchar"),
    ("Issue Method", "nvarchar"),
    ("Inventory UoM", "nvarchar"),
    ("Comment", "nvarchar"),
    ("Log Instance", "int"),
    ("Object", "nvarchar"),
    ("Distribution Rule", "nvarchar"),
    ("Distribution Rule2", "nvarchar"),
    ("Distribution Rule3", "nvarchar"),
    ("Distribution Rule4", "nvarchar"),
    ("Distribution Rule5", "nvarchar"),
    ("Principal Input", "nvarchar"),
    ("Project Code", "nvarchar"),
    ("Component Type", "int"),
    ("WIP Account Code", "nvarchar"),
    ("Additional Quantity", "decimal"),
    ("Row Text", "nvarchar"),
    ("Stage ID", "nvarchar"),
    ("Item Description", "nvarchar"),
]

DROP_ITT1 = "IF OBJECT_ID(N'dbo.ITT1_TMP', N'U') IS NOT NULL DROP TABLE dbo.ITT1_TMP;"

CREATE_ITT1 = """
CREATE TABLE dbo.ITT1_TMP (
    [#]                          INT             NULL,
    [Parent Item]                NVARCHAR(64)    NULL,
    [Component Element Number]   INT             NULL,
    [Visual Order]               INT             NULL,
    [Component Code]             NVARCHAR(64)    NULL,
    [Quantity]                   DECIMAL(18, 6)  NULL,
    [Warehouse]                  NVARCHAR(16)    NULL,
    [Price]                      DECIMAL(18, 6)  NULL,
    [Currency]                   NVARCHAR(16)    NULL,
    [Price List]                 INT             NULL,
    [Original Price]             DECIMAL(18, 6)  NULL,
    [Original Currency]          NVARCHAR(16)    NULL,
    [Issue Method]               NVARCHAR(8)     NULL,
    [Inventory UoM]              NVARCHAR(32)    NULL,
    [Comment]                    NVARCHAR(512)   NULL,
    [Log Instance]               INT             NULL,
    [Object]                     NVARCHAR(16)    NULL,
    [Distribution Rule]          NVARCHAR(128)   NULL,
    [Distribution Rule2]         NVARCHAR(128)   NULL,
    [Distribution Rule3]         NVARCHAR(128)   NULL,
    [Distribution Rule4]         NVARCHAR(128)   NULL,
    [Distribution Rule5]         NVARCHAR(128)   NULL,
    [Principal Input]            NVARCHAR(8)     NULL,
    [Project Code]               NVARCHAR(64)    NULL,
    [Component Type]             INT             NULL,
    [WIP Account Code]           NVARCHAR(64)    NULL,
    [Additional Quantity]        DECIMAL(18, 6)  NULL,
    [Row Text]                   NVARCHAR(512)   NULL,
    [Stage ID]                   NVARCHAR(64)    NULL,
    [Item Description]           NVARCHAR(512)   NULL
);
"""


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _coerce(value: Any, kind: str) -> Any:
    if _is_empty(value):
        return None
    if kind == "int":
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value != value:  # NaN
                return None
            return int(round(float(value)))
        s = str(value).strip()
        if s == "":
            return None
        return int(round(float(s)))
    if kind == "decimal":
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value != value:
                return None
            return Decimal(str(value))
        s = str(value).strip()
        if s == "":
            return None
        return Decimal(s)
    # nvarchar
    if isinstance(value, float):
        if value != value:
            return None
        if value == int(value):
            return str(int(value))
        return str(value).rstrip("0").rstrip(".") if "." in str(value) else str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip() or None


def _connect(server: str, database: str, user: str, password: str):
    import pyodbc

    drivers = [
        "ODBC Driver 18 for SQL Server",
        "ODBC Driver 17 for SQL Server",
        "SQL Server",
    ]
    last_err: Exception | None = None
    for drv in drivers:
        conn_s = (
            f"DRIVER={{{drv}}};"
            f"SERVER={server};"
            f"DATABASE={database};"
            f"UID={user};"
            f"PWD={password};"
            "TrustServerCertificate=yes;"
        )
        try:
            return pyodbc.connect(conn_s, timeout=15)
        except Exception as e:  # noqa: BLE001
            last_err = e
            continue
    raise RuntimeError(f"Could not connect with any ODBC driver. Last error: {last_err}") from last_err


def _read_rows(path: str) -> list[list[Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        expected = [name for name, _ in COL_SPECS]
        if headers != expected:
            raise ValueError(
                "Excel header row does not match ITT1_TMP layout.\n"
                f"Expected: {expected}\nGot:      {headers}"
            )
        out: list[list[Any]] = []
        for row in rows_iter:
            if row is None or all(_is_empty(c) for c in row):
                continue
            out.append(list(row))
        return out
    finally:
        wb.close()


def main() -> int:
    p = argparse.ArgumentParser(description="Drop/recreate dbo.ITT1_TMP and load from Excel.")
    p.add_argument("--server", default=os.environ.get("SQLSERVER_SERVER", ""))
    p.add_argument("--database", default=os.environ.get("SQLSERVER_DATABASE", "SW"))
    p.add_argument("--user", default=os.environ.get("SQLSERVER_USER", ""))
    p.add_argument(
        "--xlsx",
        default=os.environ.get(
            "ITT1_TMP_XLSX",
            r"c:\Users\lenovo\Downloads\ITT1_TMP.xlsx",
        ),
    )
    args = p.parse_args()

    password = os.environ.get("SQLSERVER_PASSWORD", "")
    if not args.server or not args.user or not password:
        print(
            "Set SQLSERVER_PASSWORD and pass --server / --user (or SQLSERVER_* env vars).",
            file=sys.stderr,
        )
        return 1
    if not os.path.isfile(args.xlsx):
        print(f"Excel file not found: {args.xlsx}", file=sys.stderr)
        return 1

    data_rows = _read_rows(args.xlsx)
    col_names = [name for name, _ in COL_SPECS]
    kinds = [k for _, k in COL_SPECS]
    bracketed = ", ".join(f"[{n}]" for n in col_names)
    placeholders = ", ".join("?" * len(col_names))
    insert_sql = f"INSERT INTO dbo.ITT1_TMP ({bracketed}) VALUES ({placeholders})"

    conn = _connect(args.server, args.database, args.user, password)
    try:
        cur = conn.cursor()
        cur.execute(DROP_ITT1)
        cur.execute(CREATE_ITT1)
        batch: list[tuple[Any, ...]] = []
        for raw in data_rows:
            if len(raw) < len(col_names):
                raw = list(raw) + [None] * (len(col_names) - len(raw))
            elif len(raw) > len(col_names):
                raw = raw[: len(col_names)]
            batch.append(tuple(_coerce(raw[i], kinds[i]) for i in range(len(col_names))))
        cur.fast_executemany = True
        cur.executemany(insert_sql, batch)
        conn.commit()
        print(f"OK: dbo.ITT1_TMP recreated and loaded {len(batch)} row(s) from {args.xlsx}.")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
