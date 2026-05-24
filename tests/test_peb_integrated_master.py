"""PEB Integrated Master builder tests."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook

from core.erp.excel.peb_integrated_master import (
    PEB_IMASTER_HEADERS,
    build_peb_integrated_master_records,
)

_MASTER_XLSX = Path(
    r"c:\Users\lenovo\OneDrive\0_BAHTERA_WORK\PEB\7\master data.xlsx"
)


def _mini_workbook_bytes() -> bytes:
    wb = Workbook()
    ws_ml = wb.active
    ws_ml.title = "Master main line by model"
    ws_ml.append(["Model", "Mainline", "Cycle Time"])
    ws_ml.append([41418223, "MC", 30])

    ws_mc = wb.create_sheet("master code")
    ws_mc.append(
        [
            "FGcode",
            "Model",
            "process",
            "Infusion code",
            "size",
            "Cycle time infusion",
            "Ink Code",
            "PackType",
            "IMaster_TimeConstraintMin",
        ]
    )
    ws_mc.append(
        [
            "TESTFG001",
            41418223,
            "Main line",
            "INF001",
            "1L",
            14,
            "INK01",
            1,
            "4D",
        ]
    )

    ws_ink = wb.create_sheet("master infusion line by Ink")
    ws_ink.append(["Ink Code", "Ink Description", "Line"])
    ws_ink.append(["INK01", "Test ink", "MP", "MQ"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_build_main_line_i_and_u_rows() -> None:
    records = build_peb_integrated_master_records(_mini_workbook_bytes())
    main = [r for r in records if r["IMaster_ProcCode"] == "Main Line"]
    qa = [r for r in records if r["IMaster_ProcCode"] == "QA Output"]
    infusion = [r for r in records if r["IMaster_ProcCode"] == "Infusion"]
    assert len(main) == 2
    assert len(qa) == 1
    assert len(infusion) == 2
    i_row = next(r for r in main if r["IMaster_InstructionType"] == "I")
    u_row = next(r for r in main if r["IMaster_InstructionType"] == "U")
    assert i_row["IMaster_FinalItemCode"] == "TESTFG001"
    assert i_row["IMaster_ItemCodeOrResourceCode"] == "INF001"
    assert i_row["IMaster_Task2Expr"] == "1.00"
    assert u_row["IMaster_ItemCodeOrResourceCode"] == "MC"
    assert u_row["IMaster_Task2Expr"] == "30PH"
    assert qa[0]["IMaster_ItemCodeOrResourceCode"] == "MC QA Output"
    assert qa[0]["IMaster_TimeConstraintMin"] == "4D"
    assert u_row["IMaster_TimeConstraintMin"] == ""
    assert all(r["IMaster_InstructionType"] == "U" for r in infusion)
    assert infusion[0]["IMaster_FinalItemCode"] == "INF001"
    assert {r["IMaster_ItemCodeOrResourceCode"] for r in infusion} == {"MP", "MQ"}
    assert infusion[0]["IMaster_Task2Expr"] == "14PH"


def test_real_master_data_row_count() -> None:
    if not _MASTER_XLSX.is_file():
        return
    records = build_peb_integrated_master_records(_MASTER_XLSX.read_bytes())
    assert len(records) > 0
    infusion = [r for r in records if r["IMaster_ProcCode"] == "Infusion"]
    main_i = [
        r
        for r in records
        if r["IMaster_ProcCode"] == "Main Line" and r["IMaster_InstructionType"] == "I"
    ]
    assert len(main_i) > 0
    assert len(infusion) > 0
    assert all(r["IMaster_FinalItemCode"] != "" for r in infusion)
    assert all(r["IMaster_InstructionType"] == "U" for r in infusion)
    for h in PEB_IMASTER_HEADERS:
        assert h in records[0]
    qa_with_4d = [
        r
        for r in records
        if r["IMaster_ProcCode"] == "QA Output"
        and r["IMaster_TimeConstraintMin"] == "4D"
    ]
    assert len(qa_with_4d) >= 1
