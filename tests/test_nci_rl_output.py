"""NCI R/L Output (IntegratedMaster_OutputInstruction) tests."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook

from core.erp.excel.nci_rl_output import (
    NCI_RL_OUTPUT_HEADERS,
    build_nci_part_code_lookup,
    build_nci_rl_output_records,
)

_REFERENCE_DELIVERY = Path(
    r"c:\Users\lenovo\OneDrive\0_BAHTERA_WORK\NCI\Prototype\data\Delivery date_①_May.xlsx"
)
_REFERENCE_ITEM = Path(
    r"c:\Users\lenovo\OneDrive\0_BAHTERA_WORK\NCI\Prototype\data\Item.xlsx"
)
_EXPECTED_PAIR = (
    "NCI_TG4_71112_K500_0_NH696L_A",
    "NCI_TG4_71117_K500_0_NH696L_A",
)


def _mini_item_lookup_bytes(mapping: dict[str, str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    header = [None] * 120
    header[0] = "Item CD"
    header[119] = "Notes"
    ws.append(header)
    for part_code, item_cd in mapping.items():
        row = [None] * 120
        row[0] = item_cd
        row[119] = part_code
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mini_delivery_workbook_bytes(*, include_mcframe: bool = True) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BRIO PLANT 2"
    if include_mcframe:
        ws.append([None, "PART CODE", "PART NAME", "mcframe Item code"])
        ws.append(
            [
                None,
                "71112TG4 K500",
                "COVER R, FR FOG",
                "NCI_TG4_71112_K500_0_NH696L_A",
            ]
        )
        ws.append([None, None, None, None, None, "3UB RH IN"])
        ws.append(
            [
                None,
                "71117TG4 K500",
                "COVER L, FR FOG",
                "NCI_TG4_71117_K500_0_NH696L_A",
            ]
        )
    else:
        ws.append([None, "PART CODE", "PART NAME", "COLOR"])
        ws.append([None, "71112TG4 K500", "COVER R, FR FOG", None])
        ws.append([None, None, None, None, None, "3UB RH IN"])
        ws.append([None, "71117TG4 K500", "COVER L, FR FOG", None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mini_item_bytes() -> bytes:
    return _mini_item_lookup_bytes(
        {
            "71112TG4 K500": _EXPECTED_PAIR[0],
            "71117TG4 K500": _EXPECTED_PAIR[1],
        }
    )


def test_build_rl_pair_cross_output_rows() -> None:
    records = build_nci_rl_output_records(
        _mini_delivery_workbook_bytes(),
        _mini_item_bytes(),
    )
    assert len(records) == 2
    assert tuple(records[0][h] for h in NCI_RL_OUTPUT_HEADERS) == (
        _EXPECTED_PAIR[0],
        "10",
        "10",
        "O",
        "Out",
        _EXPECTED_PAIR[1],
        "",
        "",
    )
    assert tuple(records[1][h] for h in NCI_RL_OUTPUT_HEADERS) == (
        _EXPECTED_PAIR[1],
        "10",
        "10",
        "O",
        "Out",
        _EXPECTED_PAIR[0],
        "",
        "",
    )


def test_new_format_resolves_item_code_from_notes_lookup() -> None:
    records = build_nci_rl_output_records(
        _mini_delivery_workbook_bytes(include_mcframe=False),
        _mini_item_bytes(),
    )
    assert len(records) == 2
    finals = {r["IMaster_FinalItemCode"]: r["IMaster_ItemCodeOrResourceCode"] for r in records}
    assert finals[_EXPECTED_PAIR[0]] == _EXPECTED_PAIR[1]


def test_build_part_code_lookup_from_item_xlsx_notes_column() -> None:
    if not _REFERENCE_ITEM.is_file():
        return
    lookup = build_nci_part_code_lookup(_REFERENCE_ITEM.read_bytes())
    assert lookup["71103TG4 K500"] == "NCI_TG4_71103_K500_0_NH696L_A"


def test_same_item_code_rl_part_name_is_not_paired() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WR-V + WR-V EXP"
    ws.append([None, "PART CODE", "PART NAME", "mcframe Item code"])
    code = "NCI_3K6_75315_K110_0_NH792L_A"
    ws.append([None, "753153K6 K110M1", "GARN,R RR DOOR LWR", code])
    ws.append([None, "753353K6 K110M1", "GARN,L RR DOOR LWR", "NCI_3K6_75335_K110_0_NH792L_A"])
    ws.append([None, "X", "ADHESIVE PLASTER B R", code])
    ws.append([None, "X", "ADHESIVE PLASTER B L", code])
    buf = io.BytesIO()
    wb.save(buf)
    item_buf = _mini_item_lookup_bytes(
        {
            "753153K6 K110M1": code,
            "753353K6 K110M1": "NCI_3K6_75335_K110_0_NH792L_A",
            "X": code,
        }
    )
    records = build_nci_rl_output_records(buf.getvalue(), item_buf)
    assert not any(
        r["IMaster_FinalItemCode"] == code
        and r["IMaster_ItemCodeOrResourceCode"] == code
        for r in records
    )


def test_garn_variant_chain_does_not_cross_k001_and_k800() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WR-V + WR-V EXP"
    ws.append([None, "PART CODE", "PART NAME", "mcframe Item code"])
    rows = (
        ("71141TG4 K001", "GARN R,FR FOG", "NCI_3K6_71141_K001_0_NH696L_A"),
        ("71191TG4 K001", "GARN L,FR FOG", "NCI_3K6_71191_K001_0_NH696L_A"),
        ("71141TG4 K800", "GARN R,FR FOG", "NCI_3K6_71141_K800_0_NH792L_A"),
        ("71191TG4 K800", "GARN L,FR FOG", "NCI_3K6_71191_K800_0_NH792L_A"),
    )
    lookup = {}
    for part, _name, code in rows:
        lookup[part] = code
        ws.append([None, part, _name, code])
    buf = io.BytesIO()
    wb.save(buf)
    records = build_nci_rl_output_records(buf.getvalue(), _mini_item_lookup_bytes(lookup))
    finals = [r["IMaster_FinalItemCode"] for r in records]
    assert finals.count("NCI_3K6_71141_K800_0_NH792L_A") == 1
    k800 = next(
        r
        for r in records
        if r["IMaster_FinalItemCode"] == "NCI_3K6_71141_K800_0_NH792L_A"
    )
    assert k800["IMaster_ItemCodeOrResourceCode"] == "NCI_3K6_71191_K800_0_NH792L_A"


def test_reference_workbook_includes_cover_rl_pair() -> None:
    if not _REFERENCE_DELIVERY.is_file() or not _REFERENCE_ITEM.is_file():
        return
    item_raw = _REFERENCE_ITEM.read_bytes()
    delivery_raw = _REFERENCE_DELIVERY.read_bytes()
    records = build_nci_rl_output_records(delivery_raw, item_raw)
    finals = {r["IMaster_FinalItemCode"]: r["IMaster_ItemCodeOrResourceCode"] for r in records}
    assert finals.get(_EXPECTED_PAIR[0]) == _EXPECTED_PAIR[1]
    assert finals.get(_EXPECTED_PAIR[1]) == _EXPECTED_PAIR[0]

    legacy = build_nci_rl_output_records(delivery_raw, None)
    legacy_keys = {
        (r["IMaster_FinalItemCode"], r["IMaster_ItemCodeOrResourceCode"])
        for r in legacy
    }
    new_keys = {
        (r["IMaster_FinalItemCode"], r["IMaster_ItemCodeOrResourceCode"]) for r in records
    }
    assert legacy_keys == new_keys
