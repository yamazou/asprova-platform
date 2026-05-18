"""PEB inventory / WIP Excel import aggregation tests."""

from __future__ import annotations

import io

from openpyxl import Workbook

from core.erp.excel.peb_exports import (
    load_peb_inventory_rows_from_xlsx_bytes,
    load_peb_inventory_wip_rows_from_xlsx_bytes,
)


def _wip_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock_list_20260516"
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_wip_three_locations_sum_to_one_row() -> None:
    raw = _wip_bytes(
        [
            ["Material", "UoM", "Available", "Locatio"],
            ["10201060H", "G", 20000, "M601"],
            ["10201060H", "G", 20000, "M636"],
            ["10201060H", "G", 11355.44, "M6A1"],
        ]
    )
    rows = load_peb_inventory_wip_rows_from_xlsx_bytes(raw)
    assert len(rows) == 1
    assert rows[0][1] == "10201060H"
    assert rows[0][2] == "51355.44"


def test_wip_title_row_and_running_total_column_ignored() -> None:
    raw = _wip_bytes(
        [
            ["WIP Stock Report", None, None, None, None],
            ["Material", "UoM", "Available", "Locatio", None],
            ["10201060H", "G", 20000, "M601", 20000],
            [None, "G", 20000, "M636", 40000],
            [None, "G", 11355.44, "M6A1", 51355.44],
        ]
    )
    rows = load_peb_inventory_wip_rows_from_xlsx_bytes(raw)
    assert len(rows) == 1
    assert rows[0][2] == "51355.44"


def test_fg_stock_merged_material_forward_fill() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FG stock"
    ws.append(["Item Code", "Item Desc", "Tran", "5/16/2026"])
    ws.append(["10201060H", "desc", "On Hand", 20000])
    ws.append([None, "desc", "On Hand", 20000])
    ws.append([None, "desc", "On Hand", 11355.44])
    ws.merge_cells("A2:A4")
    buf = io.BytesIO()
    wb.save(buf)
    rows = load_peb_inventory_rows_from_xlsx_bytes(buf.getvalue())
    assert len(rows) == 1
    assert rows[0][1] == "10201060H"
    assert rows[0][2] == "51355.44"
